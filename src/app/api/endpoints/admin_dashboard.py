import asyncio
import base64
import csv
import io
import logging
import os
from datetime import datetime, timedelta
from typing import Dict, List, Optional
from zoneinfo import ZoneInfo

import pandas as pd
from fastapi import APIRouter, Body, Depends, HTTPException, Query, Response
from fastapi.responses import StreamingResponse
from sqlalchemy import case, func, or_, text
from sqlalchemy.orm import Session
from sqlalchemy.types import Float

from src.app import models
from src.app.api.deps import (
    require_admin_only,
    require_admin_or_editor,
    require_admin_token,
    require_viewer_or_editor,
)
from src.app.core.database import get_db

router = APIRouter(prefix="/admin/dashboard", tags=["Admin"])

logger = logging.getLogger("uvicorn.error")


# ---------------------------------------------------------------------------
# IMPORTANT: Register static `/.../search` routes before `/.../{job_id}` routes.
#
# Some deployments may route-match `/ingested-jobs/system/search` against the more
# generic `/ingested-jobs/system/{job_id}` and then fail validation (422) when
# trying to parse `job_id="search"` as an int.
# ---------------------------------------------------------------------------

@router.get(
    "/ingested-jobs/system/search",
    dependencies=[Depends(require_admin_token)],
    summary="Simple Text Search in System-ingested Jobs",
)
def search_system_ingested_jobs(
    q: str = Query(..., min_length=2, description="Search query (min 2 characters)"),
    page: int = Query(1, ge=1, description="Page number"),
    per_page: int = Query(25, ge=1, le=100, description="Items per page"),
    db: Session = Depends(get_db),
):
    """Admin endpoint: Simple text search in system-ingested jobs table.

    Searches across:
    - Job address
    - Permit number
    - Project description
    - Contractor name
    - Contractor company
    - Project number
    - Permit type
    """
    # Build search query
    search_term = f"%{q.lower()}%"

    base_query = db.query(models.user.Job).filter(
        models.user.Job.uploaded_by_contractor.is_(False),
        models.user.Job.uploaded_by_user_id.is_(None),
        or_(
            func.lower(models.user.Job.job_address).like(search_term),
            func.lower(models.user.Job.permit_number).like(search_term),
            func.lower(models.user.Job.project_description).like(search_term),
            func.lower(models.user.Job.contractor_name).like(search_term),
            func.lower(models.user.Job.contractor_company).like(search_term),
            func.lower(models.user.Job.project_number).like(search_term),
            func.lower(models.user.Job.audience_type_names).like(search_term),
        ),
    )

    # Get total count
    total_count = base_query.count()
    total_pages = (total_count + per_page - 1) // per_page

    # Order by status priority (pending first) then by created_at descending
    status_order = case(
        (models.user.Job.job_review_status == "pending", 1),
        (models.user.Job.job_review_status == "posted", 2),
        else_=3,
    )

    # Apply pagination
    offset = (page - 1) * per_page
    rows = (
        base_query.order_by(status_order, models.user.Job.created_at.desc())
        .limit(per_page)
        .offset(offset)
        .all()
    )

    jobs_data = []
    for j in rows:
        jobs_data.append(
            {
                "id": j.id,
                "permit_type": j.audience_type_names,
                "contact_name": j.contact_name,
                "contractor_email": j.contractor_email,
                "trs_score": j.trs_score,
                "job_review_status": j.job_review_status,
            }
        )

    # Return search results
    return {
        "jobs": jobs_data,
        "pagination": {
            "page": page,
            "per_page": per_page,
            "total_items": total_count,
            "total_pages": total_pages,
        },
        "search_query": q,
    }


@router.get(
    "/ingested-jobs/posted/search",
    dependencies=[Depends(require_admin_token)],
    summary="Simple Text Search in Posted Jobs",
)
def search_posted_jobs(
    q: str = Query(..., min_length=2, description="Search query (min 2 characters)"),
    page: int = Query(1, ge=1, description="Page number"),
    per_page: int = Query(25, ge=1, le=100, description="Items per page"),
    db: Session = Depends(get_db),
):
    """Admin endpoint: Simple text search in posted jobs table (job_review_status = 'posted').

    Searches across:
    - Job address
    - Permit number
    - Project description
    - Contractor name
    - Contractor company
    - Project number
    - Audience type (permit type)
    """
    # Build search query
    search_term = f"%{q.lower()}%"

    base_query = db.query(models.user.Job).filter(
        models.user.Job.uploaded_by_contractor == False,
        models.user.Job.job_review_status == "posted",
        or_(
            func.lower(models.user.Job.job_address).like(search_term),
            func.lower(models.user.Job.permit_number).like(search_term),
            func.lower(models.user.Job.project_description).like(search_term),
            func.lower(models.user.Job.contractor_name).like(search_term),
            func.lower(models.user.Job.contractor_company).like(search_term),
            func.lower(models.user.Job.project_number).like(search_term),
            func.lower(models.user.Job.audience_type_names).like(search_term),
        ),
    )

    # Get total count
    total_count = base_query.count()
    total_pages = (total_count + per_page - 1) // per_page

    # Order by review_posted_at descending (most recent first)
    offset = (page - 1) * per_page
    rows = (
        base_query.order_by(models.user.Job.review_posted_at.desc())
        .limit(per_page)
        .offset(offset)
        .all()
    )

    jobs_data = []
    for j in rows:
        jobs_data.append(
            {
                "id": j.id,
                "permit_type": j.audience_type_names,
                "contact_name": j.contact_name,
                "contractor_email": j.contractor_email,
                "trs_score": j.trs_score,
                "job_review_status": j.job_review_status,
            }
        )

    # Return search results
    return {
        "jobs": jobs_data,
        "pagination": {
            "page": page,
            "per_page": per_page,
            "total_items": total_count,
            "total_pages": total_pages,
        },
        "search_query": q,
    }


@router.get(
    "/jobs/{job_id}",
    dependencies=[Depends(require_admin_token)],
    summary="Get Complete Job Details (Admin)",
)
def get_job_details(job_id: int, db: Session = Depends(get_db)):
    """Admin endpoint: get complete job details by job id.

    Returns all job data points including documents, contractor info, project details, etc.
    """
    job = db.query(models.user.Job).filter(models.user.Job.id == job_id).first()
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")
    
    # Get uploaded by user details if available
    uploaded_by_user = None
    if job.uploaded_by_user_id:
        user = db.query(models.user.User).filter(models.user.User.id == job.uploaded_by_user_id).first()
        if user:
            uploaded_by_user = {
                "id": user.id,
                "email": user.email,
                "name": user.name,
                "company_name": getattr(user, "company_name", None),
            }
    
    return {
        # Basic job info
        "id": job.id,
        "created_at": job.created_at.isoformat() if job.created_at else None,
        "updated_at": job.updated_at.isoformat() if job.updated_at else None,
        
        # Queue and routing info
        "queue_id": job.queue_id,
        "rule_id": job.rule_id,
        "recipient_group": job.recipient_group,
        "recipient_group_id": job.recipient_group_id,
        "day_offset": job.day_offset,
        "anchor_event": job.anchor_event,
        "anchor_at": job.anchor_at.isoformat() if job.anchor_at else None,
        "due_at": job.due_at.isoformat() if job.due_at else None,
        "routing_anchor_at": job.routing_anchor_at.isoformat() if job.routing_anchor_at else None,
        
        # Permit info
        "permit_id": job.permit_id,
        "permit_number": job.permit_number,
        "permit_status": job.permit_status,
        "permit_type": job.permit_type,
        "permit_type_norm": job.permit_type_norm,
        "permit_raw": job.permit_raw,
        
        # Project info
        "project_number": job.project_number,
        "project_description": job.project_description,
        "project_type": job.project_type,
        "project_sub_type": job.project_sub_type,
        "project_status": job.project_status,
        "project_cost_total": job.project_cost_total,
        "project_cost": job.project_cost,
        "project_cost_source": job.project_cost_source,
        "property_type": job.property_type,
        
        # Address info
        "job_address": job.job_address,
        "project_address": job.project_address,
        "state": job.state,
        
        # Source info
        "source_county": job.source_county,
        "source_system": job.source_system,
        "first_seen_at": job.first_seen_at.isoformat() if job.first_seen_at else None,
        "last_seen_at": job.last_seen_at.isoformat() if job.last_seen_at else None,
        
        # Contractor info
        "contractor_name": job.contractor_name,
        "contractor_company": job.contractor_company,
        "contractor_email": job.contractor_email,
        "contractor_phone": job.contractor_phone,
        "contractor_company_and_address": job.contractor_company_and_address,
        "contact_name": job.contact_name,
        
        # Owner/Applicant info
        "owner_name": job.owner_name,
        "applicant_name": job.applicant_name,
        "applicant_email": job.applicant_email,
        "applicant_phone": job.applicant_phone,
        
        # Audience info
        "audience_type_slugs": job.audience_type_slugs,
        "audience_type_names": job.audience_type_names,
        
        # Additional info
        "querystring": job.querystring,
        "trs_score": job.trs_score,
        "uploaded_by_contractor": job.uploaded_by_contractor,
        "uploaded_by_user_id": job.uploaded_by_user_id,
        "uploaded_by_user": uploaded_by_user,
        "job_review_status": job.job_review_status,
        "review_posted_at": job.review_posted_at.isoformat() if job.review_posted_at else None,
        "job_group_id": job.job_group_id,
        "job_documents": job.job_documents,  # Array of document URLs/paths
        "decline_note": getattr(job, "decline_note", None),  # Admin's decline reason
    }


@router.get(
    "/contractor-uploaded-jobs/{job_id:int}",
    dependencies=[Depends(require_admin_token)],
    summary="Get Contractor-Uploaded Job Details (Admin)",
)
def get_contractor_uploaded_job_details(job_id: int, db: Session = Depends(get_db)):
    """Admin endpoint: get complete details for a contractor-uploaded job.
    
    Returns all job data points including documents for jobs uploaded by contractors.
    """
    job = db.query(models.user.Job).filter(
        models.user.Job.id == job_id,
        models.user.Job.uploaded_by_contractor == True
    ).first()
    
    if not job:
        raise HTTPException(status_code=404, detail="Contractor-uploaded job not found")
    
    # Get uploaded by user details if available
    uploaded_by_user = None
    if job.uploaded_by_user_id:
        user = db.query(models.user.User).filter(models.user.User.id == job.uploaded_by_user_id).first()
        if user:
            uploaded_by_user = {
                "id": user.id,
                "email": user.email,
                "name": getattr(user, "name", user.email),
                "company_name": getattr(user, "company_name", None),
            }
    
    return {
        # Basic job info
        "id": job.id,
        "created_at": job.created_at.isoformat() if job.created_at else None,
        "updated_at": job.updated_at.isoformat() if job.updated_at else None,
        
        # Queue and routing info
        "queue_id": job.queue_id,
        "rule_id": job.rule_id,
        "recipient_group": job.recipient_group,
        "recipient_group_id": job.recipient_group_id,
        "day_offset": job.day_offset,
        "anchor_event": job.anchor_event,
        "anchor_at": job.anchor_at.isoformat() if job.anchor_at else None,
        "due_at": job.due_at.isoformat() if job.due_at else None,
        "routing_anchor_at": job.routing_anchor_at.isoformat() if job.routing_anchor_at else None,
        
        # Permit info
        "permit_id": job.permit_id,
        "permit_number": job.permit_number,
        "permit_status": job.permit_status,
        "permit_type": job.permit_type,
        "permit_type_norm": job.permit_type_norm,
        "permit_raw": job.permit_raw,
        
        # Project info
        "project_number": job.project_number,
        "project_description": job.project_description,
        "project_type": job.project_type,
        "project_sub_type": job.project_sub_type,
        "project_status": job.project_status,
        "project_cost_total": job.project_cost_total,
        "project_cost": job.project_cost,
        "project_cost_source": job.project_cost_source,
        "property_type": job.property_type,
        
        # Address info
        "job_address": job.job_address,
        "project_address": job.project_address,
        "state": job.state,
        
        # Source info
        "source_county": job.source_county,
        "source_system": job.source_system,
        "first_seen_at": job.first_seen_at.isoformat() if job.first_seen_at else None,
        "last_seen_at": job.last_seen_at.isoformat() if job.last_seen_at else None,
        
        # Contractor info
        "contractor_name": job.contractor_name,
        "contractor_company": job.contractor_company,
        "contractor_email": job.contractor_email,
        "contractor_phone": job.contractor_phone,
        "contractor_company_and_address": job.contractor_company_and_address,
        "contact_name": job.contact_name,
        
        # Owner/Applicant info
        "owner_name": job.owner_name,
        "applicant_name": job.applicant_name,
        "applicant_email": job.applicant_email,
        "applicant_phone": job.applicant_phone,
        
        # Audience info
        "audience_type_slugs": job.audience_type_slugs,
        "audience_type_names": job.audience_type_names,
        
        # Additional info
        "querystring": job.querystring,
        "trs_score": job.trs_score,
        "uploaded_by_contractor": job.uploaded_by_contractor,
        "uploaded_by_user_id": job.uploaded_by_user_id,
        "uploaded_by_user": uploaded_by_user,
        "job_review_status": job.job_review_status,
        "review_posted_at": job.review_posted_at.isoformat() if job.review_posted_at else None,
        "job_group_id": job.job_group_id,
        "job_documents": job.job_documents,  # Include documents for contractor jobs
        "decline_note": getattr(job, "decline_note", None),  # Admin's decline reason
    }


@router.get(
    "/ingested-jobs/system/{job_id:int}",
    dependencies=[Depends(require_admin_token)],
    summary="Get System-Ingested Job Details (Admin)",
)
def get_system_ingested_job_details(job_id: int, db: Session = Depends(get_db)):
    """Admin endpoint: get complete details for a system-ingested job.
    
    Returns all job data points (excluding documents) for jobs uploaded via bulk system ingestion.
    """
    job = db.query(models.user.Job).filter(
        models.user.Job.id == job_id,
        models.user.Job.uploaded_by_contractor == False,
        models.user.Job.uploaded_by_user_id.is_(None)
    ).first()
    
    if not job:
        raise HTTPException(status_code=404, detail="System-ingested job not found")
    
    return {
        # Basic job info
        "id": job.id,
        "created_at": job.created_at.isoformat() if job.created_at else None,
        "updated_at": job.updated_at.isoformat() if job.updated_at else None,
        
        # Queue and routing info
        "queue_id": job.queue_id,
        "rule_id": job.rule_id,
        "recipient_group": job.recipient_group,
        "recipient_group_id": job.recipient_group_id,
        "day_offset": job.day_offset,
        "anchor_event": job.anchor_event,
        "anchor_at": job.anchor_at.isoformat() if job.anchor_at else None,
        "due_at": job.due_at.isoformat() if job.due_at else None,
        "routing_anchor_at": job.routing_anchor_at.isoformat() if job.routing_anchor_at else None,
        
        # Permit info
        "permit_id": job.permit_id,
        "permit_number": job.permit_number,
        "permit_status": job.permit_status,
        "permit_type": job.permit_type,
        "permit_type_norm": job.permit_type_norm,
        "permit_raw": job.permit_raw,
        
        # Project info
        "project_number": job.project_number,
        "project_description": job.project_description,
        "project_type": job.project_type,
        "project_sub_type": job.project_sub_type,
        "project_status": job.project_status,
        "project_cost_total": job.project_cost_total,
        "project_cost": job.project_cost,
        "project_cost_source": job.project_cost_source,
        "property_type": job.property_type,
        
        # Address info
        "job_address": job.job_address,
        "project_address": job.project_address,
        "state": job.state,
        
        # Source info
        "source_county": job.source_county,
        "source_system": job.source_system,
        "first_seen_at": job.first_seen_at.isoformat() if job.first_seen_at else None,
        "last_seen_at": job.last_seen_at.isoformat() if job.last_seen_at else None,
        
        # Contractor info
        "contractor_name": job.contractor_name,
        "contractor_company": job.contractor_company,
        "contractor_email": job.contractor_email,
        "contractor_phone": job.contractor_phone,
        "contractor_company_and_address": job.contractor_company_and_address,
        "contact_name": job.contact_name,
        
        # Owner/Applicant info
        "owner_name": job.owner_name,
        "applicant_name": job.applicant_name,
        "applicant_email": job.applicant_email,
        "applicant_phone": job.applicant_phone,
        
        # Audience info
        "audience_type_slugs": job.audience_type_slugs,
        "audience_type_names": job.audience_type_names,
        
        # Additional info
        "querystring": job.querystring,
        "trs_score": job.trs_score,
        "uploaded_by_contractor": job.uploaded_by_contractor,
        "uploaded_by_user_id": job.uploaded_by_user_id,
        "job_review_status": job.job_review_status,
        "review_posted_at": job.review_posted_at.isoformat() if job.review_posted_at else None,
        "job_group_id": job.job_group_id,
        # Note: job_documents excluded for system-ingested jobs
    }


@router.get(
    "/ingested-jobs/posted/{job_id:int}",
    dependencies=[Depends(require_admin_token)],
    summary="Get Posted Job Details (Admin)",
)
def get_posted_job_details(job_id: int, db: Session = Depends(get_db)):
    """Admin endpoint: get complete details for a posted job.
    
    Returns all job data points including documents for jobs with status 'posted'.
    """
    job = db.query(models.user.Job).filter(
        models.user.Job.id == job_id,
        models.user.Job.job_review_status == "posted"
    ).first()
    
    if not job:
        raise HTTPException(status_code=404, detail="Posted job not found")
    
    # Get uploaded by user details if available
    uploaded_by_user = None
    if job.uploaded_by_user_id:
        user = db.query(models.user.User).filter(models.user.User.id == job.uploaded_by_user_id).first()
        if user:
            uploaded_by_user = {
                "id": user.id,
                "email": user.email,
                "name": user.name,
                "company_name": getattr(user, "company_name", None),
            }
    
    return {
        # Basic job info
        "id": job.id,
        "created_at": job.created_at.isoformat() if job.created_at else None,
        "updated_at": job.updated_at.isoformat() if job.updated_at else None,
        
        # Queue and routing info
        "queue_id": job.queue_id,
        "rule_id": job.rule_id,
        "recipient_group": job.recipient_group,
        "recipient_group_id": job.recipient_group_id,
        "day_offset": job.day_offset,
        "anchor_event": job.anchor_event,
        "anchor_at": job.anchor_at.isoformat() if job.anchor_at else None,
        "due_at": job.due_at.isoformat() if job.due_at else None,
        "routing_anchor_at": job.routing_anchor_at.isoformat() if job.routing_anchor_at else None,
        
        # Permit info
        "permit_id": job.permit_id,
        "permit_number": job.permit_number,
        "permit_status": job.permit_status,
        "permit_type": job.permit_type,
        "permit_type_norm": job.permit_type_norm,
        "permit_raw": job.permit_raw,
        
        # Project info
        "project_number": job.project_number,
        "project_description": job.project_description,
        "project_type": job.project_type,
        "project_sub_type": job.project_sub_type,
        "project_status": job.project_status,
        "project_cost_total": job.project_cost_total,
        "project_cost": job.project_cost,
        "project_cost_source": job.project_cost_source,
        "property_type": job.property_type,
        
        # Address info
        "job_address": job.job_address,
        "project_address": job.project_address,
        "state": job.state,
        
        # Source info
        "source_county": job.source_county,
        "source_system": job.source_system,
        "first_seen_at": job.first_seen_at.isoformat() if job.first_seen_at else None,
        "last_seen_at": job.last_seen_at.isoformat() if job.last_seen_at else None,
        
        # Contractor info
        "contractor_name": job.contractor_name,
        "contractor_company": job.contractor_company,
        "contractor_email": job.contractor_email,
        "contractor_phone": job.contractor_phone,
        "contractor_company_and_address": job.contractor_company_and_address,
        "contact_name": job.contact_name,
        
        # Owner/Applicant info
        "owner_name": job.owner_name,
        "applicant_name": job.applicant_name,
        "applicant_email": job.applicant_email,
        "applicant_phone": job.applicant_phone,
        
        # Audience info
        "audience_type_slugs": job.audience_type_slugs,
        "audience_type_names": job.audience_type_names,
        
        # Additional info
        "querystring": job.querystring,
        "trs_score": job.trs_score,
        "uploaded_by_contractor": job.uploaded_by_contractor,
        "uploaded_by_user_id": job.uploaded_by_user_id,
        "uploaded_by_user": uploaded_by_user,
        "job_review_status": job.job_review_status,
        "review_posted_at": job.review_posted_at.isoformat() if job.review_posted_at else None,
        "job_group_id": job.job_group_id,
        "job_documents": job.job_documents,  # Include documents for posted jobs
        "decline_note": getattr(job, "decline_note", None),  # Admin's decline reason
    }


import uuid
from datetime import datetime, timedelta

from pydantic import BaseModel

from src.app.utils.email import send_admin_invitation_email


class UserApprovalUpdate(BaseModel):
    user_id: int
    status: str  # "approved" or "rejected"


class AdminInvite(BaseModel):
    email: str
    name: Optional[str] = None
    role: str


class ContractorApprovalUpdate(BaseModel):
    status: str  # "approved" or "rejected"
    note: Optional[str] = None  # Optional admin note


class JobUpdate(BaseModel):
    """Model for updating job fields. All fields are optional."""
    # Queue and routing info
    queue_id: Optional[int] = None
    rule_id: Optional[int] = None
    recipient_group: Optional[str] = None
    recipient_group_id: Optional[int] = None
    day_offset: Optional[int] = None
    anchor_event: Optional[str] = None
    anchor_at: Optional[datetime] = None
    due_at: Optional[datetime] = None
    routing_anchor_at: Optional[datetime] = None
    
    # Permit info
    permit_id: Optional[int] = None
    permit_number: Optional[str] = None
    permit_status: Optional[str] = None
    permit_type_norm: Optional[str] = None
    permit_raw: Optional[str] = None
    
    # Project info
    project_number: Optional[str] = None
    project_description: Optional[str] = None
    project_type: Optional[str] = None
    project_sub_type: Optional[str] = None
    project_status: Optional[str] = None
    project_cost_total: Optional[int] = None
    project_cost: Optional[int] = None
    project_cost_source: Optional[str] = None
    project_address: Optional[str] = None
    property_type: Optional[str] = None
    
    # Address info
    job_address: Optional[str] = None
    state: Optional[str] = None
    
    # Source info
    source_county: Optional[str] = None
    source_system: Optional[str] = None
    first_seen_at: Optional[datetime] = None
    last_seen_at: Optional[datetime] = None
    
    # Contractor info
    contractor_name: Optional[str] = None
    contractor_company: Optional[str] = None
    contractor_email: Optional[str] = None
    contractor_phone: Optional[str] = None
    contractor_company_and_address: Optional[str] = None
    contact_name: Optional[str] = None
    
    # Owner/Applicant info
    owner_name: Optional[str] = None
    applicant_name: Optional[str] = None
    applicant_email: Optional[str] = None
    applicant_phone: Optional[str] = None
    
    # Audience info
    audience_type_slugs: Optional[str] = None
    audience_type_names: Optional[str] = None
    
    # Additional info
    querystring: Optional[str] = None
    trs_score: Optional[int] = None
    uploaded_by_contractor: Optional[bool] = None
    uploaded_by_user_id: Optional[int] = None
    job_review_status: Optional[str] = None
    review_posted_at: Optional[datetime] = None
    job_group_id: Optional[str] = None


def _periods_for_range(time_range: str):
    """Return (periods, bucket) where periods is list of (label,start,end).

    bucket: 'month' or 'day'
    """
    now = datetime.utcnow()
    # Normalize the time_range to lowercase and remove spaces
    time_range_normalized = time_range.lower().replace(" ", "")

    if time_range_normalized in ["last6months", "last6month"]:
        # reuse _month_starts
        return _month_starts(6), "month"
    if time_range_normalized in ["last3months", "last3month"]:
        return _month_starts(3), "month"
    if time_range_normalized in ["last12months", "last12month"]:
        return _month_starts(12), "month"
    if time_range_normalized in ["thisyear"]:
        start = datetime(now.year, 1, 1)
        months = []
        for m in range(1, now.month + 1):
            s = datetime(now.year, m, 1)
            nm = m + 1
            ny = now.year
            if nm == 13:
                nm = 1
                ny += 1
            e = datetime(ny, nm, 1)
            months.append((s.strftime("%b"), s, e))
        return months, "month"
    if time_range_normalized in ["lastyear"]:
        y = now.year - 1
        months = []
        for m in range(1, 13):
            s = datetime(y, m, 1)
            nm = m + 1
            ny = y
            if nm == 13:
                nm = 1
                ny += 1
            e = datetime(ny, nm, 1)
            months.append((s.strftime("%b"), s, e))
        return months, "month"
    if time_range_normalized in ["last30days", "last30day"]:
        periods = []
        for i in range(29, -1, -1):
            d = (now - timedelta(days=i)).replace(
                hour=0, minute=0, second=0, microsecond=0
            )
            periods.append((d.strftime("%Y-%m-%d"), d, d + timedelta(days=1)))
        return periods, "day"
    if time_range_normalized in ["last90days", "last90day"]:
        periods = []
        for i in range(89, -1, -1):
            d = (now - timedelta(days=i)).replace(
                hour=0, minute=0, second=0, microsecond=0
            )
            periods.append((d.strftime("%Y-%m-%d"), d, d + timedelta(days=1)))
        return periods, "day"
    # default
    return _month_starts(6), "month"


def _month_starts(last_n: int = 6):
    """Return list of (month_label, start_dt, end_dt) for last_n months (ascending)."""
    now = datetime.utcnow().replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    months = []
    for i in range(last_n - 1, -1, -1):
        # month i months ago
        year = now.year
        month = now.month - i
        while month <= 0:
            month += 12
            year -= 1
        start = datetime(year, month, 1)
        # compute next month
        nm = month + 1
        ny = year
        if nm == 13:
            nm = 1
            ny += 1
        end = datetime(ny, nm, 1)
        label = start.strftime("%b")
        months.append((label, start, end))
    return months


def _table_exists(db: Session, table_name: str) -> bool:
    r = db.execute(
        text("SELECT 1 FROM information_schema.tables WHERE table_name = :t LIMIT 1"),
        {"t": table_name},
    ).first()
    return bool(r)


def _percent_change(current: int, previous: int):
    """Return (pct, formatted_str) where pct is a float or None when undefined.

    Rules:
    - if previous > 0: pct = (current - previous) / previous * 100
    - if previous == 0 and current == 0: pct = 0.0
    - if previous == 0 and current > 0: return a numeric value (show +100%)
    """
    try:
        if previous > 0:
            pct = round((current - previous) / previous * 100.0, 1)
            return pct, f"{pct:+}%"
        # previous == 0
        if current == 0:
            return 0.0, "+0%"
        # previous == 0 and current > 0 -> represent as +100% (numeric) rather than N/A
        pct = 100.0
        return pct, f"+{pct}%"
    except Exception:
        return 0.0, "+0%"


@router.get("/contractors-kpis", dependencies=[Depends(require_admin_token)])
def contractors_kpis(db: Session = Depends(get_db)):
    """
    Admin endpoint: Contractors KPIs only.

    Returns all KPI metrics with percentage changes.
    """
    from datetime import datetime, timedelta

    # Calculate KPIs
    # Active Subscriptions
    active_subs = (
        db.query(func.count(models.user.Subscriber.id))
        .join(models.user.User, models.user.Subscriber.user_id == models.user.User.id)
        .filter(
            models.user.Subscriber.is_active == True,
            models.user.User.role == "Contractor",
            models.user.User.approved_by_admin == "approved",
        )
        .scalar()
        or 0
    )

    # Past Due
    past_due = (
        db.query(func.count(models.user.Subscriber.id))
        .join(models.user.User, models.user.Subscriber.user_id == models.user.User.id)
        .filter(
            models.user.Subscriber.subscription_status == "past_due",
            models.user.User.role == "Contractor",
            models.user.User.approved_by_admin == "approved",
        )
        .scalar()
        or 0
    )

    # Credits Outstanding (current credits)
    credits_outstanding = (
        db.query(func.coalesce(func.sum(models.user.Subscriber.current_credits), 0))
        .join(models.user.User, models.user.Subscriber.user_id == models.user.User.id)
        .filter(
            models.user.User.role == "Contractor",
            models.user.User.approved_by_admin == "approved",
        )
        .scalar()
        or 0
    )

    # Unlocks Last 7 Days
    seven_days_ago = datetime.utcnow() - timedelta(days=7)
    unlocks_last_7d = (
        db.query(func.count(models.user.UnlockedLead.id))
        .join(models.user.User, models.user.UnlockedLead.user_id == models.user.User.id)
        .filter(
            models.user.UnlockedLead.unlocked_at >= seven_days_ago,
            models.user.User.role == "Contractor",
            models.user.User.approved_by_admin == "approved",
        )
        .scalar()
        or 0
    )

    # Credits Purchased - Total credits ever acquired (current + spent + frozen)
    credits_purchased_result = (
        db.query(
            func.coalesce(
                func.sum(
                    models.user.Subscriber.current_credits
                    + models.user.Subscriber.total_spending
                    + models.user.Subscriber.frozen_credits
                ),
                0,
            )
        )
        .join(models.user.User, models.user.Subscriber.user_id == models.user.User.id)
        .filter(
            models.user.User.role == "Contractor",
            models.user.User.approved_by_admin == "approved",
        )
        .scalar()
        or 0
    )

    credits_purchased = int(credits_purchased_result) if credits_purchased_result else 0

    # Credits Spent - Total credits spent on unlocking leads
    credits_spent = (
        db.query(func.coalesce(func.sum(models.user.Subscriber.total_spending), 0))
        .join(models.user.User, models.user.Subscriber.user_id == models.user.User.id)
        .filter(
            models.user.User.role == "Contractor",
            models.user.User.approved_by_admin == "approved",
        )
        .scalar()
        or 0
    )

    # Trial Credits Used - Total trial credits consumed by all contractors
    trial_credits_used_sum = (
        db.query(func.coalesce(func.sum(25 - models.user.Subscriber.trial_credits), 0))
        .join(models.user.User, models.user.Subscriber.user_id == models.user.User.id)
        .filter(
            models.user.Subscriber.trial_credits_used == True,
            models.user.User.role == "Contractor",
            models.user.User.approved_by_admin == "approved",
        )
        .scalar()
        or 0
    )

    trial_credits_used_count = (
        int(trial_credits_used_sum) if trial_credits_used_sum else 0
    )

    # Leads Ingested Today (jobs created today)
    today_start = datetime.utcnow().replace(hour=0, minute=0, second=0, microsecond=0)
    leads_ingested_today = (
        db.query(func.count(models.user.Job.id))
        .filter(models.user.Job.created_at >= today_start)
        .scalar()
        or 0
    )

    # Leads Delivered (total posted jobs)
    leads_delivered = (
        db.query(func.count(models.user.Job.id))
        .filter(models.user.Job.job_review_status == "posted")
        .scalar()
        or 0
    )

    # Leads Unlocked (total unlocks by contractors)
    leads_unlocked = (
        db.query(func.count(models.user.UnlockedLead.id))
        .join(models.user.User, models.user.UnlockedLead.user_id == models.user.User.id)
        .filter(
            models.user.User.role == "Contractor",
            models.user.User.approved_by_admin == "approved",
        )
        .scalar()
        or 0
    )

    # Calculate previous period values for percentage changes
    fourteen_days_ago = datetime.utcnow() - timedelta(days=14)

    # Previous Unlocks (7-14 days ago)
    unlocks_prev_7d = (
        db.query(func.count(models.user.UnlockedLead.id))
        .join(models.user.User, models.user.UnlockedLead.user_id == models.user.User.id)
        .filter(
            models.user.UnlockedLead.unlocked_at >= fourteen_days_ago,
            models.user.UnlockedLead.unlocked_at < seven_days_ago,
            models.user.User.role == "Contractor",
            models.user.User.approved_by_admin == "approved",
        )
        .scalar()
        or 0
    )

    # Previous Leads Ingested (yesterday)
    yesterday_start = today_start - timedelta(days=1)
    leads_ingested_yesterday = (
        db.query(func.count(models.user.Job.id))
        .filter(
            models.user.Job.created_at >= yesterday_start,
            models.user.Job.created_at < today_start,
        )
        .scalar()
        or 0
    )

    # Helper function to calculate percentage change
    def calc_percentage_change(current, previous):
        if previous == 0:
            return 0 if current == 0 else 100
        return round(((current - previous) / previous) * 100, 1)

    return {
        "activeSubscriptions": {"value": active_subs, "change": 0},
        "pastDue": {"value": past_due, "change": 0},
        "creditsOutstanding": {"value": credits_outstanding, "change": 0},
        "unlocksLast7d": {
            "value": unlocks_last_7d,
            "change": calc_percentage_change(unlocks_last_7d, unlocks_prev_7d),
        },
        "creditsPurchased": {"value": credits_purchased, "change": 0},
        "creditsSpent": {"value": credits_spent, "change": 0},
        "trialCreditsUsed": {"value": trial_credits_used_count, "change": 0},
        "leadsIngestedToday": {
            "value": leads_ingested_today,
            "change": calc_percentage_change(
                leads_ingested_today, leads_ingested_yesterday
            ),
        },
        "leadsDelivered": {"value": leads_delivered, "change": 0},
        "leadsUnlocked": {"value": leads_unlocked, "change": 0},
    }


@router.get("/contractors-summary", dependencies=[Depends(require_admin_token)])
def contractors_summary(
    # Account Status filters
    account_status: Optional[str] = Query(
        None, description="Account status: active, disabled"
    ),
    # Subscription Status filters
    subscription_status: Optional[str] = Query(
        None,
        description="Subscription status: active, past_due, canceled, action_required, paused, trial, trial_expired, inactive, trialing",
    ),
    # Plan Tier filters
    plan_tier: Optional[str] = Query(
        None,
        description="Plan tier: Starter, Professional, Enterprise, Custom, no_subscription",
    ),
    # Credits Balance range
    credits_min: Optional[int] = Query(None, description="Minimum credits balance"),
    credits_max: Optional[int] = Query(None, description="Maximum credits balance"),
    # Unlocks Last 7 Days
    unlocks_range: Optional[str] = Query(
        None,
        description="Unlocks range: no_unlocks, low_1_5, medium_6_10, moderate_11_20, heavy_20_plus",
    ),
    # Trade Categories (user_type)
    trade_category: Optional[str] = Query(
        None, description="Filter by trade category/user type"
    ),
    # Service Area States
    service_state: Optional[str] = Query(None, description="Filter by service state"),
    # Registration Date
    registration_date: Optional[str] = Query(
        None,
        description="Registration date: last_7_days, last_30_days, last_90_days, this_year",
    ),
    # Active Date (last login or activity)
    active_date: Optional[str] = Query(
        None, description="Active date: last_7_days, last_30_days, inactive_90_plus"
    ),
    # Pagination
    page: int = Query(1, ge=1, description="Page number"),
    per_page: int = Query(10, ge=1, le=100, description="Items per page"),
    # Search
    search: Optional[str] = Query(None, description="Search by company, email, phone"),
    db: Session = Depends(get_db),
):
    """
    Admin endpoint: Contractors table data with filters and pagination.

    Returns table data with applied filters.
    """
    from datetime import datetime, timedelta

    # Base query for contractors with users and subscribers
    base_query = (
        db.query(
            models.user.Contractor.id,
            models.user.Contractor.company_name,
            models.user.Contractor.phone_number,
            models.user.Contractor.user_type,
            models.user.Contractor.state,
            models.user.User.email,
            models.user.User.is_active,
            models.user.User.created_at,
            models.user.User.approved_by_admin,
            models.user.Subscriber.subscription_id,
            models.user.Subscriber.current_credits,
            models.user.Subscriber.subscription_status,
            models.user.Subscriber.subscription_renew_date,
            models.user.Subscriber.total_spending,
            models.user.Subscription.name.label("plan_name"),
        )
        .join(models.user.User, models.user.User.id == models.user.Contractor.user_id)
        .outerjoin(
            models.user.Subscriber,
            models.user.Subscriber.user_id == models.user.User.id,
        )
        .outerjoin(
            models.user.Subscription,
            models.user.Subscription.id == models.user.Subscriber.subscription_id,
        )
        .filter(
            models.user.User.approved_by_admin == "approved",
            models.user.User.role == "Contractor",
        )
    )

    # Apply filters
    # Account Status
    if account_status:
        if account_status.lower() == "active":
            base_query = base_query.filter(models.user.User.is_active == True)
        elif account_status.lower() == "disabled":
            base_query = base_query.filter(models.user.User.is_active == False)

    # Subscription Status
    if subscription_status:
        base_query = base_query.filter(
            models.user.Subscriber.subscription_status == subscription_status.lower()
        )

    # Plan Tier
    if plan_tier:
        if plan_tier.lower() == "no_subscription":
            base_query = base_query.filter(
                models.user.Subscriber.subscription_id.is_(None)
            )
        else:
            base_query = base_query.filter(
                func.lower(models.user.Subscription.name) == plan_tier.lower()
            )

    # Credits Balance range
    if credits_min is not None:
        base_query = base_query.filter(
            models.user.Subscriber.current_credits >= credits_min
        )
    if credits_max is not None:
        base_query = base_query.filter(
            models.user.Subscriber.current_credits <= credits_max
        )

    # Trade Category
    if trade_category:
        base_query = base_query.filter(
            models.user.Contractor.user_type.any(trade_category)
        )

    # Service Area States
    if service_state:
        base_query = base_query.filter(models.user.Contractor.state.any(service_state))

    # Registration Date
    if registration_date:
        now = datetime.utcnow()
        if registration_date == "last_7_days":
            base_query = base_query.filter(
                models.user.User.created_at >= now - timedelta(days=7)
            )
        elif registration_date == "last_30_days":
            base_query = base_query.filter(
                models.user.User.created_at >= now - timedelta(days=30)
            )
        elif registration_date == "last_90_days":
            base_query = base_query.filter(
                models.user.User.created_at >= now - timedelta(days=90)
            )
        elif registration_date == "this_year":
            base_query = base_query.filter(
                func.extract("year", models.user.User.created_at) == now.year
            )

    # Search
    if search:
        search_term = f"%{search.lower()}%"
        base_query = base_query.filter(
            or_(
                func.lower(models.user.Contractor.company_name).like(search_term),
                func.lower(models.user.User.email).like(search_term),
                func.lower(models.user.Contractor.phone_number).like(search_term),
            )
        )

    # Apply Unlocks Last 7 Days filter
    if unlocks_range:
        seven_days_ago = datetime.utcnow() - timedelta(days=7)
        # Subquery to count unlocks per user in last 7 days
        unlocks_subq = (
            db.query(
                models.user.UnlockedLead.user_id,
                func.count(models.user.UnlockedLead.id).label("unlock_count"),
            )
            .filter(models.user.UnlockedLead.unlocked_at >= seven_days_ago)
            .group_by(models.user.UnlockedLead.user_id)
            .subquery()
        )

        base_query = base_query.outerjoin(
            unlocks_subq, unlocks_subq.c.user_id == models.user.User.id
        )

        if unlocks_range == "no_unlocks":
            base_query = base_query.filter(
                or_(
                    unlocks_subq.c.unlock_count.is_(None),
                    unlocks_subq.c.unlock_count == 0,
                )
            )
        elif unlocks_range == "low_1_5":
            base_query = base_query.filter(unlocks_subq.c.unlock_count.between(1, 5))
        elif unlocks_range == "medium_6_10":
            base_query = base_query.filter(unlocks_subq.c.unlock_count.between(6, 10))
        elif unlocks_range == "moderate_11_20":
            base_query = base_query.filter(unlocks_subq.c.unlock_count.between(11, 20))
        elif unlocks_range == "heavy_20_plus":
            base_query = base_query.filter(unlocks_subq.c.unlock_count >= 21)

    # Get total count
    total_count = base_query.count()

    # Apply pagination
    contractors_data = base_query.offset((page - 1) * per_page).limit(per_page).all()

    # Build table data
    table_data = []
    for contractor in contractors_data:
        table_data.append(
            {
                "id": contractor.id,
                "company": contractor.company_name,
                "email": contractor.email,
                "phone": contractor.phone_number,
                "planTier": contractor.plan_name or "No Subscription",
                "subscriptionStatus": contractor.subscription_status or "inactive",
                "renewalDate": (
                    contractor.subscription_renew_date.strftime("%m/%d/%y")
                    if contractor.subscription_renew_date
                    else None
                ),
                "creditsBalance": contractor.current_credits or 0,
                "creditsSpent": contractor.total_spending or 0,
                "action": "disable" if contractor.is_active else "enable",
            }
        )

    return {
        "table": table_data,
        "pagination": {
            "total": total_count,
            "page": page,
            "perPage": per_page,
            "totalPages": (total_count + per_page - 1) // per_page,
        },
        "filters": {
            "accountStatus": account_status,
            "subscriptionStatus": subscription_status,
            "planTier": plan_tier,
            "creditsMin": credits_min,
            "creditsMax": credits_max,
            "unlocksRange": unlocks_range,
            "tradeCategory": trade_category,
            "serviceState": service_state,
            "registrationDate": registration_date,
            "activeDate": active_date,
            "search": search,
        },
    }


@router.get("/contractors-pending", dependencies=[Depends(require_admin_token)])
def contractors_pending(
    status: Optional[str] = Query(
        None,
        description="Filter by approval status: pending, rejected (default: shows both, pending first)",
    ),
    db: Session = Depends(get_db),
):
    """Admin endpoint: return list of contractors pending or rejected approval.

    Query param:
    - status: 'pending', 'rejected', or None/omitted (default: shows both, ordered by pending first)

    Returns contractors with: id, name, email, company, phone, license, user_type, created_at, approval_status
    """
    # Base query joining contractors and users
    base_query = db.query(
        models.user.Contractor.id,
        models.user.Contractor.company_name,
        models.user.Contractor.primary_contact_name,
        models.user.Contractor.phone_number,
        models.user.Contractor.state_license_number,
        models.user.Contractor.user_type,
        models.user.User.email,
        models.user.User.is_active,
        models.user.User.approved_by_admin,
        models.user.User.created_at,
        models.user.User.note,
    ).join(models.user.User, models.user.User.id == models.user.Contractor.user_id)

    # Apply status filter
    if status == "pending":
        base_query = base_query.filter(models.user.User.approved_by_admin == "pending")
        rows = base_query.order_by(models.user.User.created_at.desc()).all()
    elif status == "rejected":
        base_query = base_query.filter(models.user.User.approved_by_admin == "rejected")
        rows = base_query.order_by(models.user.User.created_at.desc()).all()
    else:
        # Default: show both pending and rejected, with pending first
        base_query = base_query.filter(
            models.user.User.approved_by_admin.in_(["pending", "rejected"])
        )
        # Order by: pending first (using CASE), then by created_at desc within each group
        rows = base_query.order_by(
            case(
                (models.user.User.approved_by_admin == "pending", 0),
                (models.user.User.approved_by_admin == "rejected", 1),
                else_=2,
            ),
            models.user.User.created_at.desc(),
        ).all()

    result = []
    for row in rows:
        result.append(
            {
                "id": row.id,
                "name": row.primary_contact_name,
                "email": row.email,
                "company": row.company_name,
                "phone": row.phone_number,
                "license_number": row.state_license_number,
                "user_type": row.user_type,
                "approval_status": row.approved_by_admin,
                "is_active": row.is_active,
                "admin_note": row.note,
                "created_at": (
                    row.created_at.strftime("%m/%d/%y %H:%M")
                    if row.created_at
                    else None
                ),
            }
        )

    return {"contractors": result, "count": len(result)}


@router.get(
    "/contractors/onboarding/{contractor_id}",
    dependencies=[Depends(require_admin_token)],
)
def contractor_onboarding_detail(contractor_id: int, db: Session = Depends(get_db)):
    """Admin endpoint: Get complete onboarding data for a contractor by contractor_id.

    Returns all data collected during the onboarding process:
    - User account details
    - Business information
    - License information (numbers, dates, status only - no documents)
    - Service areas

    Note: Documents (license_picture, referrals, job_photos) are uploaded via settings, not onboarding.
    """
    # Find contractor profile
    contractor = (
        db.query(models.user.Contractor)
        .filter(models.user.Contractor.id == contractor_id)
        .first()
    )
    if not contractor:
        raise HTTPException(status_code=404, detail="Contractor not found")

    # Find user
    user = (
        db.query(models.user.User)
        .filter(models.user.User.id == contractor.user_id)
        .first()
    )
    if not user:
        raise HTTPException(status_code=404, detail="User not found")

    return {
        # Step 0: User Account
        "account": {
            "user_id": user.id,
            "email": user.email,
            "email_verified": user.email_verified,
            "created_at": user.created_at.isoformat() if user.created_at else None,
            "is_active": user.is_active,
            "approved_by_admin": user.approved_by_admin,
            "admin_note": user.note,
        },
        # Step 1: Basic Business Information
        "business_info": {
            "company_name": contractor.company_name,
            "primary_contact_name": contractor.primary_contact_name,
            "phone_number": contractor.phone_number,
            "business_address": contractor.business_address,
            "website_url": contractor.website_url,
            "business_website_url": contractor.business_website_url,
        },
        # Step 2: License Information (without documents)
        "license_credentials": {
            "state_license_number": contractor.state_license_number,
            "license_expiration_date": contractor.license_expiration_date,
            "license_status": contractor.license_status,
        },
        # Step 3: Trade Information
        "trade_info": {
            "user_type": contractor.user_type,
            "trade_count": len(contractor.user_type) if contractor.user_type else 0,
        },
        # Step 4: Service Jurisdictions
        "service_areas": {
            "service_states": contractor.service_states,
            "state": contractor.state,
            "country_city": contractor.country_city,
            "states_count": (
                len(contractor.service_states)
                if contractor.service_states
                else (len(contractor.state) if contractor.state else 0)
            ),
            "cities_count": (
                len(contractor.country_city) if contractor.country_city else 0
            ),
        },
        # Onboarding Progress
        "onboarding": {
            "registration_step": contractor.registration_step,
            "is_completed": contractor.is_completed,
            "created_at": (
                contractor.created_at.isoformat() if contractor.created_at else None
            ),
            "updated_at": (
                contractor.updated_at.isoformat() if contractor.updated_at else None
            ),
        },
    }


@router.get("/contractors/search", dependencies=[Depends(require_admin_token)])
def search_contractors(
    q: str,
    status: Optional[str] = Query(
        None, description="Filter by status: active, disabled, all (default: all)"
    ),
    db: Session = Depends(get_db),
):
    """Admin endpoint: search contractors across all columns.

    Query params:
    - `q` - search string to match against any contractor field (required, min 2 chars)
    - `status` - filter by active/disabled status: 'active', 'disabled', or 'all' (optional, default: all)

    Returns matching contractors with full details.
    """
    if not q or len(q.strip()) < 2:
        raise HTTPException(
            status_code=400, detail="Search query must be at least 2 characters"
        )

    search_term = f"%{q.lower()}%"

    # Build status filter condition
    status_filter = ""
    if status == "active":
        status_filter = "AND u.is_active = true"
    elif status == "disabled":
        status_filter = "AND u.is_active = false"
    # If status is None or "all", don't add any filter

    # Build comprehensive search query across all text columns
    # Using raw SQL for flexible ILIKE search across all columns
    query = text(
        f"""
        SELECT 
            c.id,
            c.company_name,
            c.primary_contact_name,
            c.phone_number,
            c.state_license_number,
            c.user_type,
            u.email,
            u.is_active,
            u.approved_by_admin
        FROM contractors c
        JOIN users u ON u.id = c.user_id
        WHERE 
            (
                LOWER(COALESCE(c.company_name, '')) LIKE :search
                OR LOWER(COALESCE(c.primary_contact_name, '')) LIKE :search
                OR LOWER(COALESCE(c.phone_number, '')) LIKE :search
                OR LOWER(COALESCE(c.website_url, '')) LIKE :search
                OR LOWER(COALESCE(c.business_address, '')) LIKE :search
                OR LOWER(COALESCE(c.business_website_url, '')) LIKE :search
                OR EXISTS (
                    SELECT 1 FROM jsonb_array_elements_text(COALESCE(c.state_license_number::jsonb, '[]'::jsonb)) AS license
                    WHERE LOWER(license) LIKE :search
                )
                OR EXISTS (
                    SELECT 1 FROM jsonb_array_elements_text(COALESCE(c.license_status::jsonb, '[]'::jsonb)) AS status
                    WHERE LOWER(status) LIKE :search
                )
                OR LOWER(COALESCE(u.email, '')) LIKE :search
                OR LOWER(ARRAY_TO_STRING(c.user_type, ',')) LIKE :search
                OR LOWER(ARRAY_TO_STRING(c.state, ',')) LIKE :search
                OR LOWER(ARRAY_TO_STRING(c.country_city, ',')) LIKE :search
            )
            AND u.approved_by_admin = 'approved'
            {status_filter}
        ORDER BY c.id DESC
        LIMIT 100
    """
    )

    rows = db.execute(query, {"search": search_term}).fetchall()

    result = []
    for row in rows:
        action = "disable" if row.is_active else "enable"
        result.append(
            {
                "id": row.id,
                "phone_number": row.phone_number,
                "email": row.email,
                "company": row.company_name,
                "license_number": row.state_license_number,
                "user_type": row.user_type,
                "is_active": row.is_active,
                "approved_by_admin": row.approved_by_admin,
                "action": action,
            }
        )

    return {"contractors": result, "count": len(result)}


@router.get("/contractors/search-pending", dependencies=[Depends(require_admin_token)])
def search_contractors_pending(q: str, db: Session = Depends(get_db)):
    """Admin endpoint: search pending contractors across all columns.

    Query param: `q` - search string to match against any contractor field.
    Returns only contractors with approved_by_admin = 'pending'.
    """
    if not q or len(q.strip()) < 2:
        raise HTTPException(
            status_code=400, detail="Search query must be at least 2 characters"
        )

    search_term = f"%{q.lower()}%"

    # Build comprehensive search query across all text columns
    # Using raw SQL for flexible ILIKE search across all columns
    query = text(
        """
        SELECT 
            c.id,
            c.company_name,
            c.primary_contact_name,
            c.phone_number,
            c.state_license_number,
            c.user_type,
            u.email,
            u.is_active,
            u.created_at
        FROM contractors c
        JOIN users u ON u.id = c.user_id
        WHERE 
            LOWER(COALESCE(c.company_name, '')) LIKE :search
            OR LOWER(COALESCE(c.primary_contact_name, '')) LIKE :search
            OR LOWER(COALESCE(c.phone_number, '')) LIKE :search
            OR LOWER(COALESCE(c.website_url, '')) LIKE :search
            OR LOWER(COALESCE(c.business_address, '')) LIKE :search
            OR LOWER(COALESCE(c.business_website_url, '')) LIKE :search
            OR EXISTS (
                SELECT 1 FROM jsonb_array_elements_text(COALESCE(c.state_license_number::jsonb, '[]'::jsonb)) AS license
                WHERE LOWER(license) LIKE :search
            )
            OR EXISTS (
                SELECT 1 FROM jsonb_array_elements_text(COALESCE(c.license_status::jsonb, '[]'::jsonb)) AS status
                WHERE LOWER(status) LIKE :search
            )
            OR LOWER(COALESCE(u.email, '')) LIKE :search
            OR LOWER(ARRAY_TO_STRING(c.user_type, ',')) LIKE :search
            OR LOWER(ARRAY_TO_STRING(c.state, ',')) LIKE :search
            OR LOWER(ARRAY_TO_STRING(c.country_city, ',')) LIKE :search
        AND u.approved_by_admin = 'pending'
        ORDER BY c.id DESC
        LIMIT 100
    """
    )

    rows = db.execute(query, {"search": search_term}).fetchall()

    result = []
    for row in rows:
        action = "disable" if row.is_active else "enable"
        result.append(
            {
                "id": row.id,
                "phone_number": row.phone_number,
                "email": row.email,
                "company": row.company_name,
                "license_number": row.state_license_number,
                "user_type": row.user_type,
                "action": action,
                "created_at": row.created_at.isoformat() if row.created_at else None,
            }
        )

    return {"contractors": result}


@router.get(
    "/ingested-jobs",
    dependencies=[Depends(require_admin_token)],
    summary="Job Posted Requested",
)
def ingested_jobs(db: Session = Depends(get_db)):
    """Admin endpoint: list contractor-uploaded jobs requested for posting.

    Returns entries with: id, permit_type, permit_value, job_review_status,
    address_code, job_address. Jobs where `job_review_status` is `pending` or
    `declined` and `uploaded_by_contractor == True` are returned.
    """
    rows = (
        db.query(models.user.Job)
        .filter(
            models.user.Job.uploaded_by_contractor == True,
            models.user.Job.job_review_status.in_(["pending", "declined"]),
        )
        .all()
    )

    result = []
    for j in rows:
        result.append(
            {
                "id": j.id,
                "permit_type": j.permit_type,
                "permit_value": j.job_cost,
                "job_review_status": j.job_review_status,
                "address_code": j.permit_number,
                "job_address": j.job_address,
                "uploaded_by_user_id": j.uploaded_by_user_id,
                "created_at": (
                    j.created_at.isoformat() if getattr(j, "created_at", None) else None
                ),
            }
        )

    return {"ingested_jobs": result}


@router.get(
    "/contractor-uploaded-jobs",
    dependencies=[Depends(require_admin_token)],
    summary="Contractor-Uploaded Jobs with KPIs and Filters",
)
def contractor_uploaded_jobs(
    # Review Status Filters
    status_pending: Optional[bool] = Query(None, description="Filter pending jobs"),
    status_posted: Optional[bool] = Query(None, description="Filter posted jobs"),
    status_declined: Optional[bool] = Query(None, description="Filter declined jobs"),
    # Date Filters
    date_from: Optional[str] = Query(None, description="Start date (YYYY-MM-DD)"),
    date_to: Optional[str] = Query(None, description="End date (YYYY-MM-DD)"),
    # User Filter
    uploaded_by_user_id: Optional[int] = Query(None, description="Filter by user ID who uploaded"),
    # Property Type Filter
    property_type: Optional[str] = Query(None, description="Filter by property type"),
    # Pagination
    page: int = Query(1, ge=1, description="Page number"),
    per_page: int = Query(100, ge=1, le=500, description="Items per page"),
    db: Session = Depends(get_db),
):
    """Admin endpoint: Contractor-uploaded jobs screen with KPIs and filters.

    Returns KPIs and all contractor-uploaded jobs (uploaded_by_contractor == True).
    Supports filtering by status, date range, uploaded by user, and property type.
    """
    from datetime import datetime, timedelta

    # Helper function to calculate percentage change
    def calc_percentage_change(current, previous):
        if previous == 0:
            return 0 if current == 0 else 100
        return round(((current - previous) / previous) * 100, 1)

    # ========================================================================
    # KPIs Calculation (only for contractor-uploaded jobs)
    # ========================================================================

    # Total Contractor-Uploaded Jobs
    total_jobs = (
        db.query(func.count(models.user.Job.id))
        .filter(models.user.Job.uploaded_by_contractor == True)
        .scalar()
        or 0
    )

    # Submitted Today
    today_start = datetime.utcnow().replace(hour=0, minute=0, second=0, microsecond=0)
    submitted_today = (
        db.query(func.count(models.user.Job.id))
        .filter(
            models.user.Job.uploaded_by_contractor == True,
            models.user.Job.created_at >= today_start,
        )
        .scalar()
        or 0
    )

    # Submitted Yesterday (for percentage change)
    yesterday_start = today_start - timedelta(days=1)
    submitted_yesterday = (
        db.query(func.count(models.user.Job.id))
        .filter(
            models.user.Job.uploaded_by_contractor == True,
            models.user.Job.created_at >= yesterday_start,
            models.user.Job.created_at < today_start,
        )
        .scalar()
        or 0
    )

    # Pending Review
    pending_review = (
        db.query(func.count(models.user.Job.id))
        .filter(
            models.user.Job.uploaded_by_contractor == True,
            models.user.Job.job_review_status == "pending",
        )
        .scalar()
        or 0
    )

    # Previous Pending (7 days ago for comparison)
    seven_days_ago = today_start - timedelta(days=7)
    previous_pending = (
        db.query(func.count(models.user.Job.id))
        .filter(
            models.user.Job.uploaded_by_contractor == True,
            models.user.Job.job_review_status == "pending",
            models.user.Job.created_at < seven_days_ago,
        )
        .scalar()
        or 0
    )

    # Posted Jobs
    posted = (
        db.query(func.count(models.user.Job.id))
        .filter(
            models.user.Job.uploaded_by_contractor == True,
            models.user.Job.job_review_status == "posted",
        )
        .scalar()
        or 0
    )

    # Previous Posted (7 days ago)
    previous_posted = (
        db.query(func.count(models.user.Job.id))
        .filter(
            models.user.Job.uploaded_by_contractor == True,
            models.user.Job.job_review_status == "posted",
            models.user.Job.created_at < seven_days_ago,
        )
        .scalar()
        or 0
    )

    # Declined Jobs
    declined = (
        db.query(func.count(models.user.Job.id))
        .filter(
            models.user.Job.uploaded_by_contractor == True,
            models.user.Job.job_review_status == "declined",
        )
        .scalar()
        or 0
    )

    # Previous Declined (7 days ago for comparison)
    previous_declined = (
        db.query(func.count(models.user.Job.id))
        .filter(
            models.user.Job.uploaded_by_contractor == True,
            models.user.Job.job_review_status == "declined",
            models.user.Job.created_at < seven_days_ago,
        )
        .scalar()
        or 0
    )

    # Unlock Conversion Rate for contractor-uploaded jobs
    # = number of unique contractor jobs that have been unlocked / total posted * 100
    unlocked_contractor_jobs = (
        db.query(func.count(func.distinct(models.user.UnlockedLead.job_id)))
        .join(models.user.Job, models.user.UnlockedLead.job_id == models.user.Job.id)
        .filter(
            models.user.Job.uploaded_by_contractor == True,
            models.user.Job.job_review_status == "posted",
        )
        .scalar()
        or 0
    )
    unlock_conversion_rate = (
        round((unlocked_contractor_jobs / posted) * 100, 1) if posted > 0 else 0
    )

    # Previous Unlock Conversion Rate (7 days ago)
    unlocked_contractor_jobs_prev = (
        db.query(func.count(func.distinct(models.user.UnlockedLead.job_id)))
        .join(models.user.Job, models.user.UnlockedLead.job_id == models.user.Job.id)
        .filter(
            models.user.Job.uploaded_by_contractor == True,
            models.user.Job.job_review_status == "posted",
            models.user.UnlockedLead.unlocked_at < seven_days_ago,
        )
        .scalar()
        or 0
    )
    prev_unlock_conversion_rate = (
        round((unlocked_contractor_jobs_prev / previous_posted) * 100, 1)
        if previous_posted > 0
        else 0
    )

    # Average Per Day (last 30 days)
    thirty_days_ago = today_start - timedelta(days=30)
    jobs_last_30_days = (
        db.query(func.count(models.user.Job.id))
        .filter(
            models.user.Job.uploaded_by_contractor == True,
            models.user.Job.created_at >= thirty_days_ago,
        )
        .scalar()
        or 0
    )
    avg_per_day = round(jobs_last_30_days / 30, 1)

    # Previous 30 days average
    sixty_days_ago = today_start - timedelta(days=60)
    jobs_prev_30_days = (
        db.query(func.count(models.user.Job.id))
        .filter(
            models.user.Job.uploaded_by_contractor == True,
            models.user.Job.created_at >= sixty_days_ago,
            models.user.Job.created_at < thirty_days_ago,
        )
        .scalar()
        or 0
    )
    prev_avg_per_day = round(jobs_prev_30_days / 30, 1)

    # ========================================================================
    # Apply Filters to Base Query
    # ========================================================================

    # Base query: ONLY contractor-uploaded jobs
    base_query = db.query(models.user.Job).filter(
        models.user.Job.uploaded_by_contractor == True
    )

    # Review Status Filters (pending, posted, declined)
    status_filters = []
    if status_pending:
        status_filters.append(models.user.Job.job_review_status == "pending")
    if status_posted:
        status_filters.append(models.user.Job.job_review_status == "posted")
    if status_declined:
        status_filters.append(models.user.Job.job_review_status == "declined")

    if status_filters:
        base_query = base_query.filter(or_(*status_filters))

    # Date Filters
    if date_from:
        try:
            from_date = datetime.strptime(date_from, "%Y-%m-%d")
            base_query = base_query.filter(models.user.Job.created_at >= from_date)
        except ValueError:
            pass  # Invalid date format, skip filter

    if date_to:
        try:
            to_date = datetime.strptime(date_to, "%Y-%m-%d").replace(
                hour=23, minute=59, second=59
            )
            base_query = base_query.filter(models.user.Job.created_at <= to_date)
        except ValueError:
            pass  # Invalid date format, skip filter

    # User Filter
    if uploaded_by_user_id:
        base_query = base_query.filter(
            models.user.Job.uploaded_by_user_id == uploaded_by_user_id
        )

    # Property Type Filter
    if property_type:
        base_query = base_query.filter(
            func.lower(models.user.Job.property_type).like(f"%{property_type.lower()}%")
        )

    # ========================================================================
    # Pagination and Results
    # ========================================================================

    # Order by status priority (pending first) then by created_at descending
    status_order = case(
        (models.user.Job.job_review_status == "pending", 1),
        (models.user.Job.job_review_status == "posted", 2),
        else_=3,
    )

    # Get total count after filters
    total_count = base_query.count()
    total_pages = (total_count + per_page - 1) // per_page

    # Apply pagination
    offset = (page - 1) * per_page
    rows = (
        base_query.order_by(status_order, models.user.Job.created_at.desc())
        .limit(per_page)
        .offset(offset)
        .all()
    )

    jobs_data = []
    for j in rows:
        jobs_data.append(
            {
                "id": j.id,
                "permit_type": j.audience_type_names,
                "job_address": j.job_address,
                "contact_name": j.contact_name,
                "contractor_name": j.contractor_name,
                "contractor_email": j.contractor_email,
                "project_cost_total": j.project_cost_total,
                "trs_score": j.trs_score,
                "property_type": j.property_type,
                "job_review_status": j.job_review_status,
                "created_at": j.created_at.isoformat() if j.created_at else None,
            }
        )

    # ========================================================================
    # Return KPIs + Jobs Data
    # ========================================================================
    return {
        "kpis": {
            "total_jobs": {
                "value": total_jobs,
                "change": 0,  # Total is cumulative
            },
            "submitted_today": {
                "value": submitted_today,
                "change": calc_percentage_change(submitted_today, submitted_yesterday),
            },
            "pending_review": {
                "value": pending_review,
                "change": calc_percentage_change(pending_review, previous_pending),
            },
            "posted": {
                "value": posted,
                "change": calc_percentage_change(posted, previous_posted),
            },
            "declined": {
                "value": declined,
                "change": calc_percentage_change(declined, previous_declined),
            },
            "unlock_conversion_rate": {
                "value": unlock_conversion_rate,
                "change": calc_percentage_change(
                    unlock_conversion_rate, prev_unlock_conversion_rate
                ),
            },
            "avg_per_day": {
                "value": avg_per_day,
                "change": calc_percentage_change(avg_per_day, prev_avg_per_day),
            },
        },
        "jobs": jobs_data,
        "pagination": {
            "page": page,
            "per_page": per_page,
            "total_items": total_count,
            "total_pages": total_pages,
        },
    }


@router.get(
    "/contractor-uploaded-jobs/search",
    dependencies=[Depends(require_admin_token)],
    summary="Simple Text Search in Contractor-Uploaded Jobs",
)
def search_contractor_uploaded_jobs(
    q: str = Query(..., min_length=2, description="Search query (min 2 characters)"),
    page: int = Query(1, ge=1, description="Page number"),
    per_page: int = Query(25, ge=1, le=100, description="Items per page"),
    db: Session = Depends(get_db),
):
    """Admin endpoint: Simple text search in contractor-uploaded jobs.

    Searches across:
    - Job address
    - Permit number
    - Project description
    - Contractor name
    - Contractor company
    - Audience type (permit type)
    """
    # Build search query
    search_term = f"%{q.lower()}%"

    base_query = db.query(models.user.Job).filter(
        models.user.Job.uploaded_by_contractor == True,
        or_(
            func.lower(models.user.Job.job_address).like(search_term),
            func.lower(models.user.Job.permit_number).like(search_term),
            func.lower(models.user.Job.project_description).like(search_term),
            func.lower(models.user.Job.contractor_name).like(search_term),
            func.lower(models.user.Job.contractor_company).like(search_term),
            func.lower(models.user.Job.audience_type_names).like(search_term),
        ),
    )

    # Get total count
    total_count = base_query.count()
    total_pages = (total_count + per_page - 1) // per_page

    # Order by status priority (pending first) then by created_at descending
    status_order = case(
        (models.user.Job.job_review_status == "pending", 1),
        (models.user.Job.job_review_status == "posted", 2),
        else_=3,
    )

    # Apply pagination
    offset = (page - 1) * per_page
    rows = (
        base_query.order_by(status_order, models.user.Job.created_at.desc())
        .limit(per_page)
        .offset(offset)
        .all()
    )

    jobs_data = []
    for j in rows:
        jobs_data.append(
            {
                "id": j.id,
                "permit_type": j.audience_type_names,
                "contact_name": j.contact_name,
                "contractor_email": j.contractor_email,
                "trs_score": j.trs_score,
                "job_review_status": j.job_review_status,
            }
        )

    # Return search results
    return {
        "jobs": jobs_data,
        "pagination": {
            "page": page,
            "per_page": per_page,
            "total_items": total_count,
            "total_pages": total_pages,
        },
        "search_query": q,
    }


@router.patch(
    "/ingested-jobs/{job_id:int}/post",
    dependencies=[Depends(require_admin_or_editor)],
)
def post_ingested_job(job_id: int, db: Session = Depends(get_db)):
    """Admin-only: Approve a job for scheduled posting.

    Works for BOTH contractor-uploaded and system-ingested jobs:
    - Sets review_posted_at = now() (triggers scheduler)
    - Sets uploaded_by_contractor = False (converts contractor jobs to system jobs)
    - Keeps job_review_status = 'pending'
    - Scheduler posts job when: review_posted_at + day_offset <= current_time
    
    Use cases:
    1. Contractor jobs: Admin approval starts the posting schedule
    2. System jobs (toggle OFF): Admin manually approves for scheduled posting
    """
    j = db.query(models.user.Job).filter(models.user.Job.id == job_id).first()
    if not j:
        raise HTTPException(status_code=404, detail="Job not found")

    # Get current time in EST
    est_tz = ZoneInfo("America/New_York")
    now_est = datetime.now(est_tz).replace(tzinfo=None)

    # Set review timestamp and convert to system job
    j.uploaded_by_contractor = False
    j.review_posted_at = now_est
    # Keep job_review_status as 'pending' - scheduler will update to 'posted' based on day_offset
    
    db.add(j)
    db.commit()
    db.refresh(j)

    # Calculate when scheduler will post this job
    from datetime import timedelta
    day_offset = j.day_offset if j.day_offset is not None else 0
    estimated_post_time = now_est + timedelta(days=day_offset)

    return {
        "job_id": j.id,
        "job_review_status": j.job_review_status,
        "uploaded_by_contractor": j.uploaded_by_contractor,
        "review_posted_at": now_est.isoformat(),
        "day_offset": day_offset,
        "estimated_post_time": estimated_post_time.isoformat(),
        "message": f"Job approved. Scheduler will post in {day_offset} day(s) at approximately {estimated_post_time.strftime('%Y-%m-%d %I:%M %p')} EST.",
    }


class DeclineJobRequest(BaseModel):
    note: str  # Required: admin's reason for declining the job


@router.patch(
    "/ingested-jobs/{job_id:int}/decline",
    dependencies=[Depends(require_admin_or_editor)],
)
def decline_ingested_job(
    job_id: int,
    body: DeclineJobRequest,
    db: Session = Depends(get_db),
):
    """Admin/Editor: mark an ingested job as declined with a reason note.

    Sets `job_review_status` to `declined` and saves the admin's note explaining why.
    """
    # Ensure decline_note column exists (auto-migration)
    try:
        db.execute(
            text(
                "ALTER TABLE jobs ADD COLUMN IF NOT EXISTS decline_note TEXT"
            )
        )
        db.commit()
    except Exception:
        db.rollback()

    j = db.query(models.user.Job).filter(models.user.Job.id == job_id).first()
    if not j:
        raise HTTPException(status_code=404, detail="Job not found")

    j.job_review_status = "declined"
    j.decline_note = body.note
    db.add(j)
    db.commit()

    return {
        "job_id": j.id,
        "job_review_status": j.job_review_status,
        "decline_note": j.decline_note,
        "message": "Job marked as declined.",
    }



@router.get(
    "/ingested-jobs/system",
    dependencies=[Depends(require_admin_token)],
    summary="System-ingested Jobs with KPIs and Filters",
)
def system_ingested_jobs(
    # Review Status Filters (only pending and posted for system jobs)
    status_pending: Optional[bool] = Query(None, description="Filter pending jobs"),
    status_posted: Optional[bool] = Query(None, description="Filter posted jobs"),
    # Date Filters
    quick_date: Optional[str] = Query(
        None, description="Quick date filter: last_7_days, last_30_days, last_90_days"
    ),
    date_from: Optional[str] = Query(
        None, description="Custom start date (YYYY-MM-DD)"
    ),
    date_to: Optional[str] = Query(None, description="Custom end date (YYYY-MM-DD)"),
    # Permit Filters
    permit_status: Optional[str] = Query(None, description="Filter by permit status"),
    permit_type: Optional[str] = Query(None, description="Filter by permit type"),
    # Location Filters
    source_county: Optional[str] = Query(None, description="Filter by source county"),
    state: Optional[str] = Query(None, description="Filter by state"),
    # Project Cost Range
    cost_min: Optional[float] = Query(None, description="Minimum project cost"),
    cost_max: Optional[float] = Query(None, description="Maximum project cost"),
    # Audience Type
    audience_type: Optional[str] = Query(
        None, description="Filter by audience type (user type)"
    ),
    # Contractor Search
    contractor_name: Optional[str] = Query(
        None, description="Search by contractor name"
    ),
    # Pagination
    page: int = Query(1, ge=1, description="Page number"),
    per_page: int = Query(100, ge=1, le=500, description="Items per page"),
    db: Session = Depends(get_db),
):
    """Admin/Editor endpoint: list jobs NOT uploaded by contractors with KPIs and advanced filters.

    Returns KPIs and all fields from system-ingested jobs (uploaded via /jobs/upload-leads-json).
    Supports all Advanced Filter UI options including status, dates, permit, location, property type, cost range, etc.
    """
    from datetime import datetime, timedelta

    # Helper function to calculate percentage change
    def calc_percentage_change(current, previous):
        if previous == 0:
            return 0 if current == 0 else 100
        return round(((current - previous) / previous) * 100, 1)

    # ========================================================================
    # KPIs Calculation (only for system-ingested jobs)
    # ========================================================================

    # Total Jobs (system-ingested)
    total_jobs = (
        db.query(func.count(models.user.Job.id))
        .filter(
            models.user.Job.uploaded_by_contractor.is_(False),
            models.user.Job.uploaded_by_user_id.is_(None)
        )
        .scalar()
        or 0
    )

    # Submitted Today
    today_start = datetime.utcnow().replace(hour=0, minute=0, second=0, microsecond=0)
    submitted_today = (
        db.query(func.count(models.user.Job.id))
        .filter(
            models.user.Job.uploaded_by_contractor.is_(False),
            models.user.Job.uploaded_by_user_id.is_(None),
            models.user.Job.created_at >= today_start,
        )
        .scalar()
        or 0
    )

    # Submitted Yesterday (for percentage change)
    yesterday_start = today_start - timedelta(days=1)
    submitted_yesterday = (
        db.query(func.count(models.user.Job.id))
        .filter(
            models.user.Job.uploaded_by_contractor.is_(False),
            models.user.Job.uploaded_by_user_id.is_(None),
            models.user.Job.created_at >= yesterday_start,
            models.user.Job.created_at < today_start,
        )
        .scalar()
        or 0
    )

    # Pending Review
    pending_review = (
        db.query(func.count(models.user.Job.id))
        .filter(
            models.user.Job.uploaded_by_contractor.is_(False),
            models.user.Job.uploaded_by_user_id.is_(None),
            models.user.Job.job_review_status == "pending",
        )
        .scalar()
        or 0
    )

    # Previous Pending (7 days ago for comparison)
    seven_days_ago = today_start - timedelta(days=7)
    previous_pending = (
        db.query(func.count(models.user.Job.id))
        .filter(
            models.user.Job.uploaded_by_contractor.is_(False),
            models.user.Job.uploaded_by_user_id.is_(None),
            models.user.Job.job_review_status == "pending",
            models.user.Job.created_at < seven_days_ago,
        )
        .scalar()
        or 0
    )

    # Posted Jobs
    posted = (
        db.query(func.count(models.user.Job.id))
        .filter(
            models.user.Job.uploaded_by_contractor.is_(False),
            models.user.Job.uploaded_by_user_id.is_(None),
            models.user.Job.job_review_status == "posted",
        )
        .scalar()
        or 0
    )

    # Previous Posted (7 days ago)
    previous_posted = (
        db.query(func.count(models.user.Job.id))
        .filter(
            models.user.Job.uploaded_by_contractor.is_(False),
            models.user.Job.uploaded_by_user_id.is_(None),
            models.user.Job.job_review_status == "posted",
            models.user.Job.review_posted_at < seven_days_ago,
        )
        .scalar()
        or 0
    )

    # Average Per Day (last 30 days)
    thirty_days_ago = today_start - timedelta(days=30)
    jobs_last_30_days = (
        db.query(func.count(models.user.Job.id))
        .filter(
            models.user.Job.uploaded_by_contractor.is_(False),
            models.user.Job.uploaded_by_user_id.is_(None),
            models.user.Job.created_at >= thirty_days_ago,
        )
        .scalar()
        or 0
    )
    avg_per_day = round(jobs_last_30_days / 30, 1)

    # Previous 30 days average
    sixty_days_ago = today_start - timedelta(days=60)
    jobs_prev_30_days = (
        db.query(func.count(models.user.Job.id))
        .filter(
            models.user.Job.uploaded_by_contractor.is_(False),
            models.user.Job.uploaded_by_user_id.is_(None),
            models.user.Job.created_at >= sixty_days_ago,
            models.user.Job.created_at < thirty_days_ago,
        )
        .scalar()
        or 0
    )
    prev_avg_per_day = round(jobs_prev_30_days / 30, 1) if jobs_prev_30_days > 0 else 0

    # Unlock Conversion Rate (unlocks / posted jobs)
    # Count unlocks for system-ingested jobs only
    unlocked_system_jobs = (
        db.query(func.count(func.distinct(models.user.UnlockedLead.job_id)))
        .join(models.user.Job, models.user.UnlockedLead.job_id == models.user.Job.id)
        .filter(
            models.user.Job.uploaded_by_contractor.is_(False),
            models.user.Job.uploaded_by_user_id.is_(None)
        )
        .scalar()
        or 0
    )

    unlock_conversion_rate = (
        round((unlocked_system_jobs / posted) * 100, 1) if posted > 0 else 0
    )

    # Previous unlock conversion rate (7 days ago)
    unlocked_system_jobs_prev = (
        db.query(func.count(func.distinct(models.user.UnlockedLead.job_id)))
        .join(models.user.Job, models.user.UnlockedLead.job_id == models.user.Job.id)
        .filter(
            models.user.Job.uploaded_by_contractor.is_(False),
            models.user.Job.uploaded_by_user_id.is_(None),
            models.user.UnlockedLead.unlocked_at < seven_days_ago,
        )
        .scalar()
        or 0
    )

    prev_unlock_conversion_rate = (
        round((unlocked_system_jobs_prev / previous_posted) * 100, 1)
        if previous_posted > 0
        else 0
    )

    # ========================================================================
    # Apply Filters to Query
    # ========================================================================

    base_query = db.query(models.user.Job).filter(
        models.user.Job.uploaded_by_contractor.is_(False),
        models.user.Job.uploaded_by_user_id.is_(None)
    )

    # Review Status Filters (pending or posted only)
    status_filters = []
    if status_pending:
        status_filters.append(models.user.Job.job_review_status == "pending")
    if status_posted:
        status_filters.append(models.user.Job.job_review_status == "posted")

    if status_filters:
        base_query = base_query.filter(or_(*status_filters))

    # Date Filters
    if quick_date:
        if quick_date == "last_7_days":
            date_threshold = today_start - timedelta(days=7)
            base_query = base_query.filter(models.user.Job.created_at >= date_threshold)
        elif quick_date == "last_30_days":
            date_threshold = today_start - timedelta(days=30)
            base_query = base_query.filter(models.user.Job.created_at >= date_threshold)
        elif quick_date == "last_90_days":
            date_threshold = today_start - timedelta(days=90)
            base_query = base_query.filter(models.user.Job.created_at >= date_threshold)

    # Custom Date Range
    if date_from:
        try:
            from_date = datetime.strptime(date_from, "%Y-%m-%d")
            base_query = base_query.filter(models.user.Job.created_at >= from_date)
        except ValueError:
            pass

    if date_to:
        try:
            to_date = datetime.strptime(date_to, "%Y-%m-%d")
            to_date = to_date + timedelta(days=1)
            base_query = base_query.filter(models.user.Job.created_at < to_date)
        except ValueError:
            pass

    # Permit Filters
    if permit_status:
        base_query = base_query.filter(
            func.lower(models.user.Job.permit_status) == permit_status.lower()
        )

    if permit_type:
        base_query = base_query.filter(
            func.lower(models.user.Job.permit_type_norm).contains(
                permit_type.lower()
            )
        )

    # Location Filters
    if source_county:
        base_query = base_query.filter(
            func.lower(models.user.Job.source_county).contains(source_county.lower())
        )

    if state:
        base_query = base_query.filter(
            func.lower(models.user.Job.state) == state.lower()
        )

    # Project Cost Range (use project_cost_total which is numeric)
    if cost_min is not None:
        base_query = base_query.filter(models.user.Job.project_cost_total >= cost_min)

    if cost_max is not None:
        base_query = base_query.filter(models.user.Job.project_cost_total <= cost_max)

    # Audience Type Filter
    if audience_type:
        base_query = base_query.filter(
            func.lower(models.user.Job.audience_type_names).contains(
                audience_type.lower()
            )
        )

    # Contractor Name Search
    if contractor_name:
        search_term = f"%{contractor_name.lower()}%"
        base_query = base_query.filter(
            or_(
                func.lower(models.user.Job.contractor_name).like(search_term),
                func.lower(models.user.Job.contractor_company).like(search_term),
            )
        )

    # ========================================================================
    # Jobs Data with Pagination
    # ========================================================================

    # Order by status priority (pending first) then by created_at descending
    status_order = case(
        (models.user.Job.job_review_status == "pending", 1),
        (models.user.Job.job_review_status == "posted", 2),
        else_=3,
    )

    # Get total count for pagination
    total_count = base_query.count()
    total_pages = (total_count + per_page - 1) // per_page

    # Apply pagination
    offset = (page - 1) * per_page
    rows = (
        base_query.order_by(status_order, models.user.Job.created_at.desc())
        .limit(per_page)
        .offset(offset)
        .all()
    )

    jobs_data = []
    for j in rows:
        jobs_data.append(
            {
                "id": j.id,
                "permit_type": j.audience_type_names,
                "contact_name": j.contact_name,
                "contractor_email": j.contractor_email,
                "trs_score": j.trs_score,
                "job_review_status": j.job_review_status,
            }
        )

    # ========================================================================
    # Return KPIs + Jobs Data
    # ========================================================================

    return {
        "kpis": {
            "totalJobs": {
                "value": total_jobs,
                "change": 0,  # Total is cumulative, no change metric
            },
            "submittedToday": {
                "value": submitted_today,
                "change": calc_percentage_change(submitted_today, submitted_yesterday),
            },
            "pendingReview": {
                "value": pending_review,
                "change": calc_percentage_change(pending_review, previous_pending),
            },
            "posted": {
                "value": posted,
                "change": calc_percentage_change(posted, previous_posted),
            },
            "avgPerDay": {
                "value": avg_per_day,
                "change": calc_percentage_change(avg_per_day, prev_avg_per_day),
            },
            "unlockConversionRate": {
                "value": unlock_conversion_rate,
                "change": calc_percentage_change(
                    unlock_conversion_rate, prev_unlock_conversion_rate
                ),
            },
            "pendingReviewJobs": {
                "value": pending_review,
                "change": calc_percentage_change(pending_review, previous_pending),
            },
            "systemIngestedJobs": {
                "value": total_jobs,
                "change": 0,  # Total is cumulative
            },
        },
        "jobs": jobs_data,
        "pagination": {
            "page": page,
            "per_page": per_page,
            "total_items": total_count,
            "total_pages": total_pages,
        },
    }


@router.get(
    "/ingested-jobs/posted",
    dependencies=[Depends(require_admin_token)],
    summary="Posted Jobs with KPIs and Filters",
)
def posted_jobs(
    # Date Filters
    quick_date: Optional[str] = Query(
        None, description="Quick date filter: last_7_days, last_30_days, last_90_days"
    ),
    date_from: Optional[str] = Query(
        None, description="Custom start date (YYYY-MM-DD)"
    ),
    date_to: Optional[str] = Query(None, description="Custom end date (YYYY-MM-DD)"),
    # Permit Filters
    permit_status: Optional[str] = Query(None, description="Filter by permit status"),
    permit_type: Optional[str] = Query(None, description="Filter by permit type"),
    # Location Filters
    source_county: Optional[str] = Query(None, description="Filter by source county"),
    state: Optional[str] = Query(None, description="Filter by state"),
    # Project Cost Range
    cost_min: Optional[float] = Query(None, description="Minimum project cost"),
    cost_max: Optional[float] = Query(None, description="Maximum project cost"),
    # Audience Type
    audience_type: Optional[str] = Query(
        None, description="Filter by audience type (user type)"
    ),
    # Contractor Search
    contractor_name: Optional[str] = Query(
        None, description="Search by contractor name"
    ),
    # Pagination
    page: int = Query(1, ge=1, description="Page number"),
    per_page: int = Query(100, ge=1, le=500, description="Items per page"),
    db: Session = Depends(get_db),
):
    """Admin endpoint: list all POSTED jobs (job_review_status = 'posted') with KPIs and filters.

    Returns KPIs specific to posted jobs and all fields from posted jobs table.
    Supports filtering by date, permit, location, property type, cost range, etc.
    """
    from datetime import datetime, timedelta

    # Helper function to calculate percentage change
    def calc_percentage_change(current, previous):
        if previous == 0:
            return 0 if current == 0 else 100
        return round(((current - previous) / previous) * 100, 1)

    # ========================================================================
    # KPIs Calculation (only for POSTED jobs)
    # ========================================================================

    # Total Posted Jobs
    total_posted_jobs = (
        db.query(func.count(models.user.Job.id))
        .filter(
            models.user.Job.uploaded_by_contractor == False,
            models.user.Job.job_review_status == "posted",
        )
        .scalar()
        or 0
    )

    # Posted Today
    today_start = datetime.utcnow().replace(hour=0, minute=0, second=0, microsecond=0)
    posted_today = (
        db.query(func.count(models.user.Job.id))
        .filter(
            models.user.Job.uploaded_by_contractor == False,
            models.user.Job.job_review_status == "posted",
            models.user.Job.review_posted_at >= today_start,
        )
        .scalar()
        or 0
    )

    # Posted Yesterday (for percentage change)
    yesterday_start = today_start - timedelta(days=1)
    posted_yesterday = (
        db.query(func.count(models.user.Job.id))
        .filter(
            models.user.Job.uploaded_by_contractor == False,
            models.user.Job.job_review_status == "posted",
            models.user.Job.review_posted_at >= yesterday_start,
            models.user.Job.review_posted_at < today_start,
        )
        .scalar()
        or 0
    )

    # Posted Last 7 Days
    seven_days_ago = today_start - timedelta(days=7)
    posted_last_7_days = (
        db.query(func.count(models.user.Job.id))
        .filter(
            models.user.Job.uploaded_by_contractor == False,
            models.user.Job.job_review_status == "posted",
            models.user.Job.review_posted_at >= seven_days_ago,
        )
        .scalar()
        or 0
    )

    # Previous 7 days (for comparison)
    fourteen_days_ago = today_start - timedelta(days=14)
    posted_prev_7_days = (
        db.query(func.count(models.user.Job.id))
        .filter(
            models.user.Job.uploaded_by_contractor == False,
            models.user.Job.job_review_status == "posted",
            models.user.Job.review_posted_at >= fourteen_days_ago,
            models.user.Job.review_posted_at < seven_days_ago,
        )
        .scalar()
        or 0
    )

    # Posted Last 30 Days
    thirty_days_ago = today_start - timedelta(days=30)
    posted_last_30_days = (
        db.query(func.count(models.user.Job.id))
        .filter(
            models.user.Job.uploaded_by_contractor == False,
            models.user.Job.job_review_status == "posted",
            models.user.Job.review_posted_at >= thirty_days_ago,
        )
        .scalar()
        or 0
    )

    # Previous 30 days (for comparison)
    sixty_days_ago = today_start - timedelta(days=60)
    posted_prev_30_days = (
        db.query(func.count(models.user.Job.id))
        .filter(
            models.user.Job.uploaded_by_contractor == False,
            models.user.Job.job_review_status == "posted",
            models.user.Job.review_posted_at >= sixty_days_ago,
            models.user.Job.review_posted_at < thirty_days_ago,
        )
        .scalar()
        or 0
    )

    # Average Posted Per Day (last 30 days)
    avg_posted_per_day = round(posted_last_30_days / 30, 1)
    prev_avg_posted_per_day = round(posted_prev_30_days / 30, 1)

    # Total Value of Posted Jobs (sum of project costs)
    total_value = (
        db.query(func.sum(models.user.Job.project_cost_total))
        .filter(
            models.user.Job.uploaded_by_contractor == False,
            models.user.Job.job_review_status == "posted",
            models.user.Job.project_cost_total.isnot(None),
        )
        .scalar()
        or 0
    )

    # Previous total value (jobs posted before 7 days ago)
    previous_total_value = (
        db.query(func.sum(models.user.Job.project_cost_total))
        .filter(
            models.user.Job.uploaded_by_contractor == False,
            models.user.Job.job_review_status == "posted",
            models.user.Job.project_cost_total.isnot(None),
            models.user.Job.review_posted_at < seven_days_ago,
        )
        .scalar()
        or 0
    )

    # ========================================================================
    # Apply Filters to Base Query
    # ========================================================================

    # Base query: ONLY posted jobs
    base_query = db.query(models.user.Job).filter(
        models.user.Job.uploaded_by_contractor == False,
        models.user.Job.job_review_status == "posted",
    )

    # Date Filters
    if quick_date:
        if quick_date == "last_7_days":
            base_query = base_query.filter(
                models.user.Job.review_posted_at >= seven_days_ago
            )
        elif quick_date == "last_30_days":
            base_query = base_query.filter(
                models.user.Job.review_posted_at >= thirty_days_ago
            )
        elif quick_date == "last_90_days":
            ninety_days_ago = today_start - timedelta(days=90)
            base_query = base_query.filter(
                models.user.Job.review_posted_at >= ninety_days_ago
            )

    if date_from:
        try:
            from_date = datetime.strptime(date_from, "%Y-%m-%d")
            base_query = base_query.filter(models.user.Job.review_posted_at >= from_date)
        except ValueError:
            pass  # Invalid date format, skip filter

    if date_to:
        try:
            to_date = datetime.strptime(date_to, "%Y-%m-%d").replace(
                hour=23, minute=59, second=59
            )
            base_query = base_query.filter(models.user.Job.review_posted_at <= to_date)
        except ValueError:
            pass  # Invalid date format, skip filter

    # Permit Filters
    if permit_status:
        base_query = base_query.filter(
            func.lower(models.user.Job.permit_status).like(f"%{permit_status.lower()}%")
        )

    if permit_type:
        base_query = base_query.filter(
            func.lower(models.user.Job.permit_type_norm).like(
                f"%{permit_type.lower()}%"
            )
        )

    # Location Filters
    if source_county:
        base_query = base_query.filter(
            func.lower(models.user.Job.source_county).like(
                f"%{source_county.lower()}%"
            )
        )

    if state:
        base_query = base_query.filter(
            func.lower(models.user.Job.state).like(f"%{state.lower()}%")
        )

    # Cost Range Filters
    if cost_min is not None:
        base_query = base_query.filter(
            models.user.Job.project_cost_total >= cost_min
        )

    if cost_max is not None:
        base_query = base_query.filter(
            models.user.Job.project_cost_total <= cost_max
        )

    # Audience Type Filter
    if audience_type:
        base_query = base_query.filter(
            func.lower(models.user.Job.audience_type_names).like(
                f"%{audience_type.lower()}%"
            )
        )

    # Contractor Name Search
    if contractor_name:
        base_query = base_query.filter(
            or_(
                func.lower(models.user.Job.contractor_name).like(
                    f"%{contractor_name.lower()}%"
                ),
                func.lower(models.user.Job.contractor_company).like(
                    f"%{contractor_name.lower()}%"
                ),
            )
        )

    # ========================================================================
    # Pagination and Results
    # ========================================================================

    # Get total count after filters
    total_count = base_query.count()
    total_pages = (total_count + per_page - 1) // per_page

    # Order by review_posted_at descending (most recent first)
    offset = (page - 1) * per_page
    rows = (
        base_query.order_by(models.user.Job.review_posted_at.desc())
        .limit(per_page)
        .offset(offset)
        .all()
    )

    jobs_data = []
    for j in rows:
        jobs_data.append(
            {
                "id": j.id,
                "permit_type": j.audience_type_names,
                "contact_name": j.contact_name,
                "contractor_email": j.contractor_email,
                "trs_score": j.trs_score,
                "job_review_status": j.job_review_status,
            }
        )

    # ========================================================================
    # Return KPIs + Jobs Data
    # ========================================================================
    return {
        "kpis": {
            "total_posted_jobs": {
                "value": total_posted_jobs,
                "change": 0,  # Total is cumulative
            },
            "posted_today": {
                "value": posted_today,
                "change": calc_percentage_change(posted_today, posted_yesterday),
            },
            "posted_last_7_days": {
                "value": posted_last_7_days,
                "change": calc_percentage_change(posted_last_7_days, posted_prev_7_days),
            },
            "posted_last_30_days": {
                "value": posted_last_30_days,
                "change": calc_percentage_change(
                    posted_last_30_days, posted_prev_30_days
                ),
            },
            "avg_posted_per_day": {
                "value": avg_posted_per_day,
                "change": calc_percentage_change(
                    avg_posted_per_day, prev_avg_posted_per_day
                ),
            },
            "total_value": {
                "value": round(total_value, 2),
                "change": calc_percentage_change(total_value, previous_total_value),
            },
        },
        "jobs": jobs_data,
        "pagination": {
            "page": page,
            "per_page": per_page,
            "total_items": total_count,
            "total_pages": total_pages,
        },
    }


@router.delete(
    "/ingested-jobs/{job_id:int}",
    dependencies=[Depends(require_admin_or_editor)],
)
def delete_ingested_job(job_id: int, db: Session = Depends(get_db)):
    """Admin-only: permanently delete an ingested job by id."""
    j = db.query(models.user.Job).filter(models.user.Job.id == job_id).first()
    if not j:
        raise HTTPException(status_code=404, detail="Job not found")

    db.delete(j)
    db.commit()

    return {"job_id": job_id, "deleted": True, "message": "Job deleted."}


@router.patch(
    "/ingested-jobs/{job_id:int}",
    dependencies=[Depends(require_admin_token)],
    summary="Update Job Data (Admin)",
)
def update_job(
    job_id: int,
    job_update: JobUpdate = Body(...),
    db: Session = Depends(get_db),
):
    """Admin endpoint: Update any field of a job except the job ID.
    
    All fields are optional - only provided fields will be updated.
    Returns the updated job data.
    """
    # Find the job
    job = db.query(models.user.Job).filter(models.user.Job.id == job_id).first()
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")
    
    # Update only fields that are provided (not None)
    update_data = job_update.model_dump(exclude_unset=True)
    
    for field, value in update_data.items():
        if hasattr(job, field):
            setattr(job, field, value)
    
    # Automatically update the updated_at timestamp
    job.updated_at = datetime.utcnow()
    
    db.add(job)
    db.commit()
    db.refresh(job)
    
    # Return the updated job data
    return {
        "job_id": job.id,
        "message": "Job updated successfully",
        "updated_fields": list(update_data.keys()),
        "job": {
            # Basic job info
            "id": job.id,
            "created_at": job.created_at.isoformat() if job.created_at else None,
            "updated_at": job.updated_at.isoformat() if job.updated_at else None,
            # Queue and routing info
            "queue_id": job.queue_id,
            "rule_id": job.rule_id,
            "recipient_group": job.recipient_group,
            "recipient_group_id": job.recipient_group_id,
            "day_offset": job.day_offset,
            "anchor_event": job.anchor_event,
            "anchor_at": job.anchor_at.isoformat() if job.anchor_at else None,
            "due_at": job.due_at.isoformat() if job.due_at else None,
            "routing_anchor_at": (
                job.routing_anchor_at.isoformat() if job.routing_anchor_at else None
            ),
            # Permit info
            "permit_id": job.permit_id,
            "permit_number": job.permit_number,
            "permit_status": job.permit_status,
            "permit_type_norm": job.permit_type_norm,
            "permit_raw": job.permit_raw,
            # Project info
            "project_number": job.project_number,
            "project_description": job.project_description,
            "project_type": job.project_type,
            "project_sub_type": job.project_sub_type,
            "project_status": job.project_status,
            "project_cost_total": job.project_cost_total,
            "project_cost": job.project_cost,
            "project_cost_source": job.project_cost_source,
            "property_type": job.property_type,
            # Address info
            "job_address": job.job_address,
            "project_address": job.project_address,
            "state": job.state,
            # Source info
            "source_county": job.source_county,
            "source_system": job.source_system,
            "first_seen_at": (
                job.first_seen_at.isoformat() if job.first_seen_at else None
            ),
            "last_seen_at": job.last_seen_at.isoformat() if job.last_seen_at else None,
            # Contractor info
            "contractor_name": job.contractor_name,
            "contractor_company": job.contractor_company,
            "contractor_email": job.contractor_email,
            "contractor_phone": job.contractor_phone,
            "contractor_company_and_address": job.contractor_company_and_address,
            # Owner/Applicant info
            "owner_name": job.owner_name,
            "applicant_name": job.applicant_name,
            "applicant_email": job.applicant_email,
            "applicant_phone": job.applicant_phone,
            # Audience info
            "audience_type_slugs": job.audience_type_slugs,
            "audience_type_names": job.audience_type_names,
            # Additional info
            "querystring": job.querystring,
            "trs_score": job.trs_score,
            "uploaded_by_contractor": job.uploaded_by_contractor,
            "uploaded_by_user_id": job.uploaded_by_user_id,
            "job_review_status": job.job_review_status,
            "review_posted_at": (
                job.review_posted_at.isoformat() if job.review_posted_at else None
            ),
            "job_group_id": job.job_group_id,
            "job_documents": job.job_documents,
            "contact_name": job.contact_name,
        },
    }


@router.get("/suppliers-kpis", dependencies=[Depends(require_admin_token)])
def suppliers_kpis(db: Session = Depends(get_db)):
    """
    Admin endpoint: Suppliers KPIs only.

    Returns all KPI metrics with percentage changes for suppliers.
    Note: Suppliers have product types (user_type) instead of trade categories.
    """
    from datetime import datetime, timedelta

    # Calculate KPIs for Suppliers
    # Active Subscriptions
    active_subs = (
        db.query(func.count(models.user.Subscriber.id))
        .join(models.user.User, models.user.Subscriber.user_id == models.user.User.id)
        .filter(
            models.user.Subscriber.is_active == True,
            models.user.User.role == "Supplier",
            models.user.User.approved_by_admin == "approved",
        )
        .scalar()
        or 0
    )

    # Past Due
    past_due = (
        db.query(func.count(models.user.Subscriber.id))
        .join(models.user.User, models.user.Subscriber.user_id == models.user.User.id)
        .filter(
            models.user.Subscriber.subscription_status == "past_due",
            models.user.User.role == "Supplier",
            models.user.User.approved_by_admin == "approved",
        )
        .scalar()
        or 0
    )

    # Credits Outstanding (current credits)
    credits_outstanding = (
        db.query(func.coalesce(func.sum(models.user.Subscriber.current_credits), 0))
        .join(models.user.User, models.user.Subscriber.user_id == models.user.User.id)
        .filter(
            models.user.User.role == "Supplier",
            models.user.User.approved_by_admin == "approved",
        )
        .scalar()
        or 0
    )

    # Unlocks Last 7 Days (leads unlocked by suppliers)
    seven_days_ago = datetime.utcnow() - timedelta(days=7)
    unlocks_last_7d = (
        db.query(func.count(models.user.UnlockedLead.id))
        .join(models.user.User, models.user.UnlockedLead.user_id == models.user.User.id)
        .filter(
            models.user.UnlockedLead.unlocked_at >= seven_days_ago,
            models.user.User.role == "Supplier",
            models.user.User.approved_by_admin == "approved",
        )
        .scalar()
        or 0
    )

    # Credits Purchased - Total credits ever acquired (current + spent + frozen)
    credits_purchased_result = (
        db.query(
            func.coalesce(
                func.sum(
                    models.user.Subscriber.current_credits
                    + models.user.Subscriber.total_spending
                    + models.user.Subscriber.frozen_credits
                ),
                0,
            )
        )
        .join(models.user.User, models.user.Subscriber.user_id == models.user.User.id)
        .filter(
            models.user.User.role == "Supplier",
            models.user.User.approved_by_admin == "approved",
        )
        .scalar()
        or 0
    )

    credits_purchased = int(credits_purchased_result) if credits_purchased_result else 0

    # Credits Spent - Total credits spent on unlocking leads
    credits_spent = (
        db.query(func.coalesce(func.sum(models.user.Subscriber.total_spending), 0))
        .join(models.user.User, models.user.Subscriber.user_id == models.user.User.id)
        .filter(
            models.user.User.role == "Supplier",
            models.user.User.approved_by_admin == "approved",
        )
        .scalar()
        or 0
    )

    # Trial Credits Used - Total trial credits consumed by all suppliers
    trial_credits_used_sum = (
        db.query(func.coalesce(func.sum(25 - models.user.Subscriber.trial_credits), 0))
        .join(models.user.User, models.user.Subscriber.user_id == models.user.User.id)
        .filter(
            models.user.Subscriber.trial_credits_used == True,
            models.user.User.role == "Supplier",
            models.user.User.approved_by_admin == "approved",
        )
        .scalar()
        or 0
    )

    trial_credits_used_count = (
        int(trial_credits_used_sum) if trial_credits_used_sum else 0
    )

    # Leads Ingested Today (jobs created today)
    today_start = datetime.utcnow().replace(hour=0, minute=0, second=0, microsecond=0)
    leads_ingested_today = (
        db.query(func.count(models.user.Job.id))
        .filter(models.user.Job.created_at >= today_start)
        .scalar()
        or 0
    )

    # Leads Delivered (total posted jobs)
    leads_delivered = (
        db.query(func.count(models.user.Job.id))
        .filter(models.user.Job.job_review_status == "posted")
        .scalar()
        or 0
    )

    # Leads Unlocked (total unlocks by suppliers)
    leads_unlocked = (
        db.query(func.count(models.user.UnlockedLead.id))
        .join(models.user.User, models.user.UnlockedLead.user_id == models.user.User.id)
        .filter(
            models.user.User.role == "Supplier",
            models.user.User.approved_by_admin == "approved",
        )
        .scalar()
        or 0
    )

    # Calculate previous period values for percentage changes
    fourteen_days_ago = datetime.utcnow() - timedelta(days=14)

    # Previous Unlocks (7-14 days ago)
    unlocks_prev_7d = (
        db.query(func.count(models.user.UnlockedLead.id))
        .join(models.user.User, models.user.UnlockedLead.user_id == models.user.User.id)
        .filter(
            models.user.UnlockedLead.unlocked_at >= fourteen_days_ago,
            models.user.UnlockedLead.unlocked_at < seven_days_ago,
            models.user.User.role == "Supplier",
            models.user.User.approved_by_admin == "approved",
        )
        .scalar()
        or 0
    )

    # Previous Leads Ingested (yesterday)
    yesterday_start = today_start - timedelta(days=1)
    leads_ingested_yesterday = (
        db.query(func.count(models.user.Job.id))
        .filter(
            models.user.Job.created_at >= yesterday_start,
            models.user.Job.created_at < today_start,
        )
        .scalar()
        or 0
    )

    # Helper function to calculate percentage change
    def calc_percentage_change(current, previous):
        if previous == 0:
            return 0 if current == 0 else 100
        return round(((current - previous) / previous) * 100, 1)

    return {
        "activeSubscriptions": {"value": active_subs, "change": 0},
        "pastDue": {"value": past_due, "change": 0},
        "creditsOutstanding": {"value": credits_outstanding, "change": 0},
        "unlocksLast7d": {
            "value": unlocks_last_7d,
            "change": calc_percentage_change(unlocks_last_7d, unlocks_prev_7d),
        },
        "creditsPurchased": {"value": credits_purchased, "change": 0},
        "creditsSpent": {"value": credits_spent, "change": 0},
        "trialCreditsUsed": {"value": trial_credits_used_count, "change": 0},
        "leadsIngestedToday": {
            "value": leads_ingested_today,
            "change": calc_percentage_change(
                leads_ingested_today, leads_ingested_yesterday
            ),
        },
        "leadsDelivered": {"value": leads_delivered, "change": 0},
        "leadsUnlocked": {"value": leads_unlocked, "change": 0},
    }


@router.get("/suppliers-summary", dependencies=[Depends(require_admin_token)])
def suppliers_summary(
    # Account Status filters
    account_status: Optional[str] = Query(
        None, description="Account status: active, disabled"
    ),
    # Subscription Status filters
    subscription_status: Optional[str] = Query(
        None,
        description="Subscription status: active, past_due, canceled, action_required, paused, trial, trial_expired, inactive, trialing",
    ),
    # Plan Tier filters
    plan_tier: Optional[str] = Query(
        None,
        description="Plan tier: Starter, Professional, Enterprise, Custom, no_subscription",
    ),
    # Credits Balance range
    credits_min: Optional[int] = Query(None, description="Minimum credits balance"),
    credits_max: Optional[int] = Query(None, description="Maximum credits balance"),
    # Unlocks Last 7 Days
    unlocks_range: Optional[str] = Query(
        None,
        description="Unlocks range: no_unlocks, low_1_5, medium_6_10, moderate_11_20, heavy_20_plus",
    ),
    # Product Type (user_type for suppliers)
    product_type: Optional[str] = Query(
        None, description="Filter by product type/user type"
    ),
    # Service Area States
    service_state: Optional[str] = Query(None, description="Filter by service state"),
    # Registration Date
    registration_date: Optional[str] = Query(
        None,
        description="Registration date: last_7_days, last_30_days, last_90_days, this_year",
    ),
    # Active Date (last login or activity)
    active_date: Optional[str] = Query(
        None, description="Active date: last_7_days, last_30_days, inactive_90_plus"
    ),
    # Pagination
    page: int = Query(1, ge=1, description="Page number"),
    per_page: int = Query(10, ge=1, le=100, description="Items per page"),
    # Search
    search: Optional[str] = Query(None, description="Search by company, email, phone"),
    db: Session = Depends(get_db),
):
    """
    Admin endpoint: Suppliers table data with filters and pagination.

    Returns table data with applied filters.
    Note: Suppliers have product_type (user_type) instead of trade_category.
    """
    from datetime import datetime, timedelta

    # Base query for suppliers with users and subscribers
    base_query = (
        db.query(
            models.user.Supplier.id,
            models.user.Supplier.company_name,
            models.user.Supplier.phone_number,
            models.user.Supplier.user_type,
            models.user.Supplier.service_states,
            models.user.User.email,
            models.user.User.is_active,
            models.user.User.created_at,
            models.user.User.approved_by_admin,
            models.user.Subscriber.subscription_id,
            models.user.Subscriber.current_credits,
            models.user.Subscriber.subscription_status,
            models.user.Subscriber.subscription_renew_date,
            models.user.Subscriber.total_spending,
            models.user.Subscription.name.label("plan_name"),
        )
        .join(models.user.User, models.user.User.id == models.user.Supplier.user_id)
        .outerjoin(
            models.user.Subscriber,
            models.user.Subscriber.user_id == models.user.User.id,
        )
        .outerjoin(
            models.user.Subscription,
            models.user.Subscription.id == models.user.Subscriber.subscription_id,
        )
        .filter(
            models.user.User.approved_by_admin == "approved",
            models.user.User.role == "Supplier",
        )
    )

    # Apply filters
    # Account Status
    if account_status:
        if account_status.lower() == "active":
            base_query = base_query.filter(models.user.User.is_active == True)
        elif account_status.lower() == "disabled":
            base_query = base_query.filter(models.user.User.is_active == False)

    # Subscription Status
    if subscription_status:
        base_query = base_query.filter(
            models.user.Subscriber.subscription_status == subscription_status.lower()
        )

    # Plan Tier
    if plan_tier:
        if plan_tier.lower() == "no_subscription":
            base_query = base_query.filter(
                models.user.Subscriber.subscription_id.is_(None)
            )
        else:
            base_query = base_query.filter(
                func.lower(models.user.Subscription.name) == plan_tier.lower()
            )

    # Credits Balance range
    if credits_min is not None:
        base_query = base_query.filter(
            models.user.Subscriber.current_credits >= credits_min
        )
    if credits_max is not None:
        base_query = base_query.filter(
            models.user.Subscriber.current_credits <= credits_max
        )

    # Product Type (user_type for suppliers)
    if product_type:
        base_query = base_query.filter(models.user.Supplier.user_type.any(product_type))

    # Service Area States
    if service_state:
        base_query = base_query.filter(
            models.user.Supplier.service_states.any(service_state)
        )

    # Registration Date
    if registration_date:
        now = datetime.utcnow()
        if registration_date == "last_7_days":
            base_query = base_query.filter(
                models.user.User.created_at >= now - timedelta(days=7)
            )
        elif registration_date == "last_30_days":
            base_query = base_query.filter(
                models.user.User.created_at >= now - timedelta(days=30)
            )
        elif registration_date == "last_90_days":
            base_query = base_query.filter(
                models.user.User.created_at >= now - timedelta(days=90)
            )
        elif registration_date == "this_year":
            base_query = base_query.filter(
                func.extract("year", models.user.User.created_at) == now.year
            )

    # Search
    if search:
        search_term = f"%{search.lower()}%"
        base_query = base_query.filter(
            or_(
                func.lower(models.user.Supplier.company_name).like(search_term),
                func.lower(models.user.User.email).like(search_term),
                func.lower(models.user.Supplier.phone_number).like(search_term),
            )
        )

    # Apply Unlocks Last 7 Days filter
    if unlocks_range:
        seven_days_ago = datetime.utcnow() - timedelta(days=7)
        # Subquery to count unlocks per user in last 7 days
        unlocks_subq = (
            db.query(
                models.user.UnlockedLead.user_id,
                func.count(models.user.UnlockedLead.id).label("unlock_count"),
            )
            .filter(models.user.UnlockedLead.unlocked_at >= seven_days_ago)
            .group_by(models.user.UnlockedLead.user_id)
            .subquery()
        )

        base_query = base_query.outerjoin(
            unlocks_subq, unlocks_subq.c.user_id == models.user.User.id
        )

        if unlocks_range == "no_unlocks":
            base_query = base_query.filter(
                or_(
                    unlocks_subq.c.unlock_count.is_(None),
                    unlocks_subq.c.unlock_count == 0,
                )
            )
        elif unlocks_range == "low_1_5":
            base_query = base_query.filter(unlocks_subq.c.unlock_count.between(1, 5))
        elif unlocks_range == "medium_6_10":
            base_query = base_query.filter(unlocks_subq.c.unlock_count.between(6, 10))
        elif unlocks_range == "moderate_11_20":
            base_query = base_query.filter(unlocks_subq.c.unlock_count.between(11, 20))
        elif unlocks_range == "heavy_20_plus":
            base_query = base_query.filter(unlocks_subq.c.unlock_count >= 21)

    # Get total count
    total_count = base_query.count()

    # Apply pagination
    suppliers_data = base_query.offset((page - 1) * per_page).limit(per_page).all()

    # Build table data
    table_data = []
    for supplier in suppliers_data:
        table_data.append(
            {
                "id": supplier.id,
                "company": supplier.company_name,
                "email": supplier.email,
                "phone": supplier.phone_number,
                "planTier": supplier.plan_name or "No Subscription",
                "subscriptionStatus": supplier.subscription_status or "inactive",
                "renewalDate": (
                    supplier.subscription_renew_date.strftime("%m/%d/%y")
                    if supplier.subscription_renew_date
                    else None
                ),
                "creditsBalance": supplier.current_credits or 0,
                "creditsSpent": supplier.total_spending or 0,
                "action": "disable" if supplier.is_active else "enable",
            }
        )

    return {
        "table": table_data,
        "pagination": {
            "total": total_count,
            "page": page,
            "perPage": per_page,
            "totalPages": (total_count + per_page - 1) // per_page,
        },
        "filters": {
            "accountStatus": account_status,
            "subscriptionStatus": subscription_status,
            "planTier": plan_tier,
            "creditsMin": credits_min,
            "creditsMax": credits_max,
            "unlocksRange": unlocks_range,
            "productType": product_type,
            "serviceState": service_state,
            "registrationDate": registration_date,
            "activeDate": active_date,
            "search": search,
        },
    }


@router.get("/suppliers-pending", dependencies=[Depends(require_admin_token)])
def suppliers_pending(
    status: Optional[str] = Query(
        None,
        description="Filter by approval status: pending, rejected (default: shows both, pending first)",
    ),
    db: Session = Depends(get_db),
):
    """Admin endpoint: return list of suppliers pending or rejected approval.

    Query param:
    - status: 'pending', 'rejected', or None/omitted (default: shows both, ordered by pending first)

    Returns suppliers with: id, company, email, phone, license, user_type, created_at, approval_status
    """
    # Base query joining suppliers and users
    base_query = db.query(
        models.user.Supplier.id,
        models.user.Supplier.company_name,
        models.user.Supplier.phone_number,
        models.user.Supplier.state_license_number,
        models.user.Supplier.service_states,
        models.user.Supplier.user_type,
        models.user.User.email,
        models.user.User.is_active,
        models.user.User.approved_by_admin,
        models.user.User.created_at,
        models.user.User.note,
    ).join(models.user.User, models.user.User.id == models.user.Supplier.user_id)

    # Apply status filter
    if status == "pending":
        base_query = base_query.filter(models.user.User.approved_by_admin == "pending")
        rows = base_query.order_by(models.user.User.created_at.desc()).all()
    elif status == "rejected":
        base_query = base_query.filter(models.user.User.approved_by_admin == "rejected")
        rows = base_query.order_by(models.user.User.created_at.desc()).all()
    else:
        # Default: show both pending and rejected, with pending first
        base_query = base_query.filter(
            models.user.User.approved_by_admin.in_(["pending", "rejected"])
        )
        # Order by: pending first (using CASE), then by created_at desc within each group
        rows = base_query.order_by(
            case(
                (models.user.User.approved_by_admin == "pending", 0),
                (models.user.User.approved_by_admin == "rejected", 1),
                else_=2,
            ),
            models.user.User.created_at.desc(),
        ).all()

    result = []
    for row in rows:
        result.append(
            {
                "id": row.id,
                "company": row.company_name,
                "email": row.email,
                "phone": row.phone_number,
                "license_number": row.state_license_number,
                "service_states": row.service_states,
                "user_type": row.user_type,
                "approval_status": row.approved_by_admin,
                "is_active": row.is_active,
                "admin_note": row.note,
                "created_at": (
                    row.created_at.strftime("%m/%d/%y %H:%M")
                    if row.created_at
                    else None
                ),
            }
        )

    return {"suppliers": result}


@router.get(
    "/suppliers/onboarding/{supplier_id}", dependencies=[Depends(require_admin_token)]
)
def supplier_onboarding_detail(supplier_id: int, db: Session = Depends(get_db)):
    """Admin endpoint: Get complete onboarding data for a supplier by supplier_id.

    Returns all data collected during the onboarding process:
    - User account details
    - Business information
    - License information (numbers, dates, status only - no documents)
    - Product types
    - Service areas

    Note: Documents (license_picture, referrals) are uploaded via settings, not onboarding.
    """
    # Find supplier profile
    supplier = (
        db.query(models.user.Supplier)
        .filter(models.user.Supplier.id == supplier_id)
        .first()
    )
    if not supplier:
        raise HTTPException(status_code=404, detail="Supplier not found")

    # Find user
    user = (
        db.query(models.user.User)
        .filter(models.user.User.id == supplier.user_id)
        .first()
    )
    if not user:
        raise HTTPException(status_code=404, detail="User not found")

    return {
        # Step 0: User Account
        "account": {
            "user_id": user.id,
            "email": user.email,
            "email_verified": user.email_verified,
            "created_at": user.created_at.isoformat() if user.created_at else None,
            "is_active": user.is_active,
            "approved_by_admin": user.approved_by_admin,
            "admin_note": user.note,
        },
        # Step 1: Basic Business Information
        "business_info": {
            "company_name": supplier.company_name,
            "phone_number": supplier.phone_number,
            "website_url": supplier.website_url,
            "country_city": supplier.country_city,
        },
        # Step 2: License Information (without documents)
        "license_credentials": {
            "state_license_number": supplier.state_license_number,
            "license_expiration_date": supplier.license_expiration_date,
            "license_status": supplier.license_status,
        },
        # Step 3: Product Type Information
        "product_info": {
            "user_type": supplier.user_type,
            "product_count": len(supplier.user_type) if supplier.user_type else 0,
        },
        # Step 4: Service Jurisdictions
        "service_areas": {
            "service_states": supplier.service_states,
            "country_city": supplier.country_city,
            "states_count": (
                len(supplier.service_states) if supplier.service_states else 0
            ),
            "cities_count": (
                len(supplier.country_city) if supplier.country_city else 0
            ),
        },
        # Onboarding Progress
        "onboarding": {
            "registration_step": supplier.registration_step,
            "is_completed": supplier.is_completed,
            "created_at": (
                supplier.created_at.isoformat() if supplier.created_at else None
            ),
            "updated_at": (
                supplier.updated_at.isoformat() if supplier.updated_at else None
            ),
        },
    }


@router.get("/suppliers/search", dependencies=[Depends(require_admin_token)])
def search_suppliers(q: str, db: Session = Depends(get_db)):
    """Admin endpoint: search suppliers across all columns.

    Query param: `q` - search string to match against any supplier field.
    Returns matching suppliers with basic info matching suppliers-summary format.
    """
    if not q or len(q.strip()) < 2:
        raise HTTPException(
            status_code=400, detail="Search query must be at least 2 characters"
        )

    search_term = f"%{q.lower()}%"

    # Build comprehensive search query across all text columns
    query = text(
        """
        SELECT 
            s.id,
            s.company_name,
            s.primary_contact_name,
            s.phone_number,
            s.state_license_number,
            s.service_states,
            s.user_type,
            u.email,
            u.is_active
        FROM suppliers s
        JOIN users u ON u.id = s.user_id
        WHERE 
            LOWER(COALESCE(s.company_name, '')) LIKE :search
            OR LOWER(COALESCE(s.primary_contact_name, '')) LIKE :search
            OR LOWER(COALESCE(s.phone_number, '')) LIKE :search
            OR LOWER(COALESCE(s.website_url, '')) LIKE :search
            OR LOWER(COALESCE(s.state_license_number, '')) LIKE :search
            OR LOWER(COALESCE(s.business_address, '')) LIKE :search
            OR LOWER(COALESCE(u.email, '')) LIKE :search
            OR LOWER(ARRAY_TO_STRING(s.service_states, ',')) LIKE :search
            OR LOWER(ARRAY_TO_STRING(s.country_city, ',')) LIKE :search
            OR LOWER(ARRAY_TO_STRING(s.user_type, ',')) LIKE :search
        AND u.approved_by_admin = 'approved'
        ORDER BY s.id DESC
        LIMIT 100
    """
    )

    rows = db.execute(query, {"search": search_term}).fetchall()

    result = []
    for row in rows:
        action = "disable" if row.is_active else "enable"
        result.append(
            {
                "id": row.id,
                "phone_number": row.phone_number,
                "email": row.email,
                "company": row.company_name,
                "license_number": row.state_license_number,
                "service_states": row.service_states,
                "user_type": row.user_type,
                "action": action,
            }
        )

    return {"suppliers": result}


@router.get("/suppliers/search-pending", dependencies=[Depends(require_admin_token)])
def search_suppliers_pending(q: str, db: Session = Depends(get_db)):
    """Admin endpoint: search pending suppliers across all columns.

    Query param: `q` - search string to match against any supplier field.
    Returns only suppliers with approved_by_admin = 'pending'.
    """
    if not q or len(q.strip()) < 2:
        raise HTTPException(
            status_code=400, detail="Search query must be at least 2 characters"
        )

    search_term = f"%{q.lower()}%"

    # Build comprehensive search query across all text columns
    query = text(
        """
        SELECT 
            s.id,
            s.company_name,
            s.primary_contact_name,
            s.phone_number,
            s.state_license_number,
            s.service_states,
            s.user_type,
            u.email,
            u.is_active,
            u.created_at
        FROM suppliers s
        JOIN users u ON u.id = s.user_id
        WHERE 
            LOWER(COALESCE(s.company_name, '')) LIKE :search
            OR LOWER(COALESCE(s.primary_contact_name, '')) LIKE :search
            OR LOWER(COALESCE(s.phone_number, '')) LIKE :search
            OR LOWER(COALESCE(s.website_url, '')) LIKE :search
            OR EXISTS (
                SELECT 1 FROM jsonb_array_elements_text(COALESCE(s.state_license_number::jsonb, '[]'::jsonb)) AS license
                WHERE LOWER(license) LIKE :search
            )
            OR LOWER(COALESCE(s.business_address, '')) LIKE :search
            OR LOWER(COALESCE(u.email, '')) LIKE :search
            OR LOWER(ARRAY_TO_STRING(s.service_states, ',')) LIKE :search
            OR LOWER(ARRAY_TO_STRING(s.country_city, ',')) LIKE :search
            OR LOWER(ARRAY_TO_STRING(s.user_type, ',')) LIKE :search
        AND u.approved_by_admin = 'pending'
        ORDER BY s.id DESC
        LIMIT 100
    """
    )

    rows = db.execute(query, {"search": search_term}).fetchall()

    result = []
    for row in rows:
        action = "disable" if row.is_active else "enable"
        result.append(
            {
                "id": row.id,
                "phone_number": row.phone_number,
                "email": row.email,
                "company": row.company_name,
                "license_number": row.state_license_number,
                "service_states": row.service_states,
                "user_type": row.user_type,
                "action": action,
                "created_at": row.created_at.isoformat() if row.created_at else None,
            }
        )

    return {"suppliers": result}


@router.get(
    "/admin-users/recipients",
    dependencies=[Depends(require_admin_or_editor)],
)
def admin_users_recipients(db: Session = Depends(get_db)):
    """Return non-admin admin_users suitable for sending invites/notifications.

    Each entry contains: id, name, email, role, status
    - status is 'active' when is_active is true
    - status is 'invited' when is_active is false
    """
    q = text(
        "SELECT id, COALESCE(name, '') AS name, email, role, is_active FROM admin_users "
        "WHERE LOWER(COALESCE(role, '')) != 'admin' ORDER BY id"
    )
    rows = db.execute(q).fetchall()

    recipients = []
    for r in rows:
        mapping = getattr(r, "_mapping", None)
        if mapping is None:
            try:
                rid = r[0]
                name = r[1]
                email = r[2]
                role = r[3]
                is_active = r[4]
            except Exception:
                continue
        else:
            rid = mapping.get("id")
            name = mapping.get("name")
            email = mapping.get("email")
            role = mapping.get("role")
            is_active = mapping.get("is_active")

        status = "active" if is_active else "invited"
        recipients.append(
            {"id": rid, "name": name, "email": email, "role": role, "status": status}
        )

    return {"recipients": recipients}


@router.get(
    "/admin-users/by-role",
    dependencies=[Depends(require_admin_or_editor)],
)
def admin_users_by_role(role: str, db: Session = Depends(get_db)):
    """Return admin users matching the given `role` (excluding 'admin').

    Query param: `role` (string). Returns same shape as `admin_users_list`.
    """
    if not role:
        raise HTTPException(status_code=400, detail="Missing role parameter")
    if role.lower() == "admin":
        # Explicitly disallow listing real admin role via this filtered endpoint
        raise HTTPException(
            status_code=400, detail="Filtering for role 'admin' is not allowed"
        )

    q = text(
        "SELECT id, COALESCE(name, '') AS name, email, role, is_active FROM admin_users "
        "WHERE lower(role) = lower(:role) AND LOWER(COALESCE(role, '')) != 'admin' ORDER BY id"
    )
    rows = db.execute(q, {"role": role}).fetchall()

    result = []
    for r in rows:
        mapping = getattr(r, "_mapping", None)
        if mapping is None:
            try:
                rid = r[0]
                name = r[1]
                email = r[2]
                role_val = r[3]
                is_active = r[4]
            except Exception:
                continue
        else:
            rid = mapping.get("id")
            name = mapping.get("name")
            email = mapping.get("email")
            role_val = mapping.get("role")
            is_active = mapping.get("is_active")
        status = "active" if is_active else "inactive"
        result.append(
            {
                "id": rid,
                "name": name,
                "email": email,
                "role": role_val,
                "status": status,
            }
        )

    return {"admin_users": result}


@router.get(
    "/admin-users/search",
    dependencies=[Depends(require_admin_or_editor)],
)
def admin_users_search(q: str, db: Session = Depends(get_db)):
    """Search admin_users by name or email (case-insensitive), excluding role 'admin'.

    Query param: `q` - substring to match against name/email (case-insensitive).
    Returns list of {id, name, email, role, status} where status is 'active'|'inactive'.
    """
    if not q:
        raise HTTPException(status_code=400, detail="Missing query parameter 'q'")

    like = f"%{q.lower()}%"
    # Use raw SQL for compatibility with optional columns
    q_text = text(
        "SELECT id, COALESCE(name, '') AS name, email, role, is_active FROM admin_users "
        "WHERE LOWER(COALESCE(role, '')) != 'admin' AND (LOWER(COALESCE(name, '')) LIKE :like OR LOWER(COALESCE(email, '')) LIKE :like) ORDER BY id"
    )
    rows = db.execute(q_text, {"like": like}).fetchall()

    result = []
    for r in rows:
        mapping = getattr(r, "_mapping", None)
        if mapping is None:
            try:
                rid = r[0]
                name = r[1]
                email = r[2]
                role_val = r[3]
                is_active = r[4]
            except Exception:
                continue
        else:
            rid = mapping.get("id")
            name = mapping.get("name")
            email = mapping.get("email")
            role_val = mapping.get("role")
            is_active = mapping.get("is_active")
        status = "active" if is_active else "inactive"
        result.append(
            {
                "id": rid,
                "name": name,
                "email": email,
                "role": role_val,
                "status": status,
            }
        )

    return {"admin_users": result}


@router.patch(
    "/admin-users/{admin_id:int}/role",
    dependencies=[Depends(require_admin_or_editor)],
)
def update_admin_user_role(
    admin_id: int,
    role: str = Body(..., embed=True),
    db: Session = Depends(get_db),
):
    """Update the role of an admin_user.

    Only callers with role 'admin' or 'editor' may perform this action.

    Request body: { "role": "editor" }
    """
    if not role or not role.strip():
        raise HTTPException(status_code=400, detail="Role cannot be empty")

    # Verify admin_user exists
    q = text("SELECT id, role FROM admin_users WHERE id = :id")
    row = db.execute(q, {"id": admin_id}).first()
    if not row:
        raise HTTPException(status_code=404, detail="Admin user not found")

    # Update the role
    try:
        update_q = text("UPDATE admin_users SET role = :role WHERE id = :id")
        db.execute(update_q, {"role": role.strip(), "id": admin_id})
        db.commit()
    except Exception as e:
        logger.exception("Failed to update admin user role: %s", str(e))
        raise HTTPException(
            status_code=500, detail=f"Failed to update admin user role: {e}"
        )

    return {
        "admin_id": admin_id,
        "role": role.strip(),
        "message": f"Admin user role updated to '{role.strip()}'",
    }


@router.delete(
    "/admin-users/{admin_id:int}",
    dependencies=[Depends(require_admin_only)],
)
def delete_admin_user(admin_id: int, db: Session = Depends(get_db)):
    """Admin-only: delete an admin_user by id.

    Only callers with role 'admin' may perform this action. Returns 404 if the
    admin_user id does not exist.
    """
    # Verify existence
    q = text("SELECT id FROM admin_users WHERE id = :id")
    row = db.execute(q, {"id": admin_id}).first()
    if not row:
        raise HTTPException(status_code=404, detail="Admin user not found")

    # Perform delete
    try:
        db.execute(text("DELETE FROM admin_users WHERE id = :id"), {"id": admin_id})
        db.commit()
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to delete admin user: {e}")

    return {"admin_id": admin_id, "deleted": True, "message": "Admin user deleted."}


@router.post(
    "/admin-users/invite",
    summary="Invite Admin User",
)
async def invite_admin_user(
    payload: AdminInvite,
    db: Session = Depends(get_db),
    inviter: object = Depends(require_admin_or_editor),
):
    """Invite a new admin user. Only callers with role 'admin' or 'editor'.

    Stores a pending admin_users row (is_active=false) and sends an email invite
    with a signup link to `https://tigerleads.vercel.app/admin/signup?invite_token=...`.
    """
    # Generate invitation token and expiry
    # Trim token to fit existing DB column (VARCHAR(10)) to avoid insertion errors
    raw_token = uuid.uuid4().hex
    token = raw_token[:10]
    expires = datetime.utcnow() + timedelta(days=7)

    logger.info(
        "Admin invite requested: email=%s role=%s by_inviter=%s",
        payload.email,
        payload.role,
        getattr(inviter, "email", None),
    )
    # Insert into admin_users (idempotent: skip if email exists)
    try:
        existing = db.execute(
            text("SELECT id FROM admin_users WHERE lower(email)=lower(:email)"),
            {"email": payload.email},
        ).first()
        if existing:
            raise HTTPException(
                status_code=409, detail="Admin user with this email already exists"
            )

        insert_q = text(
            "INSERT INTO admin_users (email, name, role, is_active, verification_code, code_expires_at, created_by, created_at) "
            "VALUES (:email, :name, :role, :is_active, :code, :expires, :created_by, :created_at)"
        )
        # Resolve `created_by` to a users.id if possible (admin.inviter may be an admin_users id)
        inviter_email = getattr(inviter, "email", None)
        created_by_user_id = None
        if inviter_email:
            row = db.execute(
                text("SELECT id FROM users WHERE lower(email)=lower(:email)"),
                {"email": inviter_email},
            ).first()
            if row:
                created_by_user_id = row.id
        logger.debug(
            "Resolved created_by_user_id=%s for inviter_email=%s",
            created_by_user_id,
            inviter_email,
        )

        params = {
            "email": payload.email,
            "name": payload.name,
            "role": payload.role,
            "is_active": False,
            "code": token,
            "expires": expires,
            "created_by": created_by_user_id,
            "created_at": datetime.utcnow(),
        }
        logger.debug("Inserting admin_users row with params: %s", params)
        db.execute(insert_q, params)
        db.commit()
    except HTTPException:
        raise
    except Exception as e:
        logger.exception(
            "Failed to create admin invite for email=%s; params=%s",
            payload.email,
            {
                "email": payload.email,
                "name": payload.name,
                "role": payload.role,
                "created_by": created_by_user_id,
            },
        )
        raise HTTPException(
            status_code=500, detail=f"Failed to create admin invite: {e}"
        )

    # Schedule sending the invitation email (fire-and-forget)
    signup_url = "https://tigerleads.vercel.app/admin/signup"
    inviter_name = getattr(inviter, "email", "Administrator")
    try:
        asyncio.create_task(
            send_admin_invitation_email(
                payload.email, inviter_name, payload.role, signup_url, token
            )
        )
    except Exception:
        # If task scheduling fails, don't block invite creation; log and continue
        pass

    return {"email": payload.email, "invited": True}


@router.get("/contractors/{contractor_id:int}", dependencies=[Depends(require_admin_token)])
def contractor_detail(contractor_id: int, db: Session = Depends(get_db)):
    """Admin endpoint: return full contractor profile with complete information and embedded file data.

    Returns:
    - All contractor business information
    - User account details (email, status, approval)
    - Subscription details (plan, credits, status)
    - Activity stats (unlocks last 30 days, all time)
    - Complete file data embedded as base64 (license_picture, referrals, job_photos)
    - Service area counts
    """
    from datetime import datetime, timedelta

    c = (
        db.query(models.user.Contractor)
        .filter(models.user.Contractor.id == contractor_id)
        .first()
    )
    if not c:
        raise HTTPException(status_code=404, detail="Contractor not found")

    # Get associated user
    user = db.query(models.user.User).filter(models.user.User.id == c.user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="Associated user not found")

    # Get subscriber details
    subscriber = (
        db.query(models.user.Subscriber)
        .filter(models.user.Subscriber.user_id == user.id)
        .first()
    )

    # Get subscription plan name and stripe subscription ID
    subscription_plan_name = None
    stripe_subscription_id = None
    if subscriber:
        stripe_subscription_id = subscriber.stripe_subscription_id
        if subscriber.subscription_id:
            subscription = (
                db.query(models.user.Subscription)
                .filter(models.user.Subscription.id == subscriber.subscription_id)
                .first()
            )
            if subscription:
                subscription_plan_name = subscription.name

    # Calculate unlock statistics
    thirty_days_ago = datetime.utcnow() - timedelta(days=30)

    # Unlocks last 30 days
    unlocks_30_days = (
        db.query(func.count(models.user.UnlockedLead.id))
        .filter(
            models.user.UnlockedLead.user_id == user.id,
            models.user.UnlockedLead.unlocked_at >= thirty_days_ago,
        )
        .scalar()
        or 0
    )

    # Unlocks all time
    unlocks_all_time = (
        db.query(func.count(models.user.UnlockedLead.id))
        .filter(models.user.UnlockedLead.user_id == user.id)
        .scalar()
        or 0
    )

    def process_files_with_data(json_array, field_name):
        """Return complete file data with embedded base64 and metadata"""
        if not json_array or not isinstance(json_array, list):
            return []

        result = []
        for idx, file_obj in enumerate(json_array):
            file_entry = {
                "filename": file_obj.get("filename"),
                "content_type": file_obj.get("content_type"),
                "file_index": idx,
                "url": f"/admin/dashboard/contractors/{contractor_id}/image/{field_name}?file_index={idx}",
                "data": file_obj.get("data"),  # Base64 encoded data
            }
            result.append(file_entry)
        return result

    # Process all file fields with embedded data
    license_picture = process_files_with_data(c.license_picture, "license_picture")
    referrals = process_files_with_data(c.referrals, "referrals")
    job_photos = process_files_with_data(c.job_photos, "job_photos")

    # Calculate service area counts
    trade_categories_count = len(c.user_type) if c.user_type else 0
    service_states_count = (
        len(c.service_states) if c.service_states else (len(c.state) if c.state else 0)
    )
    cities_counties_count = len(c.country_city) if c.country_city else 0

    return {
        # Basic Profile
        "id": c.id,
        "name": c.primary_contact_name,
        "email": user.email,
        "phone": c.phone_number,
        "account_status": "Active" if user.is_active else "Disabled",
        "approval_status": (
            user.approved_by_admin.capitalize() if user.approved_by_admin else "Pending"
        ),
        "registered_on": (
            user.created_at.strftime("%m/%d/%y") if user.created_at else None
        ),
        # Subscription Details
        "subscription": {
            "plan": subscription_plan_name or "No Subscription",
            "status": (
                subscriber.subscription_status.replace("_", " ").title()
                if subscriber and subscriber.subscription_status
                else "Inactive"
            ),
            "stripe_id": user.stripe_customer_id,
            "subscription_id": stripe_subscription_id,
            "start_date": (
                subscriber.subscription_start_date.strftime("%m/%d/%y")
                if subscriber and subscriber.subscription_start_date
                else None
            ),
            "auto_renewal": "Yes" if subscriber and subscriber.auto_renew else "No",
        },
        # Credits & Activity
        "credits_activity": {
            "current_credits": subscriber.current_credits if subscriber else 0,
            "total_spent": subscriber.total_spending if subscriber else 0,
            "trial_credits": subscriber.trial_credits if subscriber else 0,
            "unlock_30_days": unlocks_30_days,
            "unlock_all_time": unlocks_all_time,
        },
        # Business Information
        "business": {
            "company_name": c.company_name,
            "contact_name": c.primary_contact_name,
            "phone": c.phone_number,
            "email": user.email,
            "website": c.business_website_url or c.website_url,
            "business_address": c.business_address,
        },
        # License & Credentials
        "license": {
            "license_number": c.state_license_number,
            "license_status": c.license_status,
            "expire_date": c.license_expiration_date,
            "documents": {
                "license_picture": license_picture,
                "job_photos": job_photos,
                "referrals": referrals,
            },
        },
        # Service Areas
        "service_areas": {
            "trade_categories": c.user_type,
            "trade_categories_count": trade_categories_count,
            "service_states": c.service_states or c.state,
            "service_states_count": service_states_count,
            "cities_counties": c.country_city,
            "cities_counties_count": cities_counties_count,
        },
        # Additional Details (for extended views)
        "additional": {
            "user_id": user.id,
            "contractor_id": c.id,
            "email_verified": user.email_verified,
            "two_factor_enabled": user.two_factor_enabled,
            "parent_user_id": user.parent_user_id,
            "team_role": user.team_role,
            "invited_by_id": user.invited_by_id,
            "admin_note": user.note,
            "registration_step": c.registration_step,
            "is_completed": c.is_completed,
            "frozen_credits": subscriber.frozen_credits if subscriber else 0,
            "seats_used": subscriber.seats_used if subscriber else 0,
            "purchased_seats": subscriber.purchased_seats if subscriber else 0,
            "created_at": c.created_at.isoformat() if c.created_at else None,
            "updated_at": c.updated_at.isoformat() if c.updated_at else None,
        },
    }


@router.get("/suppliers/{supplier_id:int}", dependencies=[Depends(require_admin_token)])
def supplier_detail(supplier_id: int, db: Session = Depends(get_db)):
    """Admin endpoint: return full supplier profile with complete information and embedded file data.

    Returns:
    - All supplier business information
    - User account details (email, status, approval)
    - Subscription details (plan, credits, status)
    - Activity stats (unlocks last 30 days, all time)
    - Complete file data embedded as base64 (license_picture, referrals, job_photos)
    - Service area counts
    """
    from datetime import datetime, timedelta

    s = (
        db.query(models.user.Supplier)
        .filter(models.user.Supplier.id == supplier_id)
        .first()
    )
    if not s:
        raise HTTPException(status_code=404, detail="Supplier not found")

    # Get associated user
    user = db.query(models.user.User).filter(models.user.User.id == s.user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="Associated user not found")

    # Get subscriber details
    subscriber = (
        db.query(models.user.Subscriber)
        .filter(models.user.Subscriber.user_id == user.id)
        .first()
    )

    # Get subscription plan name and stripe subscription ID
    subscription_plan_name = None
    stripe_subscription_id = None
    if subscriber:
        stripe_subscription_id = subscriber.stripe_subscription_id
        if subscriber.subscription_id:
            subscription = (
                db.query(models.user.Subscription)
                .filter(models.user.Subscription.id == subscriber.subscription_id)
                .first()
            )
            if subscription:
                subscription_plan_name = subscription.name

    # Calculate unlock statistics
    thirty_days_ago = datetime.utcnow() - timedelta(days=30)

    # Unlocks last 30 days
    unlocks_30_days = (
        db.query(func.count(models.user.UnlockedLead.id))
        .filter(
            models.user.UnlockedLead.user_id == user.id,
            models.user.UnlockedLead.unlocked_at >= thirty_days_ago,
        )
        .scalar()
        or 0
    )

    # Unlocks all time
    unlocks_all_time = (
        db.query(func.count(models.user.UnlockedLead.id))
        .filter(models.user.UnlockedLead.user_id == user.id)
        .scalar()
        or 0
    )

    def process_files_with_data(json_array, field_name):
        """Return complete file data with embedded base64 and metadata"""
        if not json_array or not isinstance(json_array, list):
            return []

        result = []
        for idx, file_obj in enumerate(json_array):
            file_entry = {
                "filename": file_obj.get("filename"),
                "content_type": file_obj.get("content_type"),
                "file_index": idx,
                "url": f"/admin/dashboard/suppliers/{supplier_id}/image/{field_name}?file_index={idx}",
                "data": file_obj.get("data"),  # Base64 encoded data
            }
            result.append(file_entry)
        return result

    # Process all file fields with embedded data
    license_picture = process_files_with_data(s.license_picture, "license_picture")
    referrals = process_files_with_data(s.referrals, "referrals")
    job_photos = process_files_with_data(s.job_photos, "job_photos")

    # Calculate service area counts
    product_types_count = len(s.user_type) if s.user_type else 0
    service_states_count = len(s.service_states) if s.service_states else 0
    cities_counties_count = len(s.country_city) if s.country_city else 0

    return {
        # Basic Profile
        "id": s.id,
        "name": s.company_name,
        "email": user.email,
        "phone": s.phone_number,
        "account_status": "Active" if user.is_active else "Disabled",
        "approval_status": (
            user.approved_by_admin.capitalize() if user.approved_by_admin else "Pending"
        ),
        "registered_on": (
            user.created_at.strftime("%m/%d/%y") if user.created_at else None
        ),
        # Subscription Details
        "subscription": {
            "plan": subscription_plan_name or "No Subscription",
            "status": (
                subscriber.subscription_status.replace("_", " ").title()
                if subscriber and subscriber.subscription_status
                else "Inactive"
            ),
            "stripe_id": user.stripe_customer_id,
            "subscription_id": stripe_subscription_id,
            "start_date": (
                subscriber.subscription_start_date.strftime("%m/%d/%y")
                if subscriber and subscriber.subscription_start_date
                else None
            ),
            "auto_renewal": "Yes" if subscriber and subscriber.auto_renew else "No",
        },
        # Credits & Activity
        "credits_activity": {
            "current_credits": subscriber.current_credits if subscriber else 0,
            "total_spent": subscriber.total_spending if subscriber else 0,
            "trial_credits": subscriber.trial_credits if subscriber else 0,
            "unlock_30_days": unlocks_30_days,
            "unlock_all_time": unlocks_all_time,
        },
        # Business Information
        "business": {
            "company_name": s.company_name,
            "phone": s.phone_number,
            "email": user.email,
            "website": s.website_url,
            "business_address": s.business_address,
        },
        # License & Credentials
        "license": {
            "license_number": s.state_license_number,
            "license_status": s.license_status,
            "expire_date": s.license_expiration_date,
            "documents": {
                "license_picture": license_picture,
                "job_photos": job_photos,
                "referrals": referrals,
            },
        },
        # Service Areas
        "service_areas": {
            "product_types": s.user_type,
            "product_types_count": product_types_count,
            "service_states": s.service_states,
            "service_states_count": service_states_count,
            "cities_counties": s.country_city,
            "cities_counties_count": cities_counties_count,
        },
        # Admin Notes
        "admin_note": user.note,
    }


@router.get(
    "/suppliers/{supplier_id:int}/image/{field}",
    dependencies=[Depends(require_admin_token)],
)
def supplier_image(
    supplier_id: int,
    field: str,
    file_index: int = Query(
        0, ge=0, description="Index of file in the array (0-based)"
    ),
    db: Session = Depends(get_db),
):
    """Return binary content for a supplier image/document field.

    `field` must be one of: `license_picture`, `referrals`, `job_photos`.
    `file_index` specifies which file to retrieve from the JSON array (default: 0).

    Files are stored as JSON arrays with base64-encoded data.
    Responds with raw binary and proper Content-Type so frontend can display or open.
    """
    import base64
    import json

    allowed_fields = ["license_picture", "referrals", "job_photos"]
    if field not in allowed_fields:
        raise HTTPException(status_code=400, detail="Invalid image field")

    s = (
        db.query(models.user.Supplier)
        .filter(models.user.Supplier.id == supplier_id)
        .first()
    )
    if not s:
        raise HTTPException(status_code=404, detail="Supplier not found")

    # Get the JSON array for the field
    files_json = getattr(s, field, None)
    if not files_json:
        raise HTTPException(status_code=404, detail=f"{field} not found for supplier")

    # Parse JSON array
    try:
        if isinstance(files_json, str):
            files_array = json.loads(files_json)
        else:
            files_array = files_json
    except (json.JSONDecodeError, TypeError):
        raise HTTPException(
            status_code=500, detail=f"Invalid file data format for {field}"
        )

    # Check if file_index exists
    if not isinstance(files_array, list) or len(files_array) == 0:
        raise HTTPException(status_code=404, detail=f"No files found for {field}")

    if file_index >= len(files_array):
        raise HTTPException(
            status_code=404,
            detail=f"File index {file_index} not found. Only {len(files_array)} file(s) available.",
        )

    # Get the specific file
    file_data = files_array[file_index]

    # Decode base64 data
    try:
        blob = base64.b64decode(file_data.get("data", ""))
        content_type = file_data.get("content_type", "application/octet-stream")
        filename = file_data.get("filename", f"{field}-{supplier_id}-{file_index}")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error decoding file: {str(e)}")

    # Stream raw bytes with correct Content-Type so browsers can render via <img src="...">.
    return Response(
        content=blob,
        media_type=content_type,
        headers={"Content-Disposition": f'inline; filename="{filename}"'},
    )


@router.patch(
    "/suppliers/{supplier_id:int}/approval",
    dependencies=[Depends(require_admin_or_editor)],
)
def update_supplier_approval(
    supplier_id: int, data: ContractorApprovalUpdate, db: Session = Depends(get_db)
):
    """Admin/Editor: Approve or reject a supplier account.

    Updates the `approved_by_admin` field in the users table to "approved" or "rejected".
    Optionally adds a note to the `note` field for admin reference.

    Request body:
    {
        "status": "approved",  // or "rejected"
        "note": "Verified license and credentials"  // optional
    }
    """
    # Validate status
    if data.status not in ["approved", "rejected"]:
        raise HTTPException(
            status_code=400, detail="Status must be either 'approved' or 'rejected'"
        )

    # Find supplier
    s = (
        db.query(models.user.Supplier)
        .filter(models.user.Supplier.id == supplier_id)
        .first()
    )
    if not s:
        raise HTTPException(status_code=404, detail="Supplier not found")

    # Find associated user
    user = db.query(models.user.User).filter(models.user.User.id == s.user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="Associated user not found")

    try:
        # Update approval status
        old_status = user.approved_by_admin
        user.approved_by_admin = data.status

        # Update note if provided
        if data.note is not None:
            user.note = data.note

        db.add(user)
        db.commit()
        db.refresh(user)

        logger.info(
            f"Supplier {supplier_id} (user {user.id}) approval status changed from "
            f"'{old_status}' to '{data.status}'"
        )

        # Create notification for supplier
        notification = models.user.Notification(
            user_id=user.id,
            type="account_approval",
            message=f"Your supplier account has been {data.status} by an administrator.",
        )
        db.add(notification)
        db.commit()

        return {
            "success": True,
            "supplier_id": supplier_id,
            "user_id": user.id,
            "email": user.email,
            "approved_by_admin": data.status,
            "note": user.note,
            "message": f"Supplier account has been {data.status} successfully.",
        }

    except Exception as e:
        db.rollback()
        logger.error(f"Failed to update supplier approval status: {str(e)}")
        raise HTTPException(
            status_code=500, detail=f"Failed to update approval status: {str(e)}"
        )


@router.get(
    "/contractors/{contractor_id:int}/image/{field}",
    dependencies=[Depends(require_admin_token)],
)
def contractor_image(
    contractor_id: int,
    field: str,
    file_index: int = Query(
        0, ge=0, description="Index of file in the array (0-based)"
    ),
    db: Session = Depends(get_db),
):
    """Return binary content for a contractor image/document field.

    `field` must be one of: `license_picture`, `referrals`, `job_photos`.
    `file_index` specifies which file to retrieve from the JSON array (default: 0).

    Files are now stored as JSON arrays with base64-encoded data.
    Responds with raw binary and proper Content-Type so frontend can display or open.
    """
    import base64
    import json

    allowed_fields = ["license_picture", "referrals", "job_photos"]
    if field not in allowed_fields:
        raise HTTPException(status_code=400, detail="Invalid image field")

    c = (
        db.query(models.user.Contractor)
        .filter(models.user.Contractor.id == contractor_id)
        .first()
    )
    if not c:
        raise HTTPException(status_code=404, detail="Contractor not found")

    # Get the JSON array for the field
    files_json = getattr(c, field, None)
    if not files_json:
        raise HTTPException(status_code=404, detail=f"{field} not found for contractor")

    # Parse JSON array
    try:
        if isinstance(files_json, str):
            files_array = json.loads(files_json)
        else:
            files_array = files_json
    except (json.JSONDecodeError, TypeError):
        raise HTTPException(
            status_code=500, detail=f"Invalid file data format for {field}"
        )

    # Check if file_index exists
    if not isinstance(files_array, list) or len(files_array) == 0:
        raise HTTPException(status_code=404, detail=f"No files found for {field}")

    if file_index >= len(files_array):
        raise HTTPException(
            status_code=404,
            detail=f"File index {file_index} not found. Only {len(files_array)} file(s) available.",
        )

    # Get the specific file
    file_data = files_array[file_index]

    # Decode base64 data
    try:
        blob = base64.b64decode(file_data.get("data", ""))
        content_type = file_data.get("content_type", "application/octet-stream")
        filename = file_data.get("filename", f"{field}-{contractor_id}-{file_index}")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error decoding file: {str(e)}")

    # Stream raw bytes with correct Content-Type so browsers can render via <img src="...">.
    return Response(
        content=blob,
        media_type=content_type,
        headers={"Content-Disposition": f'inline; filename="{filename}"'},
    )


# NOTE: The signed public image endpoints were removed. If you need temporary
# public URLs for contractor images, consider implementing cloud storage
# presigned URLs or a revocable token mechanism.


@router.patch(
    "/contractors/{contractor_id:int}/active",
    dependencies=[Depends(require_admin_or_editor)],
)
def set_contractor_active(contractor_id: int, db: Session = Depends(get_db)):
    """Admin-only: toggle the contractor's user `is_active` flag.

    The endpoint requires only the `contractor_id` path parameter. It will
    fetch the associated `users.is_active` value and flip it (true -> false,
    false -> true), commit the change, and return the new state.
    """
    c = (
        db.query(models.user.Contractor)
        .filter(models.user.Contractor.id == contractor_id)
        .first()
    )
    if not c:
        raise HTTPException(status_code=404, detail="Contractor not found")

    user = db.query(models.user.User).filter(models.user.User.id == c.user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="Associated user not found")

    # Flip the active flag
    user.is_active = not bool(user.is_active)
    db.add(user)
    db.commit()

    message = (
        "Contractor account has been enabled."
        if user.is_active
        else "Contractor account has been disabled by an administrator."
    )

    return {"user_id": user.id, "is_active": user.is_active, "message": message}


@router.patch(
    "/contractors/{contractor_id:int}/approval",
    dependencies=[Depends(require_admin_or_editor)],
)
def update_contractor_approval(
    contractor_id: int, data: ContractorApprovalUpdate, db: Session = Depends(get_db)
):
    """Admin/Editor: Approve or reject a contractor account.

    Updates the `approved_by_admin` field in the users table to "approved" or "rejected".
    Optionally adds a note to the `note` field for admin reference.

    Request body:
    {
        "status": "approved",  // or "rejected"
        "note": "Verified license and credentials"  // optional
    }
    """
    # Validate status
    if data.status not in ["approved", "rejected"]:
        raise HTTPException(
            status_code=400, detail="Status must be either 'approved' or 'rejected'"
        )

    # Find contractor
    c = (
        db.query(models.user.Contractor)
        .filter(models.user.Contractor.id == contractor_id)
        .first()
    )
    if not c:
        raise HTTPException(status_code=404, detail="Contractor not found")

    # Find associated user
    user = db.query(models.user.User).filter(models.user.User.id == c.user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="Associated user not found")

    try:
        # Update approval status
        old_status = user.approved_by_admin
        user.approved_by_admin = data.status

        # Update note if provided
        if data.note is not None:
            user.note = data.note

        db.add(user)
        db.commit()
        db.refresh(user)

        logger.info(
            f"Contractor {contractor_id} (user {user.id}) approval status changed from "
            f"'{old_status}' to '{data.status}'"
        )

        # Create notification for contractor
        notification = models.user.Notification(
            user_id=user.id,
            type="account_approval",
            message=f"Your contractor account has been {data.status} by an administrator.",
        )
        db.add(notification)
        db.commit()

        return {
            "success": True,
            "contractor_id": contractor_id,
            "user_id": user.id,
            "email": user.email,
            "approved_by_admin": data.status,
            "note": user.note,
            "message": f"Contractor account has been {data.status} successfully.",
        }

    except Exception as e:
        db.rollback()
        logger.error(f"Failed to update contractor approval status: {str(e)}")
        raise HTTPException(
            status_code=500, detail=f"Failed to update approval status: {str(e)}"
        )


@router.patch(
    "/suppliers/{supplier_id:int}/active",
    dependencies=[Depends(require_admin_or_editor)],
)
def set_supplier_active(supplier_id: int, db: Session = Depends(get_db)):
    """Admin-only: toggle the supplier's user `is_active` flag.

    The endpoint requires only the `supplier_id` path parameter. It will
    fetch the associated `users.is_active` value and flip it (true -> false,
    false -> true), commit the change, and return the new state.
    """
    s = (
        db.query(models.user.Supplier)
        .filter(models.user.Supplier.id == supplier_id)
        .first()
    )
    if not s:
        raise HTTPException(status_code=404, detail="Supplier not found")

    user = db.query(models.user.User).filter(models.user.User.id == s.user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="Associated user not found")

    # Flip the active flag
    user.is_active = not bool(user.is_active)
    db.add(user)
    db.commit()

    message = (
        "Supplier account has been enabled."
        if user.is_active
        else "Supplier account has been disabled by an administrator."
    )

    return {"user_id": user.id, "is_active": user.is_active, "message": message}


@router.delete(
    "/account",
    summary="Delete Own Account (Viewer/Editor)",
)
def delete_user_account(
    admin_user=Depends(require_viewer_or_editor),
    db: Session = Depends(get_db),
):
    """Delete the authenticated user's account. Only accessible by admin users with 'viewer' or 'editor' role.

    This endpoint will:
    - Delete the authenticated user's account from the database
    - Automatically determined from the authentication token
    - Cascade delete all related data (jobs, subscriptions, etc.)
    - Prevent deletion of actual admin users

    Returns:
        Success message with deleted user information

    Raises:
        HTTPException 401: If unable to identify user
        HTTPException 404: If user not found
        HTTPException 500: If deletion fails
    """
    # Get the authenticated admin user's email
    admin_email = getattr(admin_user, "email", None)
    if not admin_email:
        raise HTTPException(
            status_code=401, detail="Unable to identify authenticated user"
        )

    # Find the user by email
    user = (
        db.query(models.user.User).filter(models.user.User.email == admin_email).first()
    )
    if not user:
        raise HTTPException(status_code=404, detail=f"User account not found")

    # Check if the user is an actual admin user - admin accounts cannot be deleted
    admin_check = db.execute(
        text("SELECT id FROM admin_users WHERE email = :email"), {"email": user.email}
    ).first()

    if admin_check:
        return {
            "success": False,
            "deleted": False,
            "message": "Admin accounts cannot be deleted through this endpoint.",
            "detail": "Please use the appropriate admin management system or contact your system administrator for assistance.",
            "user_email": user.email,
        }

    try:
        # Store email and ID for response message
        user_email = user.email
        user_id = user.id

        # Delete the user (cascade will handle related records)
        db.delete(user)
        db.commit()

        logger.info(f"User account deleted: {user_email} (ID: {user_id})")

        return {
            "success": True,
            "user_id": user_id,
            "email": user_email,
            "deleted": True,
            "message": f"Your account '{user_email}' has been successfully deleted.",
        }

    except Exception as e:
        db.rollback()
        logger.error(f"Failed to delete user {user_id}: {str(e)}")
        raise HTTPException(
            status_code=500, detail=f"Failed to delete user account: {str(e)}"
        )


@router.put(
    "/users/approve",
    dependencies=[Depends(require_admin_only)],
    summary="Approve or Reject User",
)
def update_user_approval(
    data: UserApprovalUpdate,
    admin: models.user.AdminUser = Depends(require_admin_only),
    db: Session = Depends(get_db),
):
    """
    Admin endpoint to approve or reject a user account.

    Status options:
    - "approved": User can access the platform
    - "rejected": User is denied access

    Restrictions:
    - Only users with 'admin' role can use this endpoint
    - Cannot approve/reject admin accounts
    """
    # Validate status
    if data.status not in ["approved", "rejected"]:
        raise HTTPException(
            status_code=400, detail="Status must be either 'approved' or 'rejected'"
        )

    # Get the user
    user = (
        db.query(models.user.User).filter(models.user.User.id == data.user_id).first()
    )

    if not user:
        raise HTTPException(status_code=404, detail="User not found")

    # Check if target user is an admin (prevent approving/rejecting admin accounts)
    try:
        admin_check = db.execute(
            text(
                "SELECT id FROM admin_users WHERE lower(email) = lower(:email) LIMIT 1"
            ),
            {"email": user.email},
        ).first()

        if admin_check:
            raise HTTPException(
                status_code=403,
                detail="Cannot approve or reject admin accounts through this endpoint.",
            )
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error checking admin status: {str(e)}")

    try:
        # Update approval status
        old_status = user.approved_by_admin
        user.approved_by_admin = data.status
        db.commit()
        db.refresh(user)

        logger.info(
            f"Admin {admin.email} changed user {user.email} (ID: {user.id}) approval status from "
            f"'{old_status}' to '{data.status}'"
        )

        # Create notification for user
        notification = models.user.Notification(
            user_id=user.id,
            type="account_approval",
            message=f"Your account has been {data.status} by an administrator.",
        )
        db.add(notification)
        db.commit()

        return {
            "success": True,
            "user_id": user.id,
            "email": user.email,
            "status": data.status,
            "message": f"User account has been {data.status} successfully.",
            "updated_by": admin.email,
        }

    except Exception as e:
        db.rollback()
        logger.error(
            f"Failed to update approval status for user {data.user_id}: {str(e)}"
        )
        raise HTTPException(
            status_code=500, detail=f"Failed to update approval status: {str(e)}"
        )


import json
from typing import Optional

# ============================================================================
# Helper Functions for Analytics
# ============================================================================


def _get_date_range_from_filter(
    time_range: str, date_from: Optional[str] = None, date_to: Optional[str] = None
):
    """
    Convert time_range filter to (start_date, end_date, periods, bucket).

    Returns:
        tuple: (start_date, end_date, periods_list, bucket_type)
        - periods_list: [(label, start, end), ...]
        - bucket_type: 'day', 'week', or 'month'
    """
    now = datetime.utcnow()

    if time_range == "custom" and date_from and date_to:
        start = datetime.fromisoformat(date_from.replace("Z", ""))
        end = datetime.fromisoformat(date_to.replace("Z", ""))
        # For custom range, use monthly buckets
        periods, bucket = _periods_for_range("last6months")
        # Filter periods to custom range
        periods = [(label, s, e) for label, s, e in periods if s >= start and e <= end]
        return start, end, periods, bucket

    # Map time_range to periods
    if time_range == "today":
        start = now.replace(hour=0, minute=0, second=0, microsecond=0)
        end = now
        periods = [
            (
                start.strftime("%H:00"),
                start + timedelta(hours=i),
                start + timedelta(hours=i + 1),
            )
            for i in range(24)
        ]
        return start, end, periods, "hour"

    elif time_range == "7days":
        start = (now - timedelta(days=6)).replace(
            hour=0, minute=0, second=0, microsecond=0
        )
        end = now
        periods = []
        for i in range(7):
            d = start + timedelta(days=i)
            periods.append((d.strftime("%a"), d, d + timedelta(days=1)))
        return start, end, periods, "day"

    elif time_range == "30days":
        periods, bucket = _periods_for_range("last30days")
        start = periods[0][1]
        end = periods[-1][2]
        return start, end, periods, bucket

    elif time_range == "1year":
        periods, bucket = _periods_for_range("last12months")
        start = periods[0][1]
        end = periods[-1][2]
        return start, end, periods, bucket

    else:  # Default: 6months
        periods, bucket = _periods_for_range("last6months")
        start = periods[0][1]
        end = periods[-1][2]
        return start, end, periods, bucket


def _apply_global_filters(query, model, filters: dict, db: Session):
    """
    Apply global filters (state, user_type, subscription_tier) to a query.

    Args:
        query: SQLAlchemy query object
        model: The main model being queried (Job, User, etc.)
        filters: Dict with keys: state, user_type, subscription_tier
        db: Database session

    Returns:
        Modified query with filters applied
    """
    # State filter
    if filters.get("state") and filters["state"] not in ["All", "all"]:
        state = filters["state"]
        if hasattr(model, "state"):
            query = query.filter(model.state == state)

    # User type filter (Contractors vs Suppliers)
    if filters.get("user_type") and filters["user_type"] not in ["All", "all"]:
        user_type = filters["user_type"]
        if user_type == "Contractors":
            # Join with contractors table
            query = query.join(
                models.user.Contractor,
                models.user.Contractor.user_id == models.user.User.id,
            )
        elif user_type == "Suppliers":
            # Join with suppliers table
            query = query.join(
                models.user.Supplier,
                models.user.Supplier.user_id == models.user.User.id,
            )

    # Subscription tier filter
    if filters.get("subscription_tier") and filters["subscription_tier"] not in [
        "All",
        "all",
    ]:
        tier = filters["subscription_tier"]
        # Join with subscribers and subscriptions
        query = query.join(
            models.user.Subscriber,
            models.user.Subscriber.user_id == models.user.User.id,
        )
        query = query.join(
            models.user.Subscription,
            models.user.Subscription.id == models.user.Subscriber.subscription_id,
        )
        query = query.filter(models.user.Subscription.name == tier)

    return query


def _calculate_credits_flow(db: Session, period_start, period_end, filters: dict):
    """
    Calculate credits flow for a period: granted, purchased, spent, frozen.

    Returns:
        dict: {"granted": int, "purchased": int, "spent": int, "frozen": int}
    """
    # 1. Credits Granted (trial credits given to new users in this period)
    granted_query = (
        db.query(func.coalesce(func.sum(models.user.Subscriber.trial_credits), 0))
        .join(models.user.User, models.user.Subscriber.user_id == models.user.User.id)
        .join(
            models.user.Subscription,
            models.user.Subscriber.subscription_id == models.user.Subscription.id,
        )
        .filter(
            models.user.Subscriber.subscription_start_date >= period_start,
            models.user.Subscriber.subscription_start_date < period_end,
            models.user.Subscriber.trial_credits_used == True,
        )
    )
    granted_query = _apply_user_filters(
        db, granted_query, filters, subscriber_joined=True
    )
    granted = granted_query.scalar() or 0

    # 2. Credits Purchased (from subscription purchases in this period)
    purchased_query = (
        db.query(func.coalesce(func.sum(models.user.Subscription.credits), 0))
        .join(
            models.user.Subscriber,
            models.user.Subscriber.subscription_id == models.user.Subscription.id,
        )
        .join(models.user.User, models.user.Subscriber.user_id == models.user.User.id)
        .filter(
            models.user.Subscriber.subscription_start_date >= period_start,
            models.user.Subscriber.subscription_start_date < period_end,
        )
    )
    purchased_query = _apply_user_filters(
        db, purchased_query, filters, subscriber_joined=True
    )
    purchased = purchased_query.scalar() or 0

    # 3. Credits Spent (from unlocked leads in this period)
    spent_query = (
        db.query(func.coalesce(func.sum(models.user.UnlockedLead.credits_spent), 0))
        .join(models.user.User, models.user.UnlockedLead.user_id == models.user.User.id)
        .filter(
            models.user.UnlockedLead.unlocked_at >= period_start,
            models.user.UnlockedLead.unlocked_at < period_end,
        )
    )
    spent_query = _apply_user_filters(db, spent_query, filters)
    spent = spent_query.scalar() or 0

    # 4. Credits Frozen (subscriptions that lapsed in this period)
    frozen_query = (
        db.query(func.coalesce(func.sum(models.user.Subscriber.frozen_credits), 0))
        .join(models.user.User, models.user.Subscriber.user_id == models.user.User.id)
        .join(
            models.user.Subscription,
            models.user.Subscriber.subscription_id == models.user.Subscription.id,
        )
        .filter(
            models.user.Subscriber.frozen_at >= period_start,
            models.user.Subscriber.frozen_at < period_end,
        )
    )
    frozen_query = _apply_user_filters(
        db, frozen_query, filters, subscriber_joined=True
    )
    frozen = frozen_query.scalar() or 0

    return {
        "granted": int(granted),
        "purchased": int(purchased),
        "spent": int(spent),
        "frozen": int(frozen),
    }


def _apply_user_filters(
    db: Session, query, filters: dict, base_model=None, subscriber_joined=False
):
    """
    Apply state, country_city, user_type, and subscription_tier filters to a user query.

    Args:
        db: Database session
        query: SQLAlchemy query object
        filters: dict with keys: state, country_city, user_type, subscription_tier
        base_model: The base model being queried (User, Contractor, or Supplier)
        subscriber_joined: Set to True if Subscriber table is already joined in the query

    Returns:
        Filtered query
    """
    state = filters.get("state")
    country_city = filters.get("country_city")
    user_type = filters.get("user_type")
    subscription_tier = filters.get("subscription_tier")

    # Apply user_type filter first (determines which table to join)
    if user_type and user_type != "All":
        if user_type == "Contractors":
            # Join Contractor table if not already joined
            if base_model != models.user.Contractor:
                query = query.join(
                    models.user.Contractor,
                    models.user.User.id == models.user.Contractor.user_id,
                )

            # Apply state filter for contractors
            if state and state != "All":
                query = query.filter(models.user.Contractor.state.any(state))

            # Apply country_city filter for contractors
            if country_city and country_city != "All":
                query = query.filter(
                    models.user.Contractor.country_city.any(country_city)
                )

        elif user_type == "Suppliers":
            # Join Supplier table if not already joined
            if base_model != models.user.Supplier:
                query = query.join(
                    models.user.Supplier,
                    models.user.User.id == models.user.Supplier.user_id,
                )

            # Apply state filter for suppliers
            if state and state != "All":
                query = query.filter(models.user.Supplier.service_states.any(state))

            # Apply country_city filter for suppliers
            if country_city and country_city != "All":
                query = query.filter(
                    models.user.Supplier.country_city.any(country_city)
                )
    else:
        # No user_type filter - need to check both contractors and suppliers
        if state and state != "All" or (country_city and country_city != "All"):
            # Use OR condition to match either contractors or suppliers in the location
            location_conditions = []

            if base_model != models.user.Contractor:
                contractor_exists = db.query(models.user.Contractor).filter(
                    models.user.Contractor.user_id == models.user.User.id
                )
                if state and state != "All":
                    contractor_exists = contractor_exists.filter(
                        models.user.Contractor.state.any(state)
                    )
                if country_city and country_city != "All":
                    contractor_exists = contractor_exists.filter(
                        models.user.Contractor.country_city.any(country_city)
                    )
                location_conditions.append(contractor_exists.exists())

            if base_model != models.user.Supplier:
                supplier_exists = db.query(models.user.Supplier).filter(
                    models.user.Supplier.user_id == models.user.User.id
                )
                if state and state != "All":
                    supplier_exists = supplier_exists.filter(
                        models.user.Supplier.service_states.any(state)
                    )
                if country_city and country_city != "All":
                    supplier_exists = supplier_exists.filter(
                        models.user.Supplier.country_city.any(country_city)
                    )
                location_conditions.append(supplier_exists.exists())

            if location_conditions:
                query = query.filter(or_(*location_conditions))

    # Apply subscription_tier filter
    if subscription_tier and subscription_tier != "All":
        # Only join Subscriber/Subscription if not already joined
        if not subscriber_joined:
            query = (
                query.join(
                    models.user.Subscriber,
                    models.user.User.id == models.user.Subscriber.user_id,
                )
                .join(
                    models.user.Subscription,
                    models.user.Subscriber.subscription_id
                    == models.user.Subscription.id,
                )
                .filter(models.user.Subscription.name == subscription_tier)
            )
        else:
            # Subscriber/Subscription already joined, just add filter
            query = query.filter(models.user.Subscription.name == subscription_tier)

    return query


def _apply_job_filters(query, filters: dict):
    """
    Apply state and country_city filters to job queries.

    Args:
        query: SQLAlchemy query object
        filters: dict with keys: state, country_city

    Returns:
        Filtered query
    """
    state = filters.get("state")
    country_city = filters.get("country_city")

    # Apply state filter
    if state and state != "All":
        query = query.filter(models.user.Job.state == state)

    # Apply country_city filter (Job.source_county)
    if country_city and country_city != "All":
        query = query.filter(models.user.Job.source_county == country_city)

    return query


# ============================================================================
# Main Analytics Endpoint
# ============================================================================


@router.get("/analytics", dependencies=[Depends(require_admin_token)])
def get_admin_analytics(
    time_range: str = Query(
        "6months",
        description="Time range: today, 7days, 30days, 6months, 1year, custom",
    ),
    state: str = Query("All", description="State filter or 'All'"),
    country_city: Optional[str] = Query(
        None, description="County/City filter (optional)"
    ),
    user_type: str = Query("All", description="User type: All, Contractors, Suppliers"),
    subscription_tier: str = Query(
        "All",
        description="Subscription tier: All, Starter, Professional, Enterprise, Custom",
    ),
    date_from: Optional[str] = Query(
        None, description="Custom range start (ISO format)"
    ),
    date_to: Optional[str] = Query(None, description="Custom range end (ISO format)"),
    page: int = Query(1, ge=1, description="Page number for tables"),
    per_page: int = Query(10, ge=1, le=100, description="Items per page"),
    db: Session = Depends(get_db),
    response: Response = None,  # Add Response parameter for headers
):
    """
    Comprehensive admin analytics dashboard.

    Returns:
        - 4 KPIs with growth metrics
        - 8 Charts (revenue, jobs, users, credits flow, funnel, subscriptions, categories, geography)
        - 2 Data tables (top categories, top jurisdictions) with pagination
        - Applied filters metadata
    """

    # Get date range and periods
    start_date, end_date, periods, bucket = _get_date_range_from_filter(
        time_range, date_from, date_to
    )

    # Filters dict for reuse
    filters = {
        "state": state,
        "country_city": country_city,
        "user_type": user_type,
        "subscription_tier": subscription_tier,
    }

    # ========================================================================
    # KPIs Calculation
    # ========================================================================

    # Total Users (cumulative at end of period)
    total_users_query = db.query(func.count(models.user.User.id)).filter(
        models.user.User.created_at < end_date
    )
    total_users_query = _apply_user_filters(db, total_users_query, filters)
    total_users = total_users_query.scalar() or 0

    # Previous period users (for growth %)
    period_length = end_date - start_date
    prev_period_end = start_date
    prev_period_start = start_date - period_length
    prev_users_query = db.query(func.count(models.user.User.id)).filter(
        models.user.User.created_at < prev_period_end
    )
    prev_users_query = _apply_user_filters(db, prev_users_query, filters)
    prev_users = prev_users_query.scalar() or 0

    users_change = total_users - prev_users
    users_growth_pct = (
        ((users_change / prev_users * 100) if prev_users > 0 else 100.0)
        if users_change != 0
        else 0.0
    )

    # Total Revenue (from payments table if exists, else from subscriber.total_spending)
    if _table_exists(db, "payments"):
        revenue_query = text(
            "SELECT COALESCE(SUM(amount), 0) FROM payments WHERE payment_date >= :start AND payment_date < :end"
        )
        total_revenue = (
            db.execute(revenue_query, {"start": start_date, "end": end_date}).scalar()
            or 0
        )
        prev_revenue = (
            db.execute(
                revenue_query, {"start": prev_period_start, "end": prev_period_end}
            ).scalar()
            or 0
        )
    else:
        # Fallback: use total_spending from subscribers
        total_revenue = (
            db.query(
                func.coalesce(func.sum(models.user.Subscriber.total_spending), 0)
            ).scalar()
            or 0
        )
        prev_revenue = 0  # Can't calculate previous without timestamps

    revenue_change = total_revenue - prev_revenue
    revenue_growth_pct = (
        ((revenue_change / prev_revenue * 100) if prev_revenue > 0 else 100.0)
        if revenue_change != 0
        else 0.0
    )

    # Active Subscriptions
    active_subs_query = (
        db.query(func.count(models.user.Subscriber.id))
        .join(models.user.User, models.user.Subscriber.user_id == models.user.User.id)
        .join(
            models.user.Subscription,
            models.user.Subscriber.subscription_id == models.user.Subscription.id,
        )
        .filter(models.user.Subscriber.is_active == True)
    )
    active_subs_query = _apply_user_filters(
        db, active_subs_query, filters, subscriber_joined=True
    )
    active_subs = active_subs_query.scalar() or 0

    # Previous active subs (approximate - count those active before period start)
    prev_active_subs_query = (
        db.query(func.count(models.user.Subscriber.id))
        .join(models.user.User, models.user.Subscriber.user_id == models.user.User.id)
        .join(
            models.user.Subscription,
            models.user.Subscriber.subscription_id == models.user.Subscription.id,
        )
        .filter(
            models.user.Subscriber.is_active == True,
            models.user.Subscriber.subscription_start_date < prev_period_end,
        )
    )
    prev_active_subs_query = _apply_user_filters(
        db, prev_active_subs_query, filters, subscriber_joined=True
    )
    prev_active_subs = prev_active_subs_query.scalar() or 0

    subs_change = active_subs - prev_active_subs
    subs_growth_pct = (
        ((subs_change / prev_active_subs * 100) if prev_active_subs > 0 else 100.0)
        if subs_change != 0
        else 0.0
    )

    # ========================================================================
    # Charts Data
    # ========================================================================

    # Chart 1: Revenue Timeline
    revenue_data = []
    for label, p_start, p_end in periods:
        if _table_exists(db, "payments"):
            rev_q = text(
                "SELECT COALESCE(SUM(amount), 0) FROM payments WHERE payment_date >= :s AND payment_date < :e"
            )
            rev = db.execute(rev_q, {"s": p_start, "e": p_end}).scalar() or 0
        else:
            rev = 0
        revenue_data.append({"month": label, "value": int(rev)})

    revenue_total = sum(r["value"] for r in revenue_data)
    revenue_peak = (
        max(revenue_data, key=lambda x: x["value"])
        if revenue_data
        else {"month": None, "value": 0}
    )

    # Chart 2: Jobs Growth
    jobs_data = []
    for label, p_start, p_end in periods:
        jobs_query = db.query(func.count(models.user.Job.id)).filter(
            models.user.Job.created_at >= p_start, models.user.Job.created_at < p_end
        )
        jobs_query = _apply_job_filters(jobs_query, filters)
        jobs_count = jobs_query.scalar() or 0
        jobs_data.append({"month": label, "value": int(jobs_count)})

    jobs_total = sum(j["value"] for j in jobs_data)

    # Chart 3: User Growth (Cumulative)
    users_growth_data = []
    for label, p_start, p_end in periods:
        users_cum_query = db.query(func.count(models.user.User.id)).filter(
            models.user.User.created_at < p_end
        )
        users_cum_query = _apply_user_filters(db, users_cum_query, filters)
        users_cum = users_cum_query.scalar() or 0
        users_growth_data.append({"month": label, "value": int(users_cum)})

    # Chart 4: Credits Flow (with daily/weekly/monthly toggle)
    credits_flow_data = []

    # Use monthly periods for credits flow
    flow_periods = periods

    for label, p_start, p_end in flow_periods:
        flow = _calculate_credits_flow(db, p_start, p_end, filters)
        credits_flow_data.append(
            {
                "period": label,
                "granted": flow["granted"],
                "purchased": flow["purchased"],
                "spent": flow["spent"],
                "frozen": flow["frozen"],
            }
        )

    credits_totals = {
        "granted": sum(c["granted"] for c in credits_flow_data),
        "purchased": sum(c["purchased"] for c in credits_flow_data),
        "spent": sum(c["spent"] for c in credits_flow_data),
        "frozen": sum(c["frozen"] for c in credits_flow_data),
    }

    # Chart 5: Marketplace Funnel
    # Stage 1: Delivered (all posted jobs)
    delivered_query = db.query(func.count(models.user.Job.id)).filter(
        models.user.Job.job_review_status == "posted"
    )
    delivered_query = _apply_job_filters(delivered_query, filters)
    delivered_count = delivered_query.scalar() or 0

    # Stage 2: Unlocked (all unlocked leads - includes deleted jobs!)
    unlocked_query = db.query(func.count(models.user.UnlockedLead.id)).join(
        models.user.User, models.user.UnlockedLead.user_id == models.user.User.id
    )
    unlocked_query = _apply_user_filters(db, unlocked_query, filters)
    unlocked_count = unlocked_query.scalar() or 0

    unlocked_credits_query = db.query(
        func.coalesce(func.sum(models.user.UnlockedLead.credits_spent), 0)
    ).join(models.user.User, models.user.UnlockedLead.user_id == models.user.User.id)
    unlocked_credits_query = _apply_user_filters(db, unlocked_credits_query, filters)
    unlocked_credits = unlocked_credits_query.scalar() or 0

    # Stage 3: Saved
    saved_query = db.query(func.count(models.user.SavedJob.id)).join(
        models.user.User, models.user.SavedJob.user_id == models.user.User.id
    )
    saved_query = _apply_user_filters(db, saved_query, filters)
    saved_count = saved_query.scalar() or 0

    # Stage 4: Not Interested
    not_interested_query = db.query(func.count(models.user.NotInterestedJob.id)).join(
        models.user.User, models.user.NotInterestedJob.user_id == models.user.User.id
    )
    not_interested_query = _apply_user_filters(db, not_interested_query, filters)
    not_interested_count = not_interested_query.scalar() or 0

    conversion_rate = (
        (unlocked_count / delivered_count * 100) if delivered_count > 0 else 0.0
    )

    # Chart 6: Subscription Distribution (Donut)
    subscription_dist_query = (
        db.query(
            models.user.Subscription.name,
            func.count(models.user.Subscriber.id).label("count"),
        )
        .join(
            models.user.Subscriber,
            models.user.Subscriber.subscription_id == models.user.Subscription.id,
        )
        .join(models.user.User, models.user.Subscriber.user_id == models.user.User.id)
        .filter(models.user.Subscriber.is_active == True)
    )
    subscription_dist_query = _apply_user_filters(
        db, subscription_dist_query, filters, subscriber_joined=True
    )
    subscription_dist = subscription_dist_query.group_by(
        models.user.Subscription.name
    ).all()

    subscription_data = []
    for sub in subscription_dist:
        # Calculate revenue (count * price)
        sub_obj = (
            db.query(models.user.Subscription)
            .filter(models.user.Subscription.name == sub.name)
            .first()
        )
        price = (
            float(sub_obj.price.replace("$", "").replace(",", ""))
            if sub_obj and sub_obj.price
            else 0
        )
        revenue = sub.count * price
        subscription_data.append(
            {"tier": sub.name, "count": sub.count, "revenue": int(revenue)}
        )

    # Chart 7: Category Performance (by user types)
    # Group by audience_type_names
    category_query = (
        db.query(
            models.user.Job.audience_type_names.label("category"),
            func.count(models.user.Job.id).label("delivered"),
            func.count(models.user.UnlockedLead.id).label("unlocked"),
        )
        .outerjoin(
            models.user.UnlockedLead,
            models.user.Job.id == models.user.UnlockedLead.job_id,
        )
        .filter(models.user.Job.audience_type_names.isnot(None))
    )
    category_query = _apply_job_filters(category_query, filters)
    category_query = category_query.group_by(models.user.Job.audience_type_names).all()

    category_data = []
    for cat in category_query:
        conv_pct = (cat.unlocked / cat.delivered * 100) if cat.delivered > 0 else 0.0
        category_data.append(
            {
                "category": cat.category or "Unknown",
                "delivered": cat.delivered,
                "unlocked": cat.unlocked,
                "conversionPct": round(conv_pct, 1),
            }
        )

    # Sort by unlocked count (descending)
    category_data.sort(key=lambda x: x["unlocked"], reverse=True)

    # Chart 8: Geographic Distribution
    geo_jobs_query = db.query(
        models.user.Job.state.label("state"),
        func.count(models.user.Job.id).label("jobs"),
    ).filter(models.user.Job.state.isnot(None))
    geo_jobs_query = _apply_job_filters(geo_jobs_query, filters)
    geo_query = geo_jobs_query.group_by(models.user.Job.state).all()

    geographic_data = []
    for geo in geo_query:
        # Count contractors and suppliers in this state (with filters applied)
        contractors_query = (
            db.query(func.count(func.distinct(models.user.Contractor.user_id)))
            .join(
                models.user.User, models.user.Contractor.user_id == models.user.User.id
            )
            .filter(models.user.Contractor.state.any(geo.state))
        )
        contractors_query = _apply_user_filters(
            db, contractors_query, filters, base_model=models.user.Contractor
        )
        contractors = contractors_query.scalar() or 0

        suppliers_query = (
            db.query(func.count(func.distinct(models.user.Supplier.user_id)))
            .join(models.user.User, models.user.Supplier.user_id == models.user.User.id)
            .filter(models.user.Supplier.service_states.any(geo.state))
        )
        suppliers_query = _apply_user_filters(
            db, suppliers_query, filters, base_model=models.user.Supplier
        )
        suppliers = suppliers_query.scalar() or 0

        geographic_data.append(
            {
                "state": geo.state,
                "jobs": geo.jobs,
                "contractors": contractors,
                "suppliers": suppliers,
            }
        )

    # Sort by jobs count (descending)
    geographic_data.sort(key=lambda x: x["jobs"], reverse=True)

    # ========================================================================
    # Data Tables
    # ========================================================================

    # Table 1: Top Categories Performance
    categories_table_query = (
        db.query(
            models.user.Job.audience_type_names.label("category"),
            func.count(models.user.Job.id).label("delivered"),
            func.count(models.user.UnlockedLead.id).label("unlocked"),
            func.avg(models.user.UnlockedLead.credits_spent).label("avg_credits"),
            func.sum(models.user.UnlockedLead.credits_spent).label("total_revenue"),
        )
        .outerjoin(
            models.user.UnlockedLead,
            models.user.Job.id == models.user.UnlockedLead.job_id,
        )
        .filter(models.user.Job.audience_type_names.isnot(None))
    )

    # Apply user_type filter to unlocks
    if user_type and user_type != "All":
        # Join User table to filter unlocks by user role
        categories_table_query = categories_table_query.outerjoin(
            models.user.User, models.user.UnlockedLead.user_id == models.user.User.id
        )

        if user_type == "Contractors":
            categories_table_query = categories_table_query.filter(
                or_(
                    models.user.UnlockedLead.id.is_(
                        None
                    ),  # Include jobs with no unlocks
                    models.user.User.role == "Contractor",
                )
            )
        elif user_type == "Suppliers":
            categories_table_query = categories_table_query.filter(
                or_(
                    models.user.UnlockedLead.id.is_(
                        None
                    ),  # Include jobs with no unlocks
                    models.user.User.role == "Supplier",
                )
            )

    categories_table_query = _apply_job_filters(categories_table_query, filters)
    categories_table_query = categories_table_query.group_by(
        models.user.Job.audience_type_names
    )

    # Get total count for pagination
    categories_total = categories_table_query.count()

    # Apply pagination and sorting - order by delivered (highest first) to match export
    categories_table = (
        categories_table_query.order_by(func.count(models.user.Job.id).desc())
        .offset((page - 1) * per_page)
        .limit(per_page)
        .all()
    )

    categories_table_data = []
    for cat in categories_table:
        conv_pct = (cat.unlocked / cat.delivered * 100) if cat.delivered > 0 else 0.0
        categories_table_data.append(
            {
                "category": cat.category or "Unknown",
                "delivered": cat.delivered,
                "unlocked": cat.unlocked,
                "conversionPct": round(conv_pct, 1),
                "avgCredits": round(float(cat.avg_credits or 0), 1),
                "totalRevenue": int(cat.total_revenue or 0),
            }
        )

    # Table 2: Top Jurisdictions
    jurisdictions_query = (
        db.query(
            models.user.Job.state.label("location"),
            func.count(models.user.Job.id).label("jobs_delivered"),
            func.count(models.user.UnlockedLead.id).label("unlocks"),
        )
        .outerjoin(
            models.user.UnlockedLead,
            models.user.Job.id == models.user.UnlockedLead.job_id,
        )
        .filter(models.user.Job.state.isnot(None))
    )

    # Apply user_type filter to unlocks
    if user_type and user_type != "All":
        # Join User table to filter unlocks by user role
        jurisdictions_query = jurisdictions_query.outerjoin(
            models.user.User, models.user.UnlockedLead.user_id == models.user.User.id
        )

        if user_type == "Contractors":
            jurisdictions_query = jurisdictions_query.filter(
                or_(
                    models.user.UnlockedLead.id.is_(
                        None
                    ),  # Include jobs with no unlocks
                    models.user.User.role == "Contractor",
                )
            )
        elif user_type == "Suppliers":
            jurisdictions_query = jurisdictions_query.filter(
                or_(
                    models.user.UnlockedLead.id.is_(
                        None
                    ),  # Include jobs with no unlocks
                    models.user.User.role == "Supplier",
                )
            )

    jurisdictions_query = _apply_job_filters(jurisdictions_query, filters)
    jurisdictions_query = jurisdictions_query.group_by(models.user.Job.state)

    jurisdictions_total = jurisdictions_query.count()

    # Order by jobs delivered (highest first) to match export
    jurisdictions_table = (
        jurisdictions_query.order_by(func.count(models.user.Job.id).desc())
        .offset((page - 1) * per_page)
        .limit(per_page)
        .all()
    )

    jurisdictions_table_data = []
    for jur in jurisdictions_table:
        # Get contractors and suppliers count
        contractors = (
            db.query(func.count(func.distinct(models.user.Contractor.user_id)))
            .filter(models.user.Contractor.state.any(jur.location))
            .scalar()
            or 0
        )

        suppliers = (
            db.query(func.count(func.distinct(models.user.Supplier.user_id)))
            .filter(models.user.Supplier.service_states.any(jur.location))
            .scalar()
            or 0
        )

        conv_pct = (
            (jur.unlocks / jur.jobs_delivered * 100) if jur.jobs_delivered > 0 else 0.0
        )

        jurisdictions_table_data.append(
            {
                "location": jur.location,
                "jobsDelivered": jur.jobs_delivered,
                "contractors": contractors,
                "suppliers": suppliers,
                "unlocks": jur.unlocks,
                "conversionPct": round(conv_pct, 1),
            }
        )

    # ========================================================================
    # Build Response
    # ========================================================================

    response_dict = {
        "kpis": {
            "totalUsers": {
                "count": total_users,
                "growth": f"{users_growth_pct:+.1f}%",
                "changeValue": users_change,
            },
            "totalRevenue": {
                "amount": int(total_revenue),
                "growth": f"{revenue_growth_pct:+.1f}%",
                "changeValue": int(revenue_change),
            },
            "activeSubscriptions": {
                "count": active_subs,
                "growth": f"{subs_growth_pct:+.1f}%",
                "changeValue": subs_change,
            },
            "totalRevenue2": {  # Duplicate for 4th KPI slot
                "amount": int(total_revenue),
                "growth": f"{revenue_growth_pct:+.1f}%",
                "changeValue": int(revenue_change),
            },
        },
        "charts": {
            "revenueTimeline": {
                "data": revenue_data,
                "peakMonth": revenue_peak["month"],
                "total": revenue_total,
            },
            "jobsGrowth": {"data": jobs_data, "total": jobs_total},
            "userGrowth": {
                "data": users_growth_data,
                "growthRate": round(users_growth_pct, 1),
            },
            "creditsFlow": {"data": credits_flow_data, "totals": credits_totals},
            "marketplaceFunnel": {
                "delivered": {"count": delivered_count, "credits": 0},
                "unlocked": {"count": unlocked_count, "credits": int(unlocked_credits)},
                "saved": {"count": saved_count, "credits": 0},
                "notInterested": {"count": not_interested_count, "credits": 0},
                "conversionRate": round(conversion_rate, 1),
            },
            "subscriptionDistribution": {"data": subscription_data},
            "categoryPerformance": {"data": category_data[:10]},  # Top 10 for chart
            "geographicDistribution": {"data": geographic_data[:15]},  # Top 15 states
        },
        "tables": {
            "topCategories": {
                "data": categories_table_data,
                "pagination": {
                    "total": categories_total,
                    "page": page,
                    "perPage": per_page,
                    "totalPages": (categories_total + per_page - 1) // per_page,
                },
            },
            "topJurisdictions": {
                "data": jurisdictions_table_data,
                "pagination": {
                    "total": jurisdictions_total,
                    "page": page,
                    "perPage": per_page,
                    "totalPages": (jurisdictions_total + per_page - 1) // per_page,
                },
            },
        },
        "filters": {
            "applied": {
                "timeRange": time_range,
                "state": state,
                "userType": user_type,
                "subscriptionTier": subscription_tier,
            },
            "dateRange": {
                "from": start_date.isoformat() + "Z",
                "to": end_date.isoformat() + "Z",
            },
        },
        "metadata": {
            "lastUpdated": datetime.utcnow().isoformat() + "Z",
            "timezone": "UTC",
            "cacheKey": f"analytics_{time_range}_{state}_{user_type}_{subscription_tier}",
            "cacheDuration": 300,  # 5 minutes
        },
    }

    # Set HTTP cache headers for proper browser caching
    cache_key = f"analytics_{time_range}_{state}_{user_type}_{subscription_tier}"
    response.headers["Cache-Control"] = "public, max-age=300"  # Cache for 5 minutes
    response.headers["ETag"] = f'"{cache_key}"'  # Enable cache validation

    return response_dict


# ============================================================================
# Dedicated Chart Endpoints (with Toggle Parameters)
# ============================================================================


@router.get("/charts/credits-flow", dependencies=[Depends(require_admin_token)])
def get_credits_flow_chart(
    view: str = Query(..., description="View type: daily, weekly, monthly"),
    time_range: str = Query("6months", description="Time range filter"),
    state: str = Query("All", description="State filter"),
    country_city: str = Query("All", description="County/City filter"),
    user_type: str = Query("All", description="User type filter"),
    subscription_tier: str = Query("All", description="Subscription tier filter"),
    date_from: Optional[str] = Query(
        None, description="Custom range start (ISO format)"
    ),
    date_to: Optional[str] = Query(None, description="Custom range end (ISO format)"),
    db: Session = Depends(get_db),
):
    """
    Get credits flow chart data with configurable view.

    View options:
    - daily: Last 30 days, daily buckets
    - weekly: Last 12 weeks, weekly buckets
    - monthly: Last 6 months, monthly buckets

    Returns:
        - data: Array of period data with granted, purchased, spent, frozen, net credits
        - totals: Aggregated totals across all periods
        - metadata: View info, filters, and cache settings
    """
    # Validate view parameter
    if view not in ["daily", "weekly", "monthly"]:
        raise HTTPException(
            status_code=400,
            detail="Invalid view. Must be one of: daily, weekly, monthly",
        )

    # Get date range
    start_date, end_date, _, _ = _get_date_range_from_filter(
        time_range, date_from, date_to
    )
    now = datetime.utcnow()

    # Filters dict for reuse
    filters = {
        "state": state,
        "country_city": country_city,
        "user_type": user_type,
        "subscription_tier": subscription_tier,
    }

    # Determine periods based on view
    if view == "daily":
        # Last 30 days, daily buckets
        periods = []
        for i in range(29, -1, -1):
            day = (now - timedelta(days=i)).replace(
                hour=0, minute=0, second=0, microsecond=0
            )
            periods.append((day.strftime("%Y-%m-%d"), day, day + timedelta(days=1)))

    elif view == "weekly":
        # Last 12 weeks, weekly buckets
        periods = []
        for i in range(11, -1, -1):
            week_start = (now - timedelta(weeks=i)).replace(
                hour=0, minute=0, second=0, microsecond=0
            )
            week_end = week_start + timedelta(weeks=1)
            periods.append((f"Week {12-i}", week_start, week_end))

    else:  # monthly
        # Last 6 months, monthly buckets
        periods = _month_starts(6)

    # Calculate credits flow for each period
    flow_data = []
    for label, p_start, p_end in periods:
        flow = _calculate_credits_flow(db, p_start, p_end, filters)
        net = flow["granted"] + flow["purchased"] - flow["spent"] - flow["frozen"]

        flow_data.append(
            {
                "period": label,
                "granted": flow["granted"],
                "purchased": flow["purchased"],
                "spent": flow["spent"],
                "frozen": flow["frozen"],
                "net": net,
            }
        )

    # Calculate totals
    totals = {
        "granted": sum(d["granted"] for d in flow_data),
        "purchased": sum(d["purchased"] for d in flow_data),
        "spent": sum(d["spent"] for d in flow_data),
        "frozen": sum(d["frozen"] for d in flow_data),
        "net": sum(d["net"] for d in flow_data),
    }

    return {
        "data": flow_data,
        "totals": totals,
        "metadata": {
            "view": view,
            "bucketSize": (
                "1 day"
                if view == "daily"
                else ("7 days" if view == "weekly" else "1 month")
            ),
            "periodCount": len(flow_data),
            "filters": {
                "timeRange": time_range,
                "state": state,
                "userType": user_type,
                "subscriptionTier": subscription_tier,
            },
            "lastUpdated": datetime.utcnow().isoformat() + "Z",
            "cacheDuration": 300,
        },
    }


@router.get("/charts/marketplace-funnel", dependencies=[Depends(require_admin_token)])
def get_marketplace_funnel_chart(
    time_range: str = Query("6months", description="Time range filter"),
    state: Optional[str] = Query(None, description="State filter (optional toggle)"),
    country_city: Optional[str] = Query(
        None, description="Country/City filter (optional toggle)"
    ),
    user_type: Optional[str] = Query(
        None, description="User type filter (optional toggle)"
    ),
    date_from: Optional[str] = Query(
        None, description="Custom range start (ISO format)"
    ),
    date_to: Optional[str] = Query(None, description="Custom range end (ISO format)"),
    db: Session = Depends(get_db),
):
    """
    Get marketplace funnel chart data.

    Funnel stages:
    1. Delivered - All jobs visible to users
    2. Unlocked - Jobs purchased/unlocked
    3. Saved - Jobs bookmarked for later
    4. Not Interested - Jobs rejected

    Optional filter toggles:
    - state: Filter by job state or user's service state
    - country_city: Filter by job county/city or user's service county/city
    - user_type: Filter by user type (contractor/supplier category)

    Returns:
        - data: Funnel stage counts
        - conversionRates: Conversion percentages between stages
        - metadata: Filters applied and cache settings
    """
    # Get date range
    start_date, end_date, _, _ = _get_date_range_from_filter(
        time_range, date_from, date_to
    )

    # Build filters
    filters = {"state": state, "country_city": country_city, "user_type": user_type}

    # Helper function to get user IDs matching filters
    def get_filtered_user_ids():
        """Get user IDs that match state, country_city, and user_type filters"""
        user_ids = set()

        # Apply state filter using ORM
        if state and state != "All":
            # Get contractors with this state
            contractor_ids = (
                db.query(models.user.Contractor.user_id)
                .filter(models.user.Contractor.state.any(state))
                .all()
            )
            for row in contractor_ids:
                user_ids.add(row.user_id)

            # Get suppliers with this state
            supplier_ids = (
                db.query(models.user.Supplier.user_id)
                .filter(models.user.Supplier.service_states.any(state))
                .all()
            )
            for row in supplier_ids:
                user_ids.add(row.user_id)

        # Apply country_city filter using ORM
        if country_city and country_city != "All":
            # Get contractors with this city
            contractor_ids = (
                db.query(models.user.Contractor.user_id)
                .filter(models.user.Contractor.country_city.any(country_city))
                .all()
            )
            for row in contractor_ids:
                user_ids.add(row.user_id)

            # Get suppliers with this city
            supplier_ids = (
                db.query(models.user.Supplier.user_id)
                .filter(models.user.Supplier.service_county.any(country_city))
                .all()
            )
            for row in supplier_ids:
                user_ids.add(row.user_id)

        # Apply user_type filter using ORM
        if user_type:
            # Get contractors with this user type
            contractor_ids = (
                db.query(models.user.Contractor.user_id)
                .filter(models.user.Contractor.user_type.any(user_type))
                .all()
            )
            for row in contractor_ids:
                user_ids.add(row.user_id)

            # Get suppliers with this user type
            supplier_ids = (
                db.query(models.user.Supplier.user_id)
                .filter(models.user.Supplier.user_type.any(user_type))
                .all()
            )
            for row in supplier_ids:
                user_ids.add(row.user_id)

        # If no filters applied, return None to indicate "all users"
        if not user_ids and not state and not country_city and not user_type:
            return None

        return list(user_ids) if user_ids else []

    # Get filtered user IDs
    filtered_user_ids = get_filtered_user_ids()

    # Base response structure
    response = {
        "metadata": {
            "filters": {
                "timeRange": time_range,
                "state": state or "All",
                "countryCity": country_city or "All",
                "userType": user_type or "All",
            },
            "lastUpdated": datetime.utcnow().isoformat() + "Z",
            "cacheDuration": 300,
        }
    }

    # ========================================================================
    # Funnel Stage Counts
    # ========================================================================

    # Stage 1: Delivered (all posted jobs)
    delivered_query = db.query(func.count(models.user.Job.id)).filter(
        models.user.Job.job_review_status == "posted",
        models.user.Job.created_at >= start_date,
        models.user.Job.created_at < end_date,
    )

    # Apply state filter to jobs
    if state and state != "All":
        delivered_query = delivered_query.filter(models.user.Job.state == state)

    # Apply country_city filter to jobs
    if country_city and country_city != "All":
        delivered_query = delivered_query.filter(
            models.user.Job.country_city == country_city
        )

    delivered_count = delivered_query.scalar() or 0

    # Stage 2: Unlocked (all unlocked leads)
    unlocked_query = db.query(func.count(models.user.UnlockedLead.id)).filter(
        models.user.UnlockedLead.unlocked_at >= start_date,
        models.user.UnlockedLead.unlocked_at < end_date,
    )

    # Apply user filter if specified
    if filtered_user_ids is not None:
        if filtered_user_ids:  # Has specific users
            unlocked_query = unlocked_query.filter(
                models.user.UnlockedLead.user_id.in_(filtered_user_ids)
            )
        else:  # Empty list means no matching users
            unlocked_count = 0
            saved_count = 0
            not_interested_count = 0

    if filtered_user_ids != []:
        unlocked_count = unlocked_query.scalar() or 0

        # Stage 3: Saved (bookmarked jobs)
        saved_query = db.query(func.count(models.user.SavedJob.id)).filter(
            models.user.SavedJob.saved_at >= start_date,
            models.user.SavedJob.saved_at < end_date,
        )

        if filtered_user_ids is not None and filtered_user_ids:
            saved_query = saved_query.filter(
                models.user.SavedJob.user_id.in_(filtered_user_ids)
            )

        saved_count = saved_query.scalar() or 0

        # Stage 4: Not Interested (rejected jobs)
        not_interested_query = db.query(
            func.count(models.user.NotInterestedJob.id)
        ).filter(
            models.user.NotInterestedJob.marked_at >= start_date,
            models.user.NotInterestedJob.marked_at < end_date,
        )

        if filtered_user_ids is not None and filtered_user_ids:
            not_interested_query = not_interested_query.filter(
                models.user.NotInterestedJob.user_id.in_(filtered_user_ids)
            )

        not_interested_count = not_interested_query.scalar() or 0

    # Build response
    response["data"] = {
        "delivered": delivered_count,
        "unlocked": unlocked_count,
        "saved": saved_count,
        "notInterested": not_interested_count,
    }

    # Calculate conversion rates
    response["conversionRates"] = {
        "deliveredToUnlocked": round(
            (unlocked_count / delivered_count * 100) if delivered_count > 0 else 0, 1
        ),
        "unlockedToSaved": round(
            (saved_count / unlocked_count * 100) if unlocked_count > 0 else 0, 1
        ),
        "deliveredToNotInterested": round(
            (
                (not_interested_count / delivered_count * 100)
                if delivered_count > 0
                else 0
            ),
            1,
        ),
    }

    return response


# ============================================================================
# CATEGORIES SEARCH ENDPOINT
# ============================================================================


@router.get("/tables/categories/search", dependencies=[Depends(require_admin_token)])
def search_categories(
    search: Optional[str] = Query(None, description="Search by category name"),
    page: int = Query(1, ge=1, description="Page number"),
    per_page: int = Query(10, ge=1, le=100, description="Items per page"),
    state: Optional[str] = Query(None, description="Filter by state"),
    country_city: Optional[str] = Query(None, description="Filter by county/city"),
    user_type: Optional[str] = Query(None, description="Filter by user type"),
    time_range: str = Query("6months", description="Time range filter"),
    date_from: Optional[str] = Query(None, description="Custom start date"),
    date_to: Optional[str] = Query(None, description="Custom end date"),
    db: Session = Depends(get_db),
):
    """
    Search categories table with pagination and filters.

    Returns paginated list of categories with delivered, unlocked, conversion %,
    avg credits, and total revenue.
    """
    from src.app.api.endpoints.admin_dashboard import _get_date_range_from_filter

    # Get date range
    start_date, end_date, _, _ = _get_date_range_from_filter(
        time_range, date_from, date_to
    )

    # Base query for categories
    categories_query = db.query(
        models.user.Job.audience_type_names.label("category"),
        func.count(models.user.Job.id).label("delivered"),
    ).filter(
        models.user.Job.job_review_status == "posted",
        models.user.Job.created_at >= start_date,
        models.user.Job.created_at < end_date,
        models.user.Job.audience_type_names.isnot(None),
    )

    # Apply state filter
    if state and state != "All":
        categories_query = categories_query.filter(models.user.Job.state == state)

    # Apply country_city filter
    if country_city and country_city != "All":
        categories_query = categories_query.filter(
            models.user.Job.country_city == country_city
        )

    # Apply search filter (case-insensitive partial match)
    if search:
        categories_query = categories_query.filter(
            models.user.Job.audience_type_names.ilike(f"%{search}%")
        )

    # Group by category
    categories_query = categories_query.group_by(models.user.Job.audience_type_names)

    # Order by delivered count (highest first) for consistent top 10
    categories_query = categories_query.order_by(func.count(models.user.Job.id).desc())

    # Get total count for pagination
    total_count = categories_query.count()

    # Apply pagination
    offset = (page - 1) * per_page
    categories_data = categories_query.offset(offset).limit(per_page).all()

    # Build response data
    result_data = []
    for cat in categories_data:
        category = cat.category
        delivered = cat.delivered

        # Get unlocked count for this category
        unlocked_query = (
            db.query(func.count(models.user.UnlockedLead.id))
            .join(
                models.user.Job, models.user.Job.id == models.user.UnlockedLead.job_id
            )
            .filter(
                models.user.Job.audience_type_names == category,
                models.user.UnlockedLead.unlocked_at >= start_date,
                models.user.UnlockedLead.unlocked_at < end_date,
            )
        )

        # Apply user_type filter
        if user_type and user_type != "All":
            unlocked_query = unlocked_query.join(
                models.user.User,
                models.user.User.id == models.user.UnlockedLead.user_id,
            ).filter(models.user.User.role == user_type)

        unlocked = unlocked_query.scalar() or 0

        # Get credits data
        credits_query = (
            db.query(
                func.avg(models.user.UnlockedLead.credits_spent).label("avg_credits"),
                func.sum(models.user.UnlockedLead.credits_spent).label("total_credits"),
            )
            .join(
                models.user.Job, models.user.Job.id == models.user.UnlockedLead.job_id
            )
            .filter(
                models.user.Job.audience_type_names == category,
                models.user.UnlockedLead.unlocked_at >= start_date,
                models.user.UnlockedLead.unlocked_at < end_date,
            )
        )

        # Apply user_type filter
        if user_type and user_type != "All":
            credits_query = credits_query.join(
                models.user.User,
                models.user.User.id == models.user.UnlockedLead.user_id,
            ).filter(models.user.User.role == user_type)

        credits_query = credits_query.first()

        avg_credits = int(credits_query.avg_credits or 0)
        total_revenue = int(credits_query.total_credits or 0)

        # Calculate conversion percentage
        conversion_pct = round((unlocked / delivered * 100), 1) if delivered > 0 else 0

        result_data.append(
            {
                "category": category,
                "delivered": delivered,
                "unlocked": unlocked,
                "conversionPct": conversion_pct,
                "avgCredits": avg_credits,
                "totalRevenue": total_revenue,
            }
        )

    return {
        "data": result_data,
        "pagination": {
            "total": total_count,
            "page": page,
            "perPage": per_page,
            "totalPages": (total_count + per_page - 1) // per_page,
        },
        "filters": {
            "search": search or "",
            "state": state or "All",
            "countryCity": country_city or "All",
            "userType": user_type or "All",
            "timeRange": time_range,
        },
    }


# ============================================================================
# CATEGORIES CSV EXPORT ENDPOINT
# ============================================================================


@router.get("/export/categories", dependencies=[Depends(require_admin_token)])
def export_categories_excel(
    state: Optional[str] = Query(None, description="Filter by state"),
    country_city: Optional[str] = Query(None, description="Filter by county/city"),
    user_type: Optional[str] = Query(None, description="Filter by user type"),
    time_range: str = Query("6months", description="Time range filter"),
    date_from: Optional[str] = Query(None, description="Custom start date"),
    date_to: Optional[str] = Query(None, description="Custom end date"),
    db: Session = Depends(get_db),
):
    """
    Export categories data to Excel file with formatting.
    """
    from src.app.api.endpoints.admin_dashboard import _get_date_range_from_filter

    # Get date range
    start_date, end_date, _, _ = _get_date_range_from_filter(
        time_range, date_from, date_to
    )

    # Get all categories (no pagination for export)
    categories_query = db.query(
        models.user.Job.audience_type_names.label("category"),
        func.count(models.user.Job.id).label("delivered"),
    ).filter(
        models.user.Job.job_review_status == "posted",
        models.user.Job.created_at >= start_date,
        models.user.Job.created_at < end_date,
        models.user.Job.audience_type_names.isnot(None),
    )

    # Apply filters
    if state and state != "All":
        categories_query = categories_query.filter(models.user.Job.state == state)
    if country_city and country_city != "All":
        categories_query = categories_query.filter(
            models.user.Job.country_city == country_city
        )

    categories_query = categories_query.group_by(models.user.Job.audience_type_names)
    # Order by delivered count (highest first) - same as search endpoint
    categories_query = categories_query.order_by(func.count(models.user.Job.id).desc())
    categories_data = categories_query.all()

    # Prepare data for DataFrame
    data_rows = []
    for cat in categories_data:
        category = cat.category
        delivered = cat.delivered

        # Get unlocked count
        unlocked = (
            db.query(func.count(models.user.UnlockedLead.id))
            .join(
                models.user.Job, models.user.Job.id == models.user.UnlockedLead.job_id
            )
            .filter(
                models.user.Job.audience_type_names == category,
                models.user.UnlockedLead.unlocked_at >= start_date,
                models.user.UnlockedLead.unlocked_at < end_date,
            )
            .scalar()
            or 0
        )

        # Get credits data
        credits_query = (
            db.query(
                func.avg(models.user.UnlockedLead.credits_spent).label("avg_credits"),
                func.sum(models.user.UnlockedLead.credits_spent).label("total_credits"),
            )
            .join(
                models.user.Job, models.user.Job.id == models.user.UnlockedLead.job_id
            )
            .filter(
                models.user.Job.audience_type_names == category,
                models.user.UnlockedLead.unlocked_at >= start_date,
                models.user.UnlockedLead.unlocked_at < end_date,
            )
            .first()
        )

        avg_credits = int(credits_query.avg_credits or 0)
        total_revenue = int(credits_query.total_credits or 0)
        conversion_pct = round((unlocked / delivered * 100), 1) if delivered > 0 else 0

        data_rows.append(
            {
                "Category": category,
                "Delivered": delivered,
                "Unlocked": unlocked,
                "Conversion %": conversion_pct,
                "Avg. Credits": avg_credits,
                "Total Revenue": total_revenue,
            }
        )

    # Create DataFrame
    df = pd.DataFrame(data_rows)

    # Create Excel file in memory with formatting
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Categories", index=False)

        # Get workbook and worksheet for formatting
        workbook = writer.book
        worksheet = writer.sheets["Categories"]

        # Set column widths
        worksheet.column_dimensions["A"].width = 50  # Category
        worksheet.column_dimensions["B"].width = 12  # Delivered
        worksheet.column_dimensions["C"].width = 12  # Unlocked
        worksheet.column_dimensions["D"].width = 15  # Conversion %
        worksheet.column_dimensions["E"].width = 15  # Avg. Credits
        worksheet.column_dimensions["F"].width = 15  # Total Revenue

        # Format header row
        from openpyxl.styles import Alignment, Font, PatternFill

        header_fill = PatternFill(
            start_color="4472C4", end_color="4472C4", fill_type="solid"
        )
        header_font = Font(bold=True, color="FFFFFF")

        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Center align numeric columns
        for row in worksheet.iter_rows(
            min_row=2, max_row=worksheet.max_row, min_col=2, max_col=6
        ):
            for cell in row:
                cell.alignment = Alignment(horizontal="center")

    # Prepare response
    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename=categories_{time_range}.xlsx"
        },
    )


# ============================================================================
# JURISDICTIONS SEARCH ENDPOINT
# ============================================================================


@router.get("/tables/jurisdictions/search", dependencies=[Depends(require_admin_token)])
def search_jurisdictions(
    search: Optional[str] = Query(None, description="Search by state/county name"),
    page: int = Query(1, ge=1, description="Page number"),
    per_page: int = Query(10, ge=1, le=100, description="Items per page"),
    state: Optional[str] = Query(None, description="Filter by state"),
    country_city: Optional[str] = Query(None, description="Filter by county/city"),
    user_type: Optional[str] = Query(None, description="Filter by user type"),
    time_range: str = Query("6months", description="Time range filter"),
    date_from: Optional[str] = Query(None, description="Custom start date"),
    date_to: Optional[str] = Query(None, description="Custom end date"),
    db: Session = Depends(get_db),
):
    """
    Search jurisdictions table with pagination and filters.

    Returns paginated list of jurisdictions with jobs delivered, contractors,
    suppliers, unlocks, and conversion %.
    """
    from src.app.api.endpoints.admin_dashboard import _get_date_range_from_filter

    # Get date range
    start_date, end_date, _, _ = _get_date_range_from_filter(
        time_range, date_from, date_to
    )

    # Base query for jurisdictions (group by state)
    jurisdictions_query = db.query(
        models.user.Job.state.label("location"),
        func.count(models.user.Job.id).label("jobs_delivered"),
    ).filter(
        models.user.Job.job_review_status == "posted",
        models.user.Job.created_at >= start_date,
        models.user.Job.created_at < end_date,
        models.user.Job.state.isnot(None),
    )

    # Apply state filter
    if state and state != "All":
        jurisdictions_query = jurisdictions_query.filter(models.user.Job.state == state)

    # Apply country_city filter
    if country_city and country_city != "All":
        jurisdictions_query = jurisdictions_query.filter(
            models.user.Job.country_city == country_city
        )

    # Apply search filter (case-insensitive partial match)
    if search:
        jurisdictions_query = jurisdictions_query.filter(
            models.user.Job.state.ilike(f"%{search}%")
        )

    # Group by state
    jurisdictions_query = jurisdictions_query.group_by(models.user.Job.state)

    # Order by jobs delivered (highest first) for consistent top 10
    jurisdictions_query = jurisdictions_query.order_by(
        func.count(models.user.Job.id).desc()
    )

    # Get total count for pagination
    total_count = jurisdictions_query.count()

    # Apply pagination
    offset = (page - 1) * per_page
    jurisdictions_data = jurisdictions_query.offset(offset).limit(per_page).all()

    # Build response data
    result_data = []
    for jur in jurisdictions_data:
        location = jur.location
        jobs_delivered = jur.jobs_delivered

        # Get contractors count for this state
        contractors = (
            db.query(func.count(func.distinct(models.user.Contractor.user_id)))
            .filter(models.user.Contractor.state.any(location))
            .scalar()
            or 0
        )

        # Get suppliers count for this state
        suppliers = (
            db.query(func.count(func.distinct(models.user.Supplier.user_id)))
            .filter(models.user.Supplier.service_states.any(location))
            .scalar()
            or 0
        )

        # Get unlocks count for jobs in this state
        unlocks_query = (
            db.query(func.count(models.user.UnlockedLead.id))
            .join(
                models.user.Job, models.user.Job.id == models.user.UnlockedLead.job_id
            )
            .filter(
                models.user.Job.state == location,
                models.user.UnlockedLead.unlocked_at >= start_date,
                models.user.UnlockedLead.unlocked_at < end_date,
            )
        )

        # Apply user_type filter
        if user_type and user_type != "All":
            unlocks_query = unlocks_query.join(
                models.user.User,
                models.user.User.id == models.user.UnlockedLead.user_id,
            ).filter(models.user.User.role == user_type)

        unlocks = unlocks_query.scalar() or 0

        # Calculate conversion percentage
        conversion_pct = (
            round((unlocks / jobs_delivered * 100), 0) if jobs_delivered > 0 else 0
        )

        result_data.append(
            {
                "location": location,
                "jobsDelivered": jobs_delivered,
                "contractors": contractors,
                "suppliers": suppliers,
                "unlocks": unlocks,
                "conversionPct": int(conversion_pct),
            }
        )

    return {
        "data": result_data,
        "pagination": {
            "total": total_count,
            "page": page,
            "perPage": per_page,
            "totalPages": (total_count + per_page - 1) // per_page,
        },
        "filters": {
            "search": search or "",
            "state": state or "All",
            "countryCity": country_city or "All",
            "userType": user_type or "All",
            "timeRange": time_range,
        },
    }


# ============================================================================
# JURISDICTIONS CSV EXPORT ENDPOINT
# ============================================================================


@router.get("/export/jurisdictions", dependencies=[Depends(require_admin_token)])
def export_jurisdictions_excel(
    state: Optional[str] = Query(None, description="Filter by state"),
    country_city: Optional[str] = Query(None, description="Filter by county/city"),
    user_type: Optional[str] = Query(None, description="Filter by user type"),
    time_range: str = Query("6months", description="Time range filter"),
    date_from: Optional[str] = Query(None, description="Custom start date"),
    date_to: Optional[str] = Query(None, description="Custom end date"),
    db: Session = Depends(get_db),
):
    """
    Export jurisdictions data to Excel file with formatting.
    """
    from src.app.api.endpoints.admin_dashboard import _get_date_range_from_filter

    # Get date range
    start_date, end_date, _, _ = _get_date_range_from_filter(
        time_range, date_from, date_to
    )

    # Get all jurisdictions (no pagination for export)
    jurisdictions_query = db.query(
        models.user.Job.state.label("location"),
        func.count(models.user.Job.id).label("jobs_delivered"),
    ).filter(
        models.user.Job.job_review_status == "posted",
        models.user.Job.created_at >= start_date,
        models.user.Job.created_at < end_date,
        models.user.Job.state.isnot(None),
    )

    # Apply filters
    if state and state != "All":
        jurisdictions_query = jurisdictions_query.filter(models.user.Job.state == state)
    if country_city and country_city != "All":
        jurisdictions_query = jurisdictions_query.filter(
            models.user.Job.country_city == country_city
        )

    jurisdictions_query = jurisdictions_query.group_by(models.user.Job.state)
    # Order by jobs delivered (highest first) - same as search endpoint
    jurisdictions_query = jurisdictions_query.order_by(
        func.count(models.user.Job.id).desc()
    )
    jurisdictions_data = jurisdictions_query.all()

    # Prepare data for DataFrame
    data_rows = []
    for jur in jurisdictions_data:
        location = jur.location
        jobs_delivered = jur.jobs_delivered

        # Get contractors and suppliers count
        contractors = (
            db.query(func.count(func.distinct(models.user.Contractor.user_id)))
            .filter(models.user.Contractor.state.any(location))
            .scalar()
            or 0
        )

        suppliers = (
            db.query(func.count(func.distinct(models.user.Supplier.user_id)))
            .filter(models.user.Supplier.service_states.any(location))
            .scalar()
            or 0
        )

        # Get unlocks count
        unlocks = (
            db.query(func.count(models.user.UnlockedLead.id))
            .join(
                models.user.Job, models.user.Job.id == models.user.UnlockedLead.job_id
            )
            .filter(
                models.user.Job.state == location,
                models.user.UnlockedLead.unlocked_at >= start_date,
                models.user.UnlockedLead.unlocked_at < end_date,
            )
            .scalar()
            or 0
        )

        conversion_pct = (
            round((unlocks / jobs_delivered * 100), 1) if jobs_delivered > 0 else 0
        )

        data_rows.append(
            {
                "State/County": location,
                "Jobs Delivered": jobs_delivered,
                "Contractors": contractors,
                "Suppliers": suppliers,
                "Unlocks": unlocks,
                "Conversion %": conversion_pct,
            }
        )

    # Create DataFrame
    df = pd.DataFrame(data_rows)

    # Create Excel file in memory with formatting
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Jurisdictions", index=False)

        # Get workbook and worksheet for formatting
        workbook = writer.book
        worksheet = writer.sheets["Jurisdictions"]

        # Set column widths
        worksheet.column_dimensions["A"].width = 25  # State/County
        worksheet.column_dimensions["B"].width = 15  # Jobs Delivered
        worksheet.column_dimensions["C"].width = 15  # Contractors
        worksheet.column_dimensions["D"].width = 15  # Suppliers
        worksheet.column_dimensions["E"].width = 12  # Unlocks
        worksheet.column_dimensions["F"].width = 15  # Conversion %

        # Format header row
        from openpyxl.styles import Alignment, Font, PatternFill

        header_fill = PatternFill(
            start_color="4472C4", end_color="4472C4", fill_type="solid"
        )
        header_font = Font(bold=True, color="FFFFFF")

        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Center align numeric columns
        for row in worksheet.iter_rows(
            min_row=2, max_row=worksheet.max_row, min_col=2, max_col=6
        ):
            for cell in row:
                cell.alignment = Alignment(horizontal="center")

    # Prepare response
    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename=jurisdictions_{time_range}.xlsx"
        },
    )


# ============================================================================
# Admin Settings Endpoints
# ============================================================================


@router.get("/settings/auto-post-jobs", dependencies=[Depends(require_admin_token)])
def get_auto_post_jobs_setting(db: Session = Depends(get_db)):
    """
    Get current auto-post jobs setting.

    Returns whether jobs uploaded via upload-leads endpoints should be
    automatically posted based on timing logic, or held as pending for manual review.

    Returns:
        {
            "auto_post_jobs": true,
            "description": "Auto-post jobs from upload endpoints based on timing logic (true/false)",
            "updated_at": "2026-02-20T10:30:00",
            "updated_by": 1
        }
    """
    setting = (
        db.query(models.user.AdminSettings)
        .filter(models.user.AdminSettings.setting_key == "auto_post_jobs")
        .first()
    )

    if not setting:
        # Return default if not found
        return {
            "auto_post_jobs": True,
            "description": "Auto-post jobs from upload endpoints based on timing logic (default)",
            "updated_at": None,
            "updated_by": None,
        }

    return {
        "auto_post_jobs": setting.setting_value.lower() == "true",
        "description": setting.description,
        "updated_at": setting.updated_at.isoformat() if setting.updated_at else None,
        "updated_by": setting.updated_by,
    }


@router.patch("/settings/auto-post-jobs", dependencies=[Depends(require_admin_token)])
def update_auto_post_jobs_setting(
    enabled: bool = Body(..., embed=True, description="Enable or disable auto-posting"),
    admin: models.user.AdminUser = Depends(require_admin_token),
    db: Session = Depends(get_db),
):
    """
    Toggle auto-post jobs setting.

    When enabled (true):
    - Jobs uploaded via /jobs/upload-leads and /jobs/upload-leads-json will be
      automatically posted based on timing logic (anchor_at + day_offset)

    When disabled (false):
    - All uploaded jobs will have status 'pending' regardless of timing
    - Admin must manually post jobs one by one from the admin dashboard

    Request body:
        {
            "enabled": true  // or false
        }

    Returns:
        {
            "success": true,
            "auto_post_jobs": true,
            "message": "Auto-post jobs setting updated successfully",
            "updated_at": "2026-02-20T10:30:00",
            "updated_by": 1
        }
    """
    try:
        # Find or create setting
        setting = (
            db.query(models.user.AdminSettings)
            .filter(models.user.AdminSettings.setting_key == "auto_post_jobs")
            .first()
        )

        if not setting:
            # Create new setting
            setting = models.user.AdminSettings(
                setting_key="auto_post_jobs",
                setting_value="true" if enabled else "false",
                description="Auto-post jobs from upload endpoints based on timing logic (true/false)",
                updated_by=admin.id,
            )
            db.add(setting)
            logger.info(
                f"Created auto_post_jobs setting: {enabled} by admin {admin.id}"
            )
        else:
            # Update existing setting
            old_value = setting.setting_value
            setting.setting_value = "true" if enabled else "false"
            setting.updated_at = datetime.utcnow()
            setting.updated_by = admin.id
            logger.info(
                f"Updated auto_post_jobs setting from {old_value} to {enabled} by admin {admin.id}"
            )

        db.commit()
        db.refresh(setting)

        return {
            "success": True,
            "auto_post_jobs": enabled,
            "message": f"Auto-post jobs setting {'enabled' if enabled else 'disabled'} successfully",
            "updated_at": (
                setting.updated_at.isoformat() if setting.updated_at else None
            ),
            "updated_by": setting.updated_by,
        }

    except Exception as e:
        db.rollback()
        logger.error(f"Failed to update auto_post_jobs setting: {str(e)}")
        raise HTTPException(
            status_code=500, detail=f"Failed to update setting: {str(e)}"
        )
